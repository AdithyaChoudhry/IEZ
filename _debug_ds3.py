"""Inspect IODB and template heading/value structure"""
import sys, io
sys.path.insert(0, '/Users/adithyachoudhrym/ProjectIEZ')
import openpyxl
from openpyxl.utils import get_column_letter

# IODB columns
wb = openpyxl.load_workbook('/Users/adithyachoudhrym/ProjectIEZ/IODB.xlsx')
print("IODB sheets:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print("\nFirst 3 rows of IODB:")
for row in ws.iter_rows(min_row=1, max_row=3):
    cols = [(cell.coordinate, repr(cell.value)) for cell in row if cell.value is not None]
    print(" ", cols)

print("\n--- Template DS NEW WWTP: all rows with Refer Annexure ---")
wb2 = openpyxl.load_workbook('/Users/adithyachoudhrym/ProjectIEZ/Pressure Transmitter.xlsx')
ws2 = wb2['DS NEW WWTP']
for row in ws2.iter_rows(min_row=1, max_row=45):
    ra_cells = [c for c in row if isinstance(c.value, str) and 'annexure' in c.value.lower()]
    if ra_cells:
        non_empty = [(get_column_letter(c.column), repr(c.value))
                     for c in row if c.value is not None and c.__class__.__name__ != 'MergedCell']
        print(f"  Row {row[0].row}: {non_empty}")
