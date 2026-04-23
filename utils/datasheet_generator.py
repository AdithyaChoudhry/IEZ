"""
datasheet_generator.py
======================
Production-quality datasheet population engine.

Spec
----
* IODB: first TWO rows are combined into a single header (multi-row header).
* Template: Sheet 2 is named "Datasheet".
  - Column D contains parameter headings.
  - The cell immediately to the RIGHT of each heading (column E, or the next
    non-empty column) contains the value placeholder.
  - Cells containing "Refer Annexure 1" (case-insensitive) are replaced with
    the matched IODB value.
* Filtering: only rows where SIGNAL I/O TYPE == "AI" are shown for selection.
* Fuzzy matching via rapidfuzz (token_set_ratio) with configurable threshold.
* Output: one .xlsx per TAG; multiple tags → zipped as Datasheets.zip.

Public API
----------
    load_iodb(file)                    → (df, column_names, error)
    normalize(text)                    → str
    fuzzy_match(query, choices, ...)   → (best_match, score) | (None, 0)
    map_data_to_template(ws, row_data, col_names, threshold, log)
    generate_datasheet(template_wb, tag, row_data, col_names, threshold)
                                       → (bytes, filename)
    batch_generate_zip(iodb_file, template_file, selected_tags, ...)
                                       → (zip_bytes, "Datasheets.zip", error)
"""

from __future__ import annotations

import io
import logging
import re
import zipfile
from copy import copy
from typing import Any, Callable

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz, process as rf_process
    _HAS_RAPIDFUZZ = True
except ImportError:  # graceful fallback to difflib
    from difflib import SequenceMatcher
    _HAS_RAPIDFUZZ = False

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
SIGNAL_COL_KEYS   = ("signal i/o type", "signal type", "signal i/o", "io type")
TAG_COL_KEYS      = ("tag no", "tag number", "tag_no", "tag")
AI_FILTER_VALUE   = "AI"
DATASHEET_SHEET   = "Datasheet"          # exact sheet name (case-insensitive search)
HEADING_COL_IDX   = 4                    # Column D (1-based)
# Regex that matches any "Refer Annexure" variant:
#   "Refer Annexure 1", "Refer Annexure-1", "Refer Annexure-3", "REFER ANNEXURE", …
_PLACEHOLDER_RE = re.compile(r'refer\s*annexure', re.IGNORECASE)
FUZZY_THRESHOLD   = 70                   # minimum rapidfuzz score (0-100)
MAX_RIGHT_SEARCH  = 6                    # how many cells right to search for placeholder

# ── Normalisation ──────────────────────────────────────────────────────────────

def normalize(text: Any) -> str:
    """
    Lowercase, strip, collapse whitespace, remove special characters.
    Used for both header matching and placeholder detection.

    >>> normalize("TAG NO.")  →  "tag no"
    >>> normalize("Signal I/O Type")  →  "signal io type"
    """
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)   # replace punctuation with space
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ── Fuzzy matching ─────────────────────────────────────────────────────────────

def fuzzy_match(
    query: str,
    choices: list[str],
    threshold: int = FUZZY_THRESHOLD,
) -> tuple[str | None, int]:
    """
    Return (best_matching_choice, score) where score is 0-100.
    Returns (None, 0) when no choice meets the threshold.

    Uses rapidfuzz.fuzz.token_set_ratio when available, falling back to
    difflib.SequenceMatcher.  token_set_ratio is robust against word-order
    differences and partial overlaps ("Line Size (NB)" ↔ "Line Size").
    """
    if not choices:
        return None, 0

    norm_query   = normalize(query)
    norm_choices = [normalize(c) for c in choices]

    if _HAS_RAPIDFUZZ:
        # WRatio handles sub-string and word-order differences well
        result = rf_process.extractOne(
            norm_query,
            norm_choices,
            scorer=fuzz.WRatio,
            score_cutoff=threshold,
        )
        if result is None:
            return None, 0
        matched_norm, score, idx = result
        return choices[idx], int(score)
    else:
        # difflib fallback
        best_score = 0
        best_idx   = -1
        for i, nc in enumerate(norm_choices):
            ratio = int(SequenceMatcher(None, norm_query, nc).ratio() * 100)
            if ratio > best_score:
                best_score = ratio
                best_idx   = i
        if best_score >= threshold and best_idx >= 0:
            return choices[best_idx], best_score
        return None, 0


# ── IODB loading ───────────────────────────────────────────────────────────────

def _find_sheet(xls: pd.ExcelFile) -> str | None:
    """Find the best sheet name for the IODB (prefers 'IODB', then tag-containing)."""
    names = xls.sheet_names
    # 1. exact
    match = next((s for s in names if s.strip().lower() == "iodb"), None)
    if match:
        return match
    # 2. substring
    match = next((s for s in names if "iodb" in s.strip().lower()), None)
    if match:
        return match
    # 3. probe for tag-like column in first sheet
    for s in names:
        try:
            sample = xls.parse(s, nrows=5)
            cols_lower = [normalize(c) for c in sample.columns]
            if any(k in cols_lower for k in TAG_COL_KEYS):
                return s
        except Exception:
            continue
    return names[0] if names else None


def _combine_two_row_header(raw_df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    When an IODB has a TWO-ROW header (row 0 = group labels, row 1 = field names),
    merge them into a single header row.

    Strategy:
      - For each column, if row-0 and row-1 are both non-empty and different,
        the combined name is "<row0> <row1>" (stripped).
      - If only one is non-empty, use that.
      - De-duplicate by appending ._1, ._2, … as needed.

    Returns (df_without_header_rows, combined_column_names).
    """
    if len(raw_df) < 2:
        return raw_df, list(raw_df.columns)

    row0 = [str(v).strip() if v is not None and not _is_nan(v) else "" for v in raw_df.iloc[0]]
    row1 = [str(v).strip() if v is not None and not _is_nan(v) else "" for v in raw_df.iloc[1]]
    old_cols = [str(c).strip() if not str(c).startswith("Unnamed") else "" for c in raw_df.columns]

    combined: list[str] = []
    for i in range(len(old_cols)):
        parts = []
        if old_cols[i]:
            parts.append(old_cols[i])
        if row0[i] and row0[i] not in parts:
            parts.append(row0[i])
        if row1[i] and row1[i] not in parts:
            parts.append(row1[i])
        combined.append(" ".join(parts) if parts else f"col_{i}")

    # De-duplicate
    seen: dict[str, int] = {}
    final: list[str] = []
    for name in combined:
        if name in seen:
            seen[name] += 1
            final.append(f"{name}.{seen[name]}")
        else:
            seen[name] = 0
            final.append(name)

    data_df = raw_df.iloc[2:].reset_index(drop=True)
    data_df.columns = final
    return data_df, final


def _is_nan(v: Any) -> bool:
    try:
        return v is None or (isinstance(v, float) and pd.isna(v))
    except Exception:
        return False


def load_iodb(
    uploaded_file,
    two_row_header: bool = True,
) -> tuple[pd.DataFrame | None, list[str], str | None]:
    """
    Load the IODB Excel file.

    Parameters
    ----------
    uploaded_file : file-like (BytesIO or UploadedFile)
    two_row_header : bool
        When True, the first two data rows are treated as a combined header.

    Returns
    -------
    (df, column_names, error_or_None)
        df            – DataFrame with data rows only
        column_names  – final list of column name strings
        error         – error message, or None on success
    """
    try:
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
        buf = io.BytesIO(uploaded_file.read()) if not isinstance(uploaded_file, io.BytesIO) else uploaded_file
        buf.seek(0)

        xls       = pd.ExcelFile(buf)
        sheet     = _find_sheet(xls)
        if sheet is None:
            return None, [], "No suitable sheet found in the IODB file."

        buf.seek(0)
        # Read without assuming any header so we control two-row merging
        raw = pd.read_excel(buf, sheet_name=sheet, header=0)
        raw = raw.dropna(how="all")

        if two_row_header:
            df, col_names = _combine_two_row_header(raw)
        else:
            raw = raw.reset_index(drop=True)
            col_names = list(raw.columns)
            df = raw

        df = df.dropna(how="all").reset_index(drop=True)
        logger.info("load_iodb: loaded %d rows × %d cols from sheet '%s'", len(df), len(col_names), sheet)
        return df, col_names, None

    except Exception as exc:
        logger.exception("load_iodb failed")
        return None, [], str(exc)


# ── Tag extraction ─────────────────────────────────────────────────────────────

def _find_col(col_names: list[str], keys: tuple[str, ...]) -> str | None:
    """Find the first column whose normalised name contains any of the (also normalised) keys."""
    norm_keys = [normalize(k) for k in keys]
    for name in col_names:
        n = normalize(name)
        if any(k in n for k in norm_keys):
            return name
    return None


def get_ai_tags(
    df: pd.DataFrame,
    col_names: list[str],
) -> tuple[list[str], str | None]:
    """
    Return list of TAG NO values where SIGNAL I/O TYPE == 'AI'.
    Returns ([], error_message) on failure.
    """
    tag_col    = _find_col(col_names, TAG_COL_KEYS)
    signal_col = _find_col(col_names, SIGNAL_COL_KEYS)

    if tag_col is None:
        return [], f"TAG NO column not found. Available: {col_names[:10]}"
    if signal_col is None:
        # If no signal column, return all tags with a warning
        logger.warning("SIGNAL I/O TYPE column not found — returning all tags")
        tags = (
            df[tag_col]
            .dropna()
            .astype(str)
            .str.strip()
            .loc[lambda s: s != ""]
            .unique()
            .tolist()
        )
        return sorted(tags), None

    # Filter AI rows
    mask = df[signal_col].astype(str).str.strip().str.upper() == AI_FILTER_VALUE
    tags = (
        df.loc[mask, tag_col]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )
    logger.info("get_ai_tags: found %d AI tags", len(tags))
    return sorted(tags), None


# ── Template helpers ───────────────────────────────────────────────────────────

def _get_datasheet_ws(template_wb: openpyxl.Workbook):
    """
    Return (worksheet, error).  Looks for 'Datasheet' (case-insensitive),
    then Sheet2, then the second sheet of the workbook.
    """
    names = template_wb.sheetnames

    # 1. exact case-insensitive match on "Datasheet"
    ws_name = next((s for s in names if s.strip().lower() == "datasheet"), None)
    if ws_name:
        return template_wb[ws_name], None

    # 2. Sheet2 or "Sheet 2"
    ws_name = next(
        (s for s in names if re.sub(r"\s+", "", s).lower() in ("sheet2", "sheet_2")),
        None,
    )
    if ws_name:
        return template_wb[ws_name], None

    # 3. second sheet in workbook
    if len(names) >= 2:
        return template_wb[names[1]], None

    return None, (
        f"Template does not contain a 'Datasheet' sheet. "
        f"Found: {names}"
    )


def _is_placeholder(value: Any) -> bool:
    """True when the cell may be overwritten (empty or any 'Refer Annexure …' text)."""
    if value is None:
        return True
    if isinstance(value, str):
        s = value.strip()
        return s == "" or bool(_PLACEHOLDER_RE.search(s))
    return False


def _is_explicit_placeholder(value: Any) -> bool:
    """
    True ONLY for cells whose text matches 'Refer Annexure …' — NOT for empty/None.
    Handles every variant: 'Refer Annexure 1', 'Refer Annexure-1', 'Refer Annexure-3', …
    """
    if not isinstance(value, str):
        return False
    return bool(_PLACEHOLDER_RE.search(value.strip()))


# ── Core mapping ───────────────────────────────────────────────────────────────

def map_data_to_template(
    ws,
    row_data: dict[str, Any],
    col_names: list[str],
    threshold: int = FUZZY_THRESHOLD,
    log: list[dict] | None = None,
) -> None:
    """
    Populate a Datasheet worksheet in-place.

    Scans Column D for parameter headings.  For each heading, looks rightward
    (columns E, F, …) for a placeholder cell and fills it with the matched
    IODB value.

    Strategy (inverted scan)
    ------------------------
    Rather than starting from Column D headings and looking rightward, this
    function scans the ENTIRE sheet for cells that explicitly contain
    "Refer Annexure 1".  For each such cell it looks LEFTWARD along the same
    row to find the nearest non-empty heading cell, matches that heading to an
    IODB column via fuzzy matching, and overwrites the "Refer Annexure 1" cell
    with the IODB value.

    This handles all template layouts regardless of how many blank columns
    sit between the heading and the placeholder.

    Parameters
    ----------
    ws        : openpyxl Worksheet (Datasheet sheet)
    row_data  : {column_name → value} for the selected TAG row
    col_names : list of IODB column names (for fuzzy matching)
    threshold : minimum fuzzy match score
    log       : if provided, mapping results are appended as dicts
    """
    # Build normalised lookup: norm_col_name → original_col_name
    norm_to_orig: dict[str, str] = {normalize(c): c for c in col_names}
    norm_choices: list[str]      = list(norm_to_orig.keys())

    for row in ws.iter_rows():
        for cell in row:
            # ── Only target explicit "Refer Annexure 1" cells ─────────────────
            if not _is_explicit_placeholder(cell.value):
                continue
            # Never touch formula cells
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                continue
            # Skip read-only merged-cell subordinates
            if cell.__class__.__name__ == "MergedCell":
                continue

            # ── Find heading by scanning LEFTWARD in the same row ─────────────
            # Collect every non-empty, non-placeholder cell to the left.
            # Reading left→right order (rightmost collected last), the LAST
            # entry is the immediate label (sub-heading like "Operating") and
            # the FIRST entry is the main paragraph heading (e.g. "Pressure").
            # We join them to form a compound key: "Pressure Operating".
            left_labels: list[str] = []
            for check_col in range(cell.column - 1, 0, -1):
                check_cell = ws.cell(row=cell.row, column=check_col)
                # Skip merged-cell subordinates (they carry no value)
                if check_cell.__class__.__name__ == "MergedCell":
                    continue
                cv = check_cell.value
                if cv is None or not str(cv).strip():
                    continue
                cv_str = str(cv).strip()
                # Skip formula cells and other placeholder cells in this row
                if cv_str.startswith("=") or _is_explicit_placeholder(cv_str):
                    continue
                left_labels.insert(0, cv_str)  # prepend → left-to-right order

            if not left_labels:
                logger.debug(
                    "map_data_to_template: no heading found left of %s%d — skipping",
                    get_column_letter(cell.column), cell.row,
                )
                continue

            # Build compound heading (e.g. "Pressure" + "Operating" → "Pressure Operating")
            # Try progressively longer combinations starting from the full compound
            # down to just the immediate label, and use whichever fuzzy-matches best.
            heading_candidates: list[str] = []
            for start in range(len(left_labels)):
                heading_candidates.append(" ".join(left_labels[start:]))

            heading_text = heading_candidates[0]  # default = full compound
            if not heading_text:
                logger.debug(
                    "map_data_to_template: no heading found left of %s%d — skipping",
                    get_column_letter(cell.column), cell.row,
                )
                continue

            # ── Resolve IODB column — try each heading candidate ───────────────
            iodb_col_orig: str | None = None
            match_score = 0
            norm_heading = normalize(heading_text)

            for candidate in heading_candidates:
                nc = normalize(candidate)
                if nc in norm_to_orig:
                    iodb_col_orig = norm_to_orig[nc]
                    match_score   = 100
                    norm_heading  = nc
                    break
                best, score = fuzzy_match(nc, norm_choices, threshold)
                if best is not None and score > match_score:
                    iodb_col_orig = norm_to_orig[best]
                    match_score   = score
                    norm_heading  = nc

            # ── Get IODB value ─────────────────────────────────────────────────
            if iodb_col_orig is None:
                iodb_value = "N/A"
                if log is not None:
                    log.append({
                        "heading":  heading_text,
                        "iodb_col": None,
                        "score":    0,
                        "value":    "N/A",
                        "status":   "UNMATCHED",
                    })
                logger.debug("Unmatched heading: '%s'", heading_text)
            else:
                raw = row_data.get(iodb_col_orig)
                iodb_value = "N/A" if (raw is None or _is_nan(raw)) else raw
                if log is not None:
                    log.append({
                        "heading":  heading_text,
                        "iodb_col": iodb_col_orig,
                        "score":    match_score,
                        "value":    iodb_value,
                        "status":   "MATCHED",
                    })

            # ── Overwrite the "Refer Annexure 1" cell ─────────────────────────
            cell.value = iodb_value


# ── Single datasheet ───────────────────────────────────────────────────────────

def generate_datasheet(
    template_bytes: bytes,
    tag: str,
    row_data: dict[str, Any],
    col_names: list[str],
    threshold: int = FUZZY_THRESHOLD,
) -> tuple[bytes, str, list[dict]]:
    """
    Produce one populated Datasheet workbook for the given tag.

    Accepts raw template bytes so that a fresh workbook is loaded per call —
    this avoids the closed-file / image-buffer error that occurs when openpyxl
    tries to re-read image data from an already-consumed stream.

    Returns
    -------
    (excel_bytes, filename, mapping_log)
    """
    mapping_log: list[dict] = []

    # Load a completely fresh copy of the template from the original bytes.
    # Never try to .save() a workbook that was loaded from a stream and then
    # re-open it — openpyxl holds lazy file references for embedded images.
    new_wb = load_workbook(io.BytesIO(template_bytes))

    ws, err = _get_datasheet_ws(new_wb)
    if err:
        logger.warning("generate_datasheet (%s): %s", tag, err)
    else:
        map_data_to_template(ws, row_data, col_names, threshold, mapping_log)

    out = io.BytesIO()
    new_wb.save(out)
    out.seek(0)
    safe_tag = re.sub(r'[\\/*?:\[\]]', "_", str(tag))
    filename = f"{safe_tag}_Datasheet.xlsx"

    logger.info(
        "generate_datasheet: tag=%s, file=%s, matched=%d/%d headings",
        tag, filename,
        sum(1 for e in mapping_log if e["status"] == "MATCHED"),
        len(mapping_log),
    )
    return out.read(), filename, mapping_log


# ── Batch generation ───────────────────────────────────────────────────────────

def batch_generate_zip(
    iodb_file,
    template_file,
    selected_tags: list[str],
    two_row_header: bool = True,
    threshold: int = FUZZY_THRESHOLD,
    progress_callback: Callable[[int, int], None] | None = None,
) -> tuple[bytes | None, str, str | None, list[dict]]:
    """
    Full pipeline: load IODB → load template → generate one xlsx per tag →
    package into ZIP.

    Parameters
    ----------
    iodb_file       : file-like
    template_file   : file-like
    selected_tags   : list of TAG NO strings to generate
    two_row_header  : whether IODB uses a two-row combined header
    threshold       : fuzzy match threshold (0-100)
    progress_callback : optional fn(current_idx, total)

    Returns
    -------
    (zip_bytes, "Datasheets.zip", error_or_None, full_mapping_log)
    """
    all_logs: list[dict] = []

    # ── Load IODB ──────────────────────────────────────────────────────────────
    df, col_names, err = load_iodb(iodb_file, two_row_header=two_row_header)
    if err:
        return None, "Datasheets.zip", f"Failed to load IODB: {err}", []

    tag_col = _find_col(col_names, TAG_COL_KEYS)
    if tag_col is None:
        return None, "Datasheets.zip", "TAG NO column not found in IODB.", []

    # ── Load template (keep raw bytes for per-tag reloading) ───────────────────
    try:
        if hasattr(template_file, "seek"):
            template_file.seek(0)
        tmpl_bytes = template_file.read()
    except Exception as exc:
        return None, "Datasheets.zip", f"Failed to read template: {exc}", []

    # Validate template has a Datasheet sheet (use a throw-away workbook)
    try:
        _validation_wb = load_workbook(io.BytesIO(tmpl_bytes))
    except Exception as exc:
        return None, "Datasheets.zip", f"Failed to load template: {exc}", []
    _, sheet_err = _get_datasheet_ws(_validation_wb)
    if sheet_err:
        return None, "Datasheets.zip", sheet_err, []

    # ── Generate ───────────────────────────────────────────────────────────────
    results: list[tuple[bytes, str]] = []
    total = len(selected_tags)

    for idx, tag in enumerate(selected_tags):
        if progress_callback:
            progress_callback(idx, total)

        # Locate IODB row for this tag
        mask = df[tag_col].astype(str).str.strip() == str(tag).strip()
        if not mask.any():
            logger.warning("batch_generate_zip: tag '%s' not found in IODB — skipping", tag)
            continue

        row_series = df[mask].iloc[0]
        row_data   = {c: row_series[c] for c in col_names if c in row_series.index}

        xlsx_bytes, filename, log = generate_datasheet(
            tmpl_bytes, tag, row_data, col_names, threshold
        )
        for entry in log:
            entry["tag"] = tag
        all_logs.extend(log)
        results.append((xlsx_bytes, filename))

    if progress_callback:
        progress_callback(total, total)

    if not results:
        return None, "Datasheets.zip", "No datasheets generated — check selected tags exist in IODB.", all_logs

    # ── Zip ────────────────────────────────────────────────────────────────────
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_bytes, fname in results:
            zf.writestr(fname, file_bytes)
    zip_buf.seek(0)

    logger.info("batch_generate_zip: packaged %d files", len(results))
    return zip_buf.read(), "Datasheets.zip", None, all_logs
