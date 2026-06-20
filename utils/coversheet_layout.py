"""
coversheet_layout.py
Builds a JSON-serialisable layout (grid + styles + images) for the cover
sheet preview, and applies preview edits (cell style overrides, image
placements) back onto an openpyxl worksheet for final generation.
"""

import base64
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

EMU_PER_PX = 9525
DEFAULT_COL_WIDTH_CHARS = 8.43
DEFAULT_ROW_HEIGHT_PT = 15.0

MAX_COL = 16  # A..P
MAX_ROW = 24


def col_width_px(width_chars):
    if width_chars is None:
        width_chars = DEFAULT_COL_WIDTH_CHARS
    return round(width_chars * 7 + 5)


def row_height_px(height_pt):
    if height_pt is None:
        height_pt = DEFAULT_ROW_HEIGHT_PT
    return round(height_pt * 4 / 3)


def get_col_widths_px(ws):
    widths = []
    for col in range(1, MAX_COL + 1):
        letter = get_column_letter(col)
        dim = ws.column_dimensions.get(letter)
        widths.append(col_width_px(dim.width if dim and dim.width else None))
    return widths


def get_row_heights_px(ws):
    heights = []
    for row in range(1, MAX_ROW + 1):
        dim = ws.row_dimensions.get(row)
        heights.append(row_height_px(dim.height if dim and dim.height else None))
    return heights


def _color_to_hex(color):
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) == 8:
        return "#" + rgb[2:]
    return None


def _border_side(side):
    if side is None or side.style is None:
        return None
    return {"style": side.style, "color": _color_to_hex(side.color) or "#000000"}


def extract_style(cell):
    font = cell.font
    alignment = cell.alignment
    fill = cell.fill
    border = cell.border
    return {
        "font": {
            "name": font.name,
            "size": font.size,
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "underline": bool(font.underline),
            "color": _color_to_hex(font.color),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrapText": bool(alignment.wrap_text),
            "shrinkToFit": bool(alignment.shrink_to_fit),
        },
        "fill": _color_to_hex(fill.fgColor) if fill and fill.fill_type == "solid" else None,
        "border": {
            "top": _border_side(border.top),
            "right": _border_side(border.right),
            "bottom": _border_side(border.bottom),
            "left": _border_side(border.left),
        },
    }


def build_grid(ws):
    """Return cell grid + merge spans for the A1:P24 range."""
    merge_master = {}  # (row, col) -> (rowspan, colspan)
    merged_away = set()  # (row, col) cells covered by a merge but not the master

    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        if min_row > MAX_ROW or min_col > MAX_COL:
            continue
        rowspan = min(max_row, MAX_ROW) - min_row + 1
        colspan = min(max_col, MAX_COL) - min_col + 1
        merge_master[(min_row, min_col)] = (rowspan, colspan)
        for r in range(min_row, min(max_row, MAX_ROW) + 1):
            for c in range(min_col, min(max_col, MAX_COL) + 1):
                if (r, c) != (min_row, min_col):
                    merged_away.add((r, c))

    cells = []
    for row in range(1, MAX_ROW + 1):
        for col in range(1, MAX_COL + 1):
            if (row, col) in merged_away:
                continue
            cell = ws.cell(row=row, column=col)
            rowspan, colspan = merge_master.get((row, col), (1, 1))
            cells.append({
                "row": row,
                "col": col,
                "coord": cell.coordinate,
                "value": "" if cell.value is None else str(cell.value),
                "rowspan": rowspan,
                "colspan": colspan,
                "style": extract_style(cell),
            })

    return cells


def _cumulative(sizes):
    """Return list of cumulative offsets (offset[i] = sum of sizes[0..i-1])."""
    offsets = [0]
    for s in sizes:
        offsets.append(offsets[-1] + s)
    return offsets


def build_layout(ws):
    col_widths = get_col_widths_px(ws)
    row_heights = get_row_heights_px(ws)
    return {
        "cols": col_widths,
        "rows": row_heights,
        "width": sum(col_widths),
        "height": sum(row_heights),
        "cells": build_grid(ws),
    }


def extract_existing_images(ws):
    """Return existing embedded images (e.g. WABAG logo) as preview entries."""
    col_widths = get_col_widths_px(ws)
    row_heights = get_row_heights_px(ws)
    col_offsets = _cumulative(col_widths)
    row_offsets = _cumulative(row_heights)

    images = []
    for img in ws._images:
        anchor = img.anchor
        frm = anchor._from
        x = col_offsets[frm.col] + round(frm.colOff / EMU_PER_PX)
        y = row_offsets[frm.row] + round(frm.rowOff / EMU_PER_PX)
        if hasattr(anchor, "to") and anchor.to is not None:
            to = anchor.to
            x2 = col_offsets[to.col] + round(to.colOff / EMU_PER_PX)
            y2 = row_offsets[to.row] + round(to.rowOff / EMU_PER_PX)
            width, height = x2 - x, y2 - y
        else:
            width = round(getattr(anchor.ext, "cx", 0) / EMU_PER_PX)
            height = round(getattr(anchor.ext, "cy", 0) / EMU_PER_PX)
        images.append({
            "x": x, "y": y, "width": width, "height": height,
            "data": base64.b64encode(img._data()).decode("ascii"),
        })
    return images


def px_to_anchor(x_px, y_px, width_px, height_px, col_widths, row_heights):
    """Convert absolute pixel position/size into a OneCellAnchor."""
    col_offsets = _cumulative(col_widths)
    row_offsets = _cumulative(row_heights)

    col_idx = 0
    for i in range(len(col_widths)):
        if x_px < col_offsets[i + 1] or i == len(col_widths) - 1:
            col_idx = i
            break
    row_idx = 0
    for i in range(len(row_heights)):
        if y_px < row_offsets[i + 1] or i == len(row_heights) - 1:
            row_idx = i
            break

    col_off_px = max(0, x_px - col_offsets[col_idx])
    row_off_px = max(0, y_px - row_offsets[row_idx])

    return OneCellAnchor(
        _from=AnchorMarker(
            col=col_idx,
            colOff=int(col_off_px * EMU_PER_PX),
            row=row_idx,
            rowOff=int(row_off_px * EMU_PER_PX),
        ),
        ext=XDRPositiveSize2D(cx=int(width_px * EMU_PER_PX), cy=int(height_px * EMU_PER_PX)),
    )


def place_image(ws, image_bytes, placement, col_widths, row_heights):
    """Add an image to the worksheet at the given pixel placement.

    placement: {"x": .., "y": .., "width": .., "height": ..}
    """
    img = XLImage(__import__("io").BytesIO(image_bytes))
    img.anchor = px_to_anchor(
        placement["x"], placement["y"], placement["width"], placement["height"],
        col_widths, row_heights,
    )
    ws.add_image(img)


def apply_cell_overrides(ws, overrides):
    """Apply value / alignment / font overrides to cells.

    overrides: {"A1": {"value": "...", "alignment": {...}, "font": {...}}, ...}
    Only the keys present in each override dict are changed; everything
    else (borders, fills, merges, sizing) is left untouched.
    """
    for coord, override in overrides.items():
        cell = ws[coord]

        # Direct value override (from canvas cell editor)
        if "value" in override:
            cell.value = override["value"]

        align_override = override.get("alignment")
        if align_override:
            current = cell.alignment
            cell.alignment = Alignment(
                horizontal=align_override.get("horizontal", current.horizontal),
                vertical=align_override.get("vertical", current.vertical),
                wrap_text=align_override.get("wrapText", current.wrap_text),
                shrink_to_fit=align_override.get("shrinkToFit", current.shrink_to_fit),
            )

        font_override = override.get("font")
        if font_override:
            current = cell.font
            cell.font = Font(
                name=current.name,
                size=font_override.get("size", current.size),
                bold=font_override.get("bold", current.bold),
                italic=font_override.get("italic", current.italic),
                underline="single" if font_override.get("underline", current.underline == "single") else None,
                color=current.color,
            )
