"""
Smart Datasheet Intelligence Engine (SDIE) router.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from fastapi.responses import StreamingResponse
import io
import json as _json
import uuid
import logging
import threading

from ..deps import get_current_user
from ..auth.models import User
from ..models.sdie import (
    ExtractedSpec,
    InstrumentSection,
    ExtractionResponse,
    ExtractionJobResponse,
    ExtractionStatusResponse,
    MappingLogEntry,
    GenerateResponse,
    GenerateJobResponse,
    GenerateStatusResponse,
)

import sys
sys.path.insert(0, '/app')
from utils.sdie_extractor import (
    extract_specifications,
    generate_populated_datasheet,
    specs_to_excel_bytes,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/sdie", tags=["Smart Datasheet Intelligence"])

extraction_jobs: dict = {}
generation_jobs: dict = {}
extraction_cache: dict = {}
generation_cache: dict = {}

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tif", ".tiff"}
ALLOWED_TEMPLATE_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


def _build_extraction_response(sections: list[dict], page_count: int, instrument_type: str) -> ExtractionResponse:
    return ExtractionResponse(
        sections=[
            InstrumentSection(
                heading=s["heading"],
                instrument_type=s["instrument_type"],
                specs=[ExtractedSpec(**spec) for spec in s["specs"]],
            )
            for s in sections
        ],
        page_count=page_count,
        message=f"Found {len(sections)} instrument section(s)",
        instrument_type=instrument_type,
    )


def _run_extraction_job(job_id: str, file_bytes: bytes, filename: str, user_id: int, template_bytes: bytes | None = None):
    try:
        sections, page_count, instrument_type = extract_specifications(file_bytes, filename, template_bytes)

        extraction_cache[f"{user_id}_sdie_extract"] = {
            "bytes": specs_to_excel_bytes(sections),
            "filename": "Extracted_Specifications.xlsx",
        }

        extraction_jobs[job_id] = {
            "status": "done",
            "result": _build_extraction_response(sections, page_count, instrument_type),
        }
        total = sum(len(s["specs"]) for s in sections)
        logger.info("SDIE extract job %s done: %d section(s), %d spec(s)", job_id, len(sections), total)
    except Exception as exc:
        logger.exception("SDIE extract job %s failed", job_id)
        extraction_jobs[job_id] = {"status": "error", "error": str(exc)}


@router.post("/extract", response_model=ExtractionJobResponse, status_code=status.HTTP_202_ACCEPTED)
async def extract_datasheet(
    file: UploadFile = File(...),
    template_file: UploadFile | None = File(default=None),
    current_user: User = Depends(get_current_user),
):
    filename = file.filename or ""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type '{ext}'. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
        )

    file_bytes = await file.read()
    template_bytes = await template_file.read() if template_file else None

    job_id = str(uuid.uuid4())
    extraction_jobs[job_id] = {"status": "processing"}

    threading.Thread(
        target=_run_extraction_job,
        args=(job_id, file_bytes, filename, current_user.id, template_bytes),
        daemon=True,
    ).start()

    return ExtractionJobResponse(job_id=job_id, status="processing")


@router.get("/extract/status/{job_id}", response_model=ExtractionStatusResponse)
async def get_extraction_status(job_id: str, current_user: User = Depends(get_current_user)):
    job = extraction_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Extraction job not found")
    return ExtractionStatusResponse(
        job_id=job_id,
        status=job["status"],
        result=job.get("result"),
        error=job.get("error"),
    )


@router.get("/extract/download")
async def download_extracted_specs(current_user: User = Depends(get_current_user)):
    cached = extraction_cache.get(f"{current_user.id}_sdie_extract")
    if not cached:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No extracted specifications found.")
    return StreamingResponse(
        io.BytesIO(cached["bytes"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={cached['filename']}"},
    )


@router.post("/map-to-template")
async def map_specs_to_template(
    specs: str = Form(...),
    template_file: UploadFile = File(...),
    heading: str = Form(default=""),
    current_user: User = Depends(get_current_user),
):
    """Map a pre-reviewed list of specs directly onto a WABAG datasheet template.

    Returns a populated Excel file immediately (no job polling needed).
    """
    try:
        specs_list = _json.loads(specs)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid specs JSON")

    template_ext = ""
    if "." in (template_file.filename or ""):
        template_ext = "." + template_file.filename.rsplit(".", 1)[-1].lower()
    if template_ext not in ALLOWED_TEMPLATE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported template type '{template_ext}'. Allowed: .xlsx .xls .xlsm",
        )

    template_bytes = await template_file.read()

    # Handle .xls format (convert to xlsx bytes via xlrd/openpyxl if needed)
    if template_ext == ".xls":
        try:
            import xlrd
            import openpyxl
            xls_wb = xlrd.open_workbook(file_contents=template_bytes)
            xlsx_wb = openpyxl.Workbook()
            xlsx_wb.remove(xlsx_wb.active)
            for sheet_name in xls_wb.sheet_names():
                xls_ws = xls_wb.sheet_by_name(sheet_name)
                xlsx_ws = xlsx_wb.create_sheet(title=sheet_name)
                for row in range(xls_ws.nrows):
                    for col in range(xls_ws.ncols):
                        xlsx_ws.cell(row=row + 1, column=col + 1, value=xls_ws.cell_value(row, col))
            buf = io.BytesIO()
            xlsx_wb.save(buf)
            template_bytes = buf.getvalue()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f".xls conversion failed: {exc}")

    out_bytes, mapping_log, err = generate_populated_datasheet(template_bytes, specs_list)
    if err:
        raise HTTPException(status_code=400, detail=err)

    safe = "".join(c for c in heading if c.isalnum() or c in " _-").strip() or "Populated"
    filename = f"Datasheet_{safe}.xlsx"

    return StreamingResponse(
        io.BytesIO(out_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _run_generation_job(job_id: str, file_bytes: bytes, filename: str, template_bytes: bytes, user_id: int):
    try:
        sections, page_count, instrument_type = extract_specifications(file_bytes, filename, template_bytes)
        # Flatten all sections for legacy single-template mapping
        flat_specs = [spec for sec in sections for spec in sec["specs"]]

        out_bytes, mapping_log, err = generate_populated_datasheet(template_bytes, flat_specs)
        if err:
            generation_jobs[job_id] = {"status": "error", "error": err}
            return

        out_filename = "Datasheet_Populated.xlsx"
        generation_cache[f"{user_id}_sdie_generate"] = {"bytes": out_bytes, "filename": out_filename}
        extraction_cache[f"{user_id}_sdie_extract"] = {
            "bytes": specs_to_excel_bytes(sections),
            "filename": "Extracted_Specifications.xlsx",
        }

        matched = sum(1 for e in mapping_log if e["status"] == "MATCHED")
        generation_jobs[job_id] = {
            "status": "done",
            "result": GenerateResponse(
                specs=[ExtractedSpec(**s) for s in flat_specs],
                page_count=page_count,
                mapping_log=[MappingLogEntry(**e) for e in mapping_log],
                filename=out_filename,
                message=f"Mapped {matched}/{len(mapping_log)} template field(s)",
                instrument_type=instrument_type,
            ),
        }
        logger.info("SDIE generate job %s done: %d spec(s), %d matched", job_id, len(flat_specs), matched)
    except Exception as exc:
        logger.exception("SDIE generate job %s failed", job_id)
        generation_jobs[job_id] = {"status": "error", "error": str(exc)}


@router.post("/generate", response_model=GenerateJobResponse, status_code=status.HTTP_202_ACCEPTED)
async def generate_datasheet(
    file: UploadFile = File(...),
    template_file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    filename = file.filename or ""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type '{ext}'.",
        )

    template_ext = ("." + (template_file.filename or "").rsplit(".", 1)[-1].lower()) if "." in (template_file.filename or "") else ""
    if template_ext not in ALLOWED_TEMPLATE_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported template type '{template_ext}'.",
        )

    file_bytes = await file.read()
    template_bytes = await template_file.read()

    job_id = str(uuid.uuid4())
    generation_jobs[job_id] = {"status": "processing"}

    threading.Thread(
        target=_run_generation_job,
        args=(job_id, file_bytes, filename, template_bytes, current_user.id),
        daemon=True,
    ).start()

    return GenerateJobResponse(job_id=job_id, status="processing")


@router.get("/generate/status/{job_id}", response_model=GenerateStatusResponse)
async def get_generation_status(job_id: str, current_user: User = Depends(get_current_user)):
    job = generation_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Generation job not found")
    return GenerateStatusResponse(
        job_id=job_id,
        status=job["status"],
        result=job.get("result"),
        error=job.get("error"),
    )


@router.get("/download")
async def download_populated_datasheet(current_user: User = Depends(get_current_user)):
    cached = generation_cache.get(f"{current_user.id}_sdie_generate")
    if not cached:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No populated datasheet found.")
    return StreamingResponse(
        io.BytesIO(cached["bytes"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={cached['filename']}"},
    )
