"""
utils/lt_radar_datasheet.py
============================
LT Non-Contact Radar Datasheet automation engine.

Responsibilities:
  1.  SPEC_FIELDS   – ordered field definitions (id, label, color, dropdown options)
  2.  iodb_lookup() – fetch GREEN field values from an uploaded IODB workbook for a tag
  3.  generate()    – produce a clean .xlsx datasheet from a values dict

Color roles (template):
  GREEN  (idx 57) → IODB auto-fill
  RED    (idx 10) → user selection / AI extraction
  GREY   (idx 13) → manual header entry (CLIENT/CONSULTANT/PROJECT/LOCATION)
  YELLOW (idx  9) → fixed pre-filled value (SS Tag Plate = PROVIDED, Transmitter Type)
  VIOLET (idx 62) → PURCHASE + NOTES manual entry
  (In this template GREY idx=13 & VIOLET idx=13 overlap – we distinguish by position.)
"""
from __future__ import annotations

import io
import re
from typing import Any

import pandas as pd

# ---------------------------------------------------------------------------
# 1. FIELD DEFINITIONS
# ---------------------------------------------------------------------------

DROPDOWN_OPTIONS: dict[str, list[str]] = {
    "frequency":          ["6 GHz", "10 GHz", "24 GHz", "26 GHz", "80 GHz"],
    "beam_angle_80":      ["3°", "4°", "5°", "6°"],
    "beam_angle_26":      ["8°", "10°", "12°", "14°"],
    "antenna_type":       ["Horn Antenna", "Lens Antenna", "Flush Antenna",
                           "PTFE Encapsulated Antenna", "Hygienic Antenna"],
    "antenna_material":   ["SS316", "SS316L", "PTFE", "PEEK", "PVDF", "Hastelloy C276"],
    "housing_moc":        ["Die Cast Aluminium", "Aluminium", "SS304", "SS316", "SS316L"],
    "process_conn_type":  ["Threaded", "Flanged", "ANSI Flanged", "DIN Flanged",
                           "JIS Flanged", "Tri Clamp", "Hygienic Connection"],
    "process_conn_size":  ["1\"", "1½\"", "2\"", "3\"", "4\"", "6\"", "8\"",
                           "DN25", "DN40", "DN50", "DN80", "DN100", "DN150"],
    "flange_rating":      ["ANSI 150#", "ANSI 300#", "ANSI 600#",
                           "PN10", "PN16", "PN25", "PN40"],
    "nozzle_size":        ["1\"", "1½\"", "2\"", "3\"", "4\"", "6\"",
                           "DN50", "DN80", "DN100", "DN150"],
    "mounting_pos":       ["Top Centre", "Top Offset", "Nozzle Mounted",
                           "Stand Pipe Mounted", "Stilling Well Mounted", "Chamber Mounted"],
    "temp_comp":          ["Not Required", "Built-In", "External Sensor"],
    "transmitter_type":   ["Integral", "Remote Mounted"],
    "power_supply":       ["12-30 VDC", "16-30 VDC", "18-36 VDC", "24 VDC"],
    "output":             ["4-20mA", "4-20mA + HART", "Foundation Fieldbus",
                           "Profibus PA", "Modbus RTU", "Wireless HART"],
    "hart_revision":      ["HART 5", "HART 6", "HART 7", "HART 8"],
    "accuracy":           ["±1 mm", "±2 mm", "±3 mm", "±5 mm"],
    "repeatability":      ["±1 mm", "±2 mm", "±3 mm", "±5 mm"],
    "display":            ["No Display", "LCD", "Backlit LCD", "Graphical LCD", "OLED"],
    "encl_protection":    ["IP65", "IP66", "IP67", "IP68", "NEMA 4X"],
    "cable_entry":        ["½\" NPT", "¾\" NPT", "M20 x 1.5", "M25 x 1.5"],
    "area_cert":          ["General Purpose", "Ex ia IIC T4", "Ex ia IIC T6",
                           "Ex d IIC T4", "Ex d IIC T6", "ATEX", "IECEx"],
    "sil_cert":           ["Not Required", "SIL 1", "SIL 2", "SIL 3"],
}

DEFAULTS: dict[str, str] = {
    "frequency":       "80 GHz",
    "housing_moc":     "Die Cast Aluminium",
    "temp_comp":       "Built-In",
    "transmitter_type": "Integral",
    "power_supply":    "24 VDC",
    "output":          "4-20mA + HART",
    "hart_revision":   "HART 7",
    "accuracy":        "±2 mm",
    "repeatability":   "±1 mm",
    "display":         "LCD",
    "encl_protection": "IP67",
    "cable_entry":     "M20 x 1.5",
    "sil_cert":        "Not Required",
    "sensor_type":     "Non Contact Radar",
}

# Ordered spec field definitions
# color: "green" | "red" | "grey" | "violet" | "yellow"
SPEC_FIELDS: list[dict] = [
    # ── Header block ─────────────────────────────────────────────────────────
    {"id": "client",        "label": "CLIENT",           "section": "HEADER",  "color": "grey",   "row": 4,  "value_col": 1},
    {"id": "consultant",    "label": "CONSULTANT",        "section": "HEADER",  "color": "grey",   "row": 5,  "value_col": 1},
    {"id": "project",       "label": "PROJECT",           "section": "HEADER",  "color": "grey",   "row": 6,  "value_col": 1},
    {"id": "location_hdr",  "label": "LOCATION",          "section": "HEADER",  "color": "grey",   "row": 7,  "value_col": 1},
    {"id": "issued_for",    "label": "ISSUED FOR",        "section": "HEADER",  "color": "grey",   "row": 4,  "value_col": 5, "merged_rows": 4},
    # ── General ──────────────────────────────────────────────────────────────
    {"id": "tag_no",        "label": "Tag No",            "section": "GENERAL", "color": "green",  "row": 8,  "value_col": 5},
    {"id": "pid_no",        "label": "P&ID No",           "section": "GENERAL", "color": "green",  "row": 9,  "value_col": 5},
    {"id": "location",      "label": "Location",          "section": "GENERAL", "color": "green",  "row": 10, "value_col": 5},
    {"id": "area_class",    "label": "Area Classification","section": "GENERAL", "color": "green",  "row": 11, "value_col": 5},
    # ── Process Conditions ────────────────────────────────────────────────────
    {"id": "fluid",         "label": "Fluid",             "section": "PROCESS CONDITIONS", "color": "green",  "row": 12, "value_col": 5},
    {"id": "specific_gravity","label":"Specific Gravity",  "section": "PROCESS CONDITIONS", "color": "green",  "row": 13, "value_col": 5},
    {"id": "viscosity",     "label": "Viscosity(cP)",     "section": "PROCESS CONDITIONS", "color": "green",  "row": 14, "value_col": 5},
    {"id": "pressure_min",  "label": "Pressure Min (Kg/Cm² g)","section":"PROCESS CONDITIONS","color":"green","row":15,"value_col":5},
    {"id": "pressure_nor",  "label": "Pressure Nor (Kg/Cm² g)","section":"PROCESS CONDITIONS","color":"green","row":15,"value_col":7},
    {"id": "pressure_max",  "label": "Pressure Max (Kg/Cm² g)","section":"PROCESS CONDITIONS","color":"green","row":15,"value_col":9},
    {"id": "temp_min",      "label": "Temp Min (°C)",     "section": "PROCESS CONDITIONS", "color": "green",  "row": 16, "value_col": 5},
    {"id": "temp_nor",      "label": "Temp Nor (°C)",     "section": "PROCESS CONDITIONS", "color": "green",  "row": 16, "value_col": 7},
    {"id": "temp_max",      "label": "Temp Max (°C)",     "section": "PROCESS CONDITIONS", "color": "green",  "row": 16, "value_col": 9},
    # ── Tank Details ──────────────────────────────────────────────────────────
    {"id": "tank_tag",      "label": "Tank Tag no",       "section": "TANK DETAILS", "color": "green",  "row": 17, "value_col": 5},
    {"id": "tank_type",     "label": "Tank Type",         "section": "TANK DETAILS", "color": "green",  "row": 18, "value_col": 5},
    {"id": "tank_moc",      "label": "Tank MOC",          "section": "TANK DETAILS", "color": "green",  "row": 19, "value_col": 5},
    {"id": "tank_height",   "label": "Tank Height",       "section": "TANK DETAILS", "color": "green",  "row": 20, "value_col": 5},
    {"id": "tank_diameter", "label": "Tank Diameter",     "section": "TANK DETAILS", "color": "green",  "row": 21, "value_col": 5},
    {"id": "agitation",     "label": "Presence of Agitation","section":"TANK DETAILS","color":"green",   "row": 22, "value_col": 5},
    {"id": "fumes",         "label": "Presence of Fumes", "section": "TANK DETAILS", "color": "green",  "row": 23, "value_col": 5},
    {"id": "max_fluid",     "label": "Max.Fluid Level",   "section": "TANK DETAILS", "color": "green",  "row": 24, "value_col": 5},
    # ── Sensor ────────────────────────────────────────────────────────────────
    {"id": "sensor_type",   "label": "Sensor Type",       "section": "SENSOR", "color": "red",    "row": 25, "value_col": 5, "fixed": "Non Contact Radar"},
    {"id": "frequency",     "label": "Frequency Range",   "section": "SENSOR", "color": "red",    "row": 26, "value_col": 5, "options": "frequency"},
    {"id": "beam_angle",    "label": "Beam Angle",        "section": "SENSOR", "color": "red",    "row": 27, "value_col": 5, "options_dynamic": "beam_angle"},
    {"id": "blocking_dist", "label": "Blocking Distance", "section": "SENSOR", "color": "red",    "row": 28, "value_col": 5},
    {"id": "housing_moc",   "label": "Housing MOC",       "section": "SENSOR", "color": "red",    "row": 29, "value_col": 5, "options": "housing_moc"},
    {"id": "enclosure_class","label":"Enclosure Class",   "section": "SENSOR", "color": "red",    "row": 30, "value_col": 5},
    {"id": "temp_comp",     "label": "Temperature Compensation","section":"SENSOR","color":"red",  "row": 31, "value_col": 5, "options": "temp_comp"},
    {"id": "proc_conn_moc", "label": "Process Connection MOC","section":"SENSOR","color":"red",   "row": 32, "value_col": 5, "options": "antenna_material"},
    {"id": "proc_conn_flange","label":"Process Connection Flange Size","section":"SENSOR","color":"red","row":33,"value_col":5,"options":"process_conn_size"},
    {"id": "nozzle_size",   "label": "Nozzle Size",       "section": "SENSOR", "color": "red",    "row": 34, "value_col": 5, "options": "nozzle_size"},
    {"id": "nozzle_height", "label": "Nozzle Height",     "section": "SENSOR", "color": "red",    "row": 35, "value_col": 5},
    {"id": "mounting_pos",  "label": "Mounting Position", "section": "SENSOR", "color": "red",    "row": 36, "value_col": 5, "options": "mounting_pos"},
    {"id": "cable_distance","label": "Cable distance from Sensor to Transmitter","section":"SENSOR","color":"red","row":37,"value_col":5},
    # ── Transmitter ───────────────────────────────────────────────────────────
    {"id": "transmitter_type","label":"Transmitter Type", "section": "TRANSMITTER", "color": "red","row": 38, "value_col": 5, "options": "transmitter_type"},
    {"id": "power_supply",  "label": "Power Supply",      "section": "TRANSMITTER", "color": "red","row": 39, "value_col": 5, "options": "power_supply"},
    {"id": "output",        "label": "Output Signal",     "section": "TRANSMITTER", "color": "red","row": 40, "value_col": 5, "options": "output"},
    {"id": "accuracy",      "label": "Accuracy",          "section": "TRANSMITTER", "color": "red","row": 41, "value_col": 5, "options": "accuracy"},
    {"id": "repeatability", "label": "Repeatability",     "section": "TRANSMITTER", "color": "red","row": 42, "value_col": 5, "options": "repeatability"},
    {"id": "inst_range",    "label": "Instrumentation Range","section":"TRANSMITTER","color":"red","row": 43, "value_col": 5},
    {"id": "calib_range",   "label": "Calibration Range", "section": "TRANSMITTER", "color": "red","row": 44, "value_col": 5},
    {"id": "encl_protection","label":"Enclosure Protection Class","section":"TRANSMITTER","color":"red","row":45,"value_col":5,"options":"encl_protection"},
    {"id": "cable_entry",   "label": "Cable Entry",       "section": "TRANSMITTER", "color": "red","row": 46, "value_col": 5, "options": "cable_entry"},
    {"id": "display",       "label": "Display",           "section": "TRANSMITTER", "color": "red","row": 47, "value_col": 5, "options": "display"},
    # ── Options ───────────────────────────────────────────────────────────────
    {"id": "mounting_acc",  "label": "Mounting Accessories","section":"OPTIONS",   "color": "red","row": 48, "value_col": 5},
    # Row 49 = SS Tag Plate → "PROVIDED" (fixed, no user input)
    # ── Certification ─────────────────────────────────────────────────────────
    {"id": "area_cert",     "label": "Area Certification","section": "CERTIFICATION","color":"red","row": 50, "value_col": 5, "options": "area_cert"},
    # ── Purchase ──────────────────────────────────────────────────────────────
    {"id": "make",          "label": "Make",              "section": "PURCHASE",  "color": "violet","row": 51, "value_col": 5},
    {"id": "model_no",      "label": "Model",             "section": "PURCHASE",  "color": "violet","row": 52, "value_col": 5},
    # ── Notes ─────────────────────────────────────────────────────────────────
    {"id": "notes",         "label": "Notes",             "section": "NOTES",     "color": "violet","row": 53, "value_col": 1},
]

# ---------------------------------------------------------------------------
# 2. IODB LOOKUP
# ---------------------------------------------------------------------------

_NORM = lambda s: re.sub(r"[\s_\-/]+", " ", str(s).strip().lower())

# Priority key lists for fuzzy IODB column matching (most-specific first)
_IODB_KEYS: dict[str, list[str]] = {
    "tag_no":          ["tag_number_new", "tag number new", "tag no", "tag number", "tag"],
    "pid_no":          ["p&id no", "pid no", "p id no", "pid number", "p and id no", "pid"],
    "location":        ["location", "plant area", "area"],
    "area_class":      ["area classification", "area class", "hazardous area class"],
    "fluid":           ["fluid", "process fluid", "fluid description", "medium"],
    "specific_gravity":["specific gravity", "sp gr", "s.g.", "sg"],
    "viscosity":       ["viscosity", "viscosity cp", "viscosity(cp)", "visc"],
    "pressure_min":    ["pressure min", "min pressure", "min. pressure", "p min"],
    "pressure_nor":    ["pressure nor", "normal pressure", "nor pressure", "oper pressure", "operating pressure"],
    "pressure_max":    ["pressure max", "max pressure", "design pressure", "p max"],
    "temp_min":        ["temp min", "min temp", "min temperature", "t min"],
    "temp_nor":        ["temp nor", "normal temperature", "nor temp", "operating temperature", "oper temperature"],
    "temp_max":        ["temp max", "max temp", "design temperature", "t max"],
    "tank_tag":        ["tank tag", "vessel tag", "tank no", "vessel no"],
    "tank_type":       ["tank type", "vessel type", "type of tank"],
    "tank_moc":        ["tank moc", "vessel moc", "tank material", "vessel material"],
    "tank_height":     ["tank height", "vessel height", "height"],
    "tank_diameter":   ["tank diameter", "vessel diameter", "diameter"],
    "agitation":       ["agitation", "presence of agitation", "agitator", "mixer"],
    "fumes":           ["fumes", "presence of fumes", "vapour", "vapor"],
    "max_fluid":       ["max fluid level", "maximum fluid level", "max level", "hh level"],
    "area_class_sensor":["area classification", "area class"],  # reuses area_class
}


def _find_col(norm_headers: list[str], keys: list[str]) -> int:
    """Return 0-based column index of first matching key; -1 if not found."""
    # Exact match first
    for key in keys:
        for i, h in enumerate(norm_headers):
            if h == key and "old" not in h:
                return i
    # Substring match (skip if "old" in header)
    for key in keys:
        for i, h in enumerate(norm_headers):
            if key in h and "old" not in h:
                return i
    return -1


def iodb_lookup(iodb_bytes: bytes, tag_no: str) -> tuple[dict[str, str], str | None]:
    """
    Given an IODB workbook bytes and a tag number, return a dict of
    field_id → value for all GREEN fields that can be found.
    Returns (values_dict, error_string_or_None).
    """
    try:
        df = pd.read_excel(io.BytesIO(iodb_bytes))
    except Exception as exc:
        return {}, f"Could not read IODB: {exc}"

    norm_cols = [_NORM(c) for c in df.columns]

    # Find tag column
    tag_col = _find_col(norm_cols, _IODB_KEYS["tag_no"])
    if tag_col == -1:
        return {}, "Tag column not found in IODB"

    # Find matching row
    tag_norm = _NORM(tag_no)
    matched = df[df.iloc[:, tag_col].astype(str).apply(_NORM) == tag_norm]
    if matched.empty:
        return {}, f"Tag '{tag_no}' not found in IODB"

    row = matched.iloc[0]
    result: dict[str, str] = {}

    for field_id, keys in _IODB_KEYS.items():
        if field_id == "area_class_sensor":
            continue
        col_idx = _find_col(norm_cols, keys)
        if col_idx != -1:
            val = str(row.iloc[col_idx])
            if val not in ("nan", "None", ""):
                result[field_id] = val

    # Always populate tag_no from the exact tag we selected
    result["tag_no"] = tag_no
    return result, None


def get_iodb_tags(iodb_bytes: bytes) -> tuple[list[str], str | None]:
    """Return list of all tag numbers from the IODB, AI+subsystem filtered."""
    try:
        df = pd.read_excel(io.BytesIO(iodb_bytes))
    except Exception as exc:
        return [], f"Could not read IODB: {exc}"

    norm_cols = [_NORM(c) for c in df.columns]
    tag_col = _find_col(norm_cols, _IODB_KEYS["tag_no"])
    if tag_col == -1:
        return [], "Tag column not found in IODB"

    # Optional AI + subsystem filter
    sig_keys = ["io type name", "io_type_name", "signal i/o type", "signal io type", "io type"]
    sub_keys = ["sub system", "sub_system", "subsystem", "sub-system"]
    sig_col = _find_col(norm_cols, sig_keys)
    sub_col = _find_col(norm_cols, sub_keys)

    filtered = df.copy()
    if sig_col != -1:
        filtered = filtered[filtered.iloc[:, sig_col].astype(str).str.strip().str.upper() == "AI"]
    if sub_col != -1:
        filtered = filtered[filtered.iloc[:, sub_col].astype(str).str.contains("-", na=False)]
    if filtered.empty:
        filtered = df  # fallback

    tags = [str(v) for v in filtered.iloc[:, tag_col].dropna().unique() if str(v) not in ("nan", "")]
    return sorted(tags), None


# ---------------------------------------------------------------------------
# 3. EXCEL GENERATION
# ---------------------------------------------------------------------------

def generate(values: dict[str, Any]) -> bytes:
    """
    Generate a clean LT Non-Contact Radar datasheet xlsx from a values dict.
    All color fills are removed. Template structure is fully preserved.
    Returns raw bytes of the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "LT-RADAR"

    # ── Styles ────────────────────────────────────────────────────────────────
    NO_FILL   = PatternFill(fill_type=None)
    thin      = Side(style="thin")
    medium    = Side(style="medium")
    no_side   = Side(style=None)

    def border(L=None, R=None, T=None, B=None):
        return Border(
            left   = L or no_side, right = R or no_side,
            top    = T or no_side, bottom= B or no_side,
        )

    def font(sz=9, bold=False, colour="000000"):
        return Font(name="Calibri", size=sz, bold=bold, color=colour)

    def align(h="left", v="center", wrap=False):
        ha = {"left":"left","center":"center","right":"right","general":"general"}.get(h,"left")
        return Alignment(horizontal=ha, vertical=v, wrap_text=wrap)

    def style(cell, sz=9, bold=False, ha="left", va="center", wrap=False,
              L=None, R=None, T=None, B=None, colour="000000"):
        cell.fill      = NO_FILL
        cell.font      = font(sz=sz, bold=bold, colour=colour)
        cell.alignment = align(h=ha, v=va, wrap=wrap)
        cell.border    = border(L=L, R=R, T=T, B=B)

    M = medium; TH = thin  # shorthand

    # ── Column widths (xlrd units → openpyxl approximate in chars) ────────────
    # xlrd width 5741 ≈ col A wide; 3145 ≈ col B-E; 2230 ≈ col F-K
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 12
    for col_letter in ["F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col_letter].width = 9

    # ── Row heights (xlrd units are in 1/20 pt; 345 ≈ 17.25 pt) ──────────────
    for r in range(1, 5):
        ws.row_dimensions[r].height = 15
    for r in range(5, 53):
        ws.row_dimensions[r].height = 17
    ws.row_dimensions[53].height = 17
    ws.row_dimensions[54].height = 48  # Notes row taller

    # ── Print / page setup ────────────────────────────────────────────────────
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── Merged cells (exact replication of template) ──────────────────────────
    # Header block
    ws.merge_cells("A1:E4")    # Company logo/address
    ws.merge_cells("F1:K4")    # DS title
    ws.merge_cells("F5:H8")    # ISSUED FOR (value spans 4 rows)
    ws.merge_cells("I5:K8")    # Approval box
    # Each header row label+value B-E
    for r in [5, 6, 7, 8]:
        ws.merge_cells(f"B{r}:E{r}")
    # Section-label merges in col A
    ws.merge_cells("A9:A12")   # GENERAL
    ws.merge_cells("A13:A17")  # PROCESS CONDITIONS
    ws.merge_cells("A18:A25")  # TANK DETAILS
    ws.merge_cells("A26:A38")  # SENSOR
    ws.merge_cells("A39:A48")  # TRANSMITTER
    ws.merge_cells("A49:A50")  # OPTIONS
    ws.merge_cells("A52:A53")  # PURCHASE
    # Data rows: label B-E, value F-K (most rows)
    for r in range(9, 55):
        if r in (16, 17):
            # Pressure / Temp rows: B=main label, C=Min, D=Nor, E=Max (all individual)
            # Value cells: F-G (Min), H-I (Nor), J-K (Max) — each a separate merge
            ws.merge_cells(f"F{r}:G{r}")
            ws.merge_cells(f"H{r}:I{r}")
            ws.merge_cells(f"J{r}:K{r}")
        elif r == 54:       # Notes
            ws.merge_cells(f"B{r}:K{r}")
        elif r == 39:       # Transmitter section: label col-A already merged,
            # row 39 in openpyxl = row 39 (1-indexed) = TRANSMITTER header row
            # B-E = "Type" label, F-K = "Non Contact radar"
            ws.merge_cells(f"B{r}:E{r}")
            ws.merge_cells(f"F{r}:K{r}")
        else:
            try:
                ws.merge_cells(f"B{r}:E{r}")
            except Exception:
                pass
            try:
                ws.merge_cells(f"F{r}:K{r}")
            except Exception:
                pass

    # ── Helper to write a cell (top-left of merged range) ─────────────────────
    def w(r, c, val, **kw):
        """Write val to cell at 1-indexed (r, c); apply style."""
        cell = ws.cell(row=r, column=c, value=val)
        style(cell, **kw)

    def w_value(row_1idx: int, val_col_0idx: int, val: Any):
        """Write user value into the appropriate value cell of a spec row."""
        col_1idx = val_col_0idx + 1  # convert to 1-indexed
        cell = ws.cell(row=row_1idx + 1, column=col_1idx, value=val)
        cell.fill      = NO_FILL
        cell.font      = font(sz=9)
        cell.alignment = align(h="left", v="center")

    # ── COMPANY HEADER (A1:E4) ────────────────────────────────────────────────
    hdr = ws.cell(row=1, column=1,
        value='VATECH WABAG LIMITED \n"WABAG HOUSE"\nNO.17, 200 FEET THORAIPAKKAM-PALLAVARAM ROAD\nPALLIKARANAI, CHENNAI – 600 100.')
    hdr.fill      = NO_FILL
    hdr.font      = font(sz=10)
    hdr.alignment = align(h="left", v="center", wrap=True)
    hdr.border    = border(L=M, T=M)

    # Datasheet title (F1:K4)
    ttl = ws.cell(row=1, column=6,
        value="INSTRUMENT DATASHEET - LEVEL TRANSMITTER\n(NON CONTACT RADAR TYPE)")
    ttl.fill      = NO_FILL
    ttl.font      = font(sz=11, bold=True)
    ttl.alignment = align(h="center", v="center", wrap=True)
    ttl.border    = border(R=M, T=M, B=M)

    # ── ISSUED FOR / APPROVAL (F5:H8 & I5:K8) ────────────────────────────────
    isf = ws.cell(row=5, column=6, value=values.get("issued_for", "ISSUED FOR APPROVAL"))
    isf.fill      = NO_FILL
    isf.font      = font(sz=9, bold=True)
    isf.alignment = align(h="center", v="center", wrap=True)
    isf.border    = border(L=M, T=M, R=TH, B=TH)

    apv = ws.cell(row=5, column=9, value="APPROVAL")
    apv.fill      = NO_FILL
    apv.font      = font(sz=9, bold=True)
    apv.alignment = align(h="center", v="center", wrap=True)
    apv.border    = border(R=M, T=M, B=TH)

    # ── HEADER ROWS (CLIENT / CONSULTANT / PROJECT / LOCATION) ───────────────
    _hdr_rows = [
        (5, "A", "CLIENT",     "B", "client"),
        (6, "A", "CONSULTANT", "B", "consultant"),
        (7, "A", "PROJECT",    "B", "project"),
        (8, "A", "LOCATION",   "B", "location_hdr"),
    ]
    for opx_r, la, lv, va, fid in _hdr_rows:
        # Label cell (col A)
        lc = ws.cell(row=opx_r, column=1, value=lv)
        lc.fill      = NO_FILL
        lc.font      = font(sz=9, bold=True)
        lc.alignment = align(h="left", v="center")
        lc.border    = border(L=M, R=TH, B=TH)
        # Value cell (col B, merged B-E)
        vc = ws.cell(row=opx_r, column=2, value=values.get(fid, ""))
        vc.fill      = NO_FILL
        vc.font      = font(sz=9)
        vc.alignment = align(h="left", v="center")
        vc.border    = border(R=M, B=TH)

    # ── SECTION DEFINITIONS ───────────────────────────────────────────────────
    # (opx_row_start, opx_row_end, section_name, col_A_border)
    _sections = [
        (9,  12, "GENERAL",             border(L=M, R=M, T=M, B=M)),
        (13, 17, "PROCESS\nCONDITIONS",border(L=M, R=M, T=M, B=M)),
        (18, 25, "TANK\nDETAILS",       border(L=M, R=M, T=M, B=M)),
        (26, 38, "SENSOR",              border(L=M, R=M, T=M, B=M)),
        (39, 48, "TRANSMITTER",         border(L=M, R=M, T=M, B=M)),
        (49, 50, "OPTIONS",             border(L=M, R=M, T=M, B=M)),
        (51, 51, "CERTIFICATION",       border(L=M, R=M, T=M, B=M)),
        (52, 53, "PURCHASE",            border(L=M, R=M, T=M, B=M)),
    ]
    for r_start, r_end, name, brd in _sections:
        sc = ws.cell(row=r_start, column=1, value=name)
        sc.fill      = NO_FILL
        sc.font      = font(sz=9, bold=True)
        sc.alignment = align(h="center", v="center", wrap=True)
        sc.border    = brd

    # ── DATA ROWS: static template labels ─────────────────────────────────────
    _labels = [
        # (opx_row, label_text)
        (9,  "Tag No"),
        (10, "P&ID No"),
        (11, "Location"),
        (12, "Area Classification"),
        (13, "Fluid"),
        (14, "Specific Gravity"),
        (15, "Viscosity(cP)"),
        (16, "Pressure\n(Kg/Cm² g)"),
        (17, "Temp(°C)"),
        (18, "Tag no"),
        (19, "Type"),
        (20, "MOC"),
        (21, "Height"),
        (22, "Diameter"),
        (23, "Presence of Agitation"),
        (24, "Presence of Fumes"),
        (25, "Max.Fluid Level"),
        (26, "Type"),
        (27, "Frequency Range"),
        (28, "Beam Angle"),
        (29, "Blocking Distance"),
        (30, "Housing MOC"),
        (31, "Enclosure Class"),
        (32, "Temperature Compensation"),
        (33, "Process Connection MOC"),
        (34, "Process Connection Flange Size"),
        (35, "Nozzle Size"),
        (36, "Nozzle Height"),
        (37, "Mounting Position"),
        (38, "Cable distance from Sensor to Transmitter"),
        (39, "Type"),          # Transmitter type header
        (40, "Power supply"),
        (41, "Output"),
        (42, "Accuracy"),
        (43, "Repeatability"),
        (44, "Instrumentation Range"),
        (45, "Calibration Range"),
        (46, "Enclosure Protection Class"),
        (47, "Cable Entry"),
        (48, "Display"),
        (49, "Mounting Accessories"),
        (50, "SS Tag Plate"),
        (51, "Area Certification"),
        (52, "Make"),
        (53, "Model"),
        (54, "NOTES : "),
    ]
    def _label(opx_r, col, val, ha="left", wrap=False, L=None, R=None, T=None, B=None):
        c = ws.cell(row=opx_r, column=col, value=val)
        c.fill      = NO_FILL
        c.font      = font(sz=9)
        c.alignment = align(h=ha, v="center", wrap=wrap)
        c.border    = border(L=L or TH, R=R or TH, T=T or TH, B=B or TH)

    for opx_r, lbl in _labels:
        if opx_r in (16, 17):
            # Split row: B=main label (individual), C=Min, D=Nor, E=Max (individual)
            _label(opx_r, 2, lbl,   ha="left",   wrap=True,  L=M)
            _label(opx_r, 3, "Min", ha="center")
            _label(opx_r, 4, "Nor", ha="center")
            _label(opx_r, 5, "Max", ha="center", R=M)
        else:
            # Normal row: col 2 is top-left of merged B-E
            lc = ws.cell(row=opx_r, column=2, value=lbl)
            lc.fill      = NO_FILL
            lc.font      = font(sz=9)
            lc.alignment = align(h="left", v="center", wrap=(opx_r == 38))
            lc.border    = border(L=M, R=M, T=TH, B=TH)

    # ── WRITE USER VALUES ──────────────────────────────────────────────────────
    def _put(opx_row, opx_col, val):
        c = ws.cell(row=opx_row, column=opx_col, value=val if val is not None else "")
        c.fill      = NO_FILL
        c.font      = font(sz=9)
        c.alignment = align(h="left", v="center")
        c.border    = border(L=TH, R=M, T=TH, B=TH)

    v = values  # shorthand

    # GENERAL
    _put(9,  6, v.get("tag_no",   ""))
    _put(10, 6, v.get("pid_no",   ""))
    _put(11, 6, v.get("location", ""))
    _put(12, 6, v.get("area_class", ""))
    # PROCESS CONDITIONS
    _put(13, 6, v.get("fluid", ""))
    _put(14, 6, v.get("specific_gravity", ""))
    _put(15, 6, v.get("viscosity", ""))
    # Pressure (split)
    _put(16, 6, v.get("pressure_min", ""))
    _put(16, 8, v.get("pressure_nor", ""))
    _put(16, 10, v.get("pressure_max", ""))
    # Temp (split)
    _put(17, 6, v.get("temp_min", ""))
    _put(17, 8, v.get("temp_nor", ""))
    _put(17, 10, v.get("temp_max", ""))
    # TANK DETAILS
    _put(18, 6, v.get("tank_tag", ""))
    _put(19, 6, v.get("tank_type", ""))
    _put(20, 6, v.get("tank_moc", ""))
    _put(21, 6, v.get("tank_height", ""))
    _put(22, 6, v.get("tank_diameter", ""))
    _put(23, 6, v.get("agitation", ""))
    _put(24, 6, v.get("fumes", ""))
    _put(25, 6, v.get("max_fluid", ""))
    # SENSOR
    _put(26, 6, v.get("sensor_type", "Non Contact Radar"))
    _put(27, 6, v.get("frequency", "80 GHz"))
    _put(28, 6, v.get("beam_angle", ""))
    _put(29, 6, v.get("blocking_dist", ""))
    _put(30, 6, v.get("housing_moc", "Die Cast Aluminium"))
    _put(31, 6, v.get("enclosure_class", ""))
    _put(32, 6, v.get("temp_comp", "Built-In"))
    _put(33, 6, v.get("proc_conn_moc", ""))
    _put(34, 6, v.get("proc_conn_flange", ""))
    _put(35, 6, v.get("nozzle_size", ""))
    _put(36, 6, v.get("nozzle_height", ""))
    _put(37, 6, v.get("mounting_pos", ""))
    _put(38, 6, v.get("cable_distance", ""))
    # TRANSMITTER header row 39 → fixed sensor type label
    _put(39, 6, "Non Contact radar")
    # TRANSMITTER specs (rows 40–48, matching template exactly)
    _put(40, 6, v.get("power_supply",   "24 VDC"))
    _put(41, 6, v.get("output",         "4-20mA + HART"))
    _put(42, 6, v.get("accuracy",       "±2 mm"))
    _put(43, 6, v.get("repeatability",  "±1 mm"))
    _put(44, 6, v.get("inst_range",     ""))
    _put(45, 6, v.get("calib_range",    ""))
    _put(46, 6, v.get("encl_protection","IP67"))
    _put(47, 6, v.get("cable_entry",    "M20 x 1.5"))
    _put(48, 6, v.get("display",        "LCD"))
    # OPTIONS (rows 49–50)
    _put(49, 6, v.get("mounting_acc",   ""))
    _put(50, 6, "PROVIDED")             # SS Tag Plate — fixed
    # CERTIFICATION (row 51)
    _put(51, 6, v.get("area_cert",      ""))
    # PURCHASE (rows 52–53)
    _put(52, 6, v.get("make",           ""))
    _put(53, 6, v.get("model_no",       ""))
    # NOTES row 54 handled below (merged B54:K54 → write to col 2)
    # NOTES (in col B of row 54 merged B:K)
    notes_cell = ws.cell(row=54, column=2, value=v.get("notes", ""))
    notes_cell.fill      = NO_FILL
    notes_cell.font      = font(sz=9)
    notes_cell.alignment = align(h="left", v="top", wrap=True)
    notes_cell.border    = border(L=M, R=M, T=TH, B=M)

    # ── APPLY BORDERS TO ANY REMAINING EMPTY VALUE CELLS ──────────────────────
    # (ensure all value cells have a border even if empty)
    for opx_r in range(9, 55):
        if opx_r == 54:  # notes row handled separately
            continue
        for opx_c in [6, 8, 10] if opx_r in (16, 17) else [6]:
            cell = ws.cell(row=opx_r, column=opx_c)
            if cell.border == Border():  # no border set yet
                cell.border = border(L=TH, R=M, T=TH, B=TH)
            cell.fill = NO_FILL  # always clear fill

    # Ensure all cells that might have been left with fills are cleared
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = NO_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 4. VALIDATION RULES
# ---------------------------------------------------------------------------

def validate_specs(values: dict[str, str]) -> list[str]:
    """Run rule engine; return list of error strings (empty = pass)."""
    errors: list[str] = []
    freq    = values.get("frequency", "")
    output  = values.get("output", "")
    area    = values.get("area_class", "")
    cert    = values.get("area_cert", "")
    tx_type = values.get("transmitter_type", "Integral")
    cable_d = values.get("cable_distance", "")
    sil     = values.get("sil_cert", "Not Required")
    encl    = values.get("encl_protection", "")
    housing = values.get("housing_moc", "")
    beam    = values.get("beam_angle", "")

    # Rule 1: Output vs HART revision
    if values.get("hart_revision") and "HART" not in output:
        errors.append("HART Revision cannot be selected because Output Signal is not HART based.")

    # Rule 2: Frequency vs Beam Angle
    if beam and freq:
        if "80" in freq and beam not in DROPDOWN_OPTIONS["beam_angle_80"]:
            errors.append(f"Selected Beam Angle '{beam}' is not valid for 80 GHz. Allowed: {', '.join(DROPDOWN_OPTIONS['beam_angle_80'])}")
        elif "26" in freq and beam not in DROPDOWN_OPTIONS["beam_angle_26"]:
            errors.append(f"Selected Beam Angle '{beam}' is not valid for 26 GHz. Allowed: {', '.join(DROPDOWN_OPTIONS['beam_angle_26'])}")

    # Rule 5: Hazardous area classification must have proper cert
    if area and "general purpose" in area.lower() and cert and cert != "General Purpose":
        pass  # General Purpose area can have any cert
    if cert and cert == "General Purpose" and area and any(
        kw in area.lower() for kw in ["hazardous", "zone", "ex", "class i", "class ii"]
    ):
        errors.append("Hazardous Area Instrument Certification Missing.")

    # Rule 12: Integral transmitter cable distance
    if tx_type == "Integral" and cable_d and cable_d.strip() not in ("", "0", "N/A", "NA"):
        errors.append("Integral Transmitter cannot have Remote Cable Length.")

    # Rule 11: Outdoor vs enclosure rating
    if encl and encl in ("IP65",):
        if area and "outdoor" in area.lower():
            errors.append("Warning: Selected Enclosure Rating (IP65) may not be suitable for Outdoor Installation.")

    return errors
