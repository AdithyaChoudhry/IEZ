"""
Vendor Analysis & Technical Bid Evaluation (TBE) router.
"""
import io
import os
import uuid
import logging
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Optional

from ..deps import get_db, get_current_user
from ..auth.models import User, AdminUser, TBEApprovalLog
from ..auth.utils import verify_password
from ..data.tbe_vendors import VENDOR_DB, TYPE_ALIASES

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/tbe", tags=["TBE"])

TBE_TMP = Path("/tmp/iez_tbe")
TBE_TMP.mkdir(parents=True, exist_ok=True)

SKIP_SHEETS = {"cover sheet", "index", "checklist", "spare", "revision"}
ANNEXURE_KEYWORDS = {"annexure", "annex"}

# Junk values that appear in empty template cells
JUNK_VALUES = {"xxxxxx", "0.0", "0", "n/a", "", "-", "tbd", "tbc",
               "min", "nor", "max", "min.", "nor.", "max."}

# ── helpers ───────────────────────────────────────────────────────────────────

def _fuzzy(a: str, b: str) -> int:
    from rapidfuzz import fuzz
    return fuzz.partial_ratio(a.lower(), b.lower())


def _load_workbook_sheets(content: bytes, filename: str) -> dict:
    """
    Load all sheets from .xls, .xlsx, or .xlsm.
    Returns {sheet_name: [[row_cells], ...]} where each cell is a string.
    """
    fname = (filename or "").lower()

    if fname.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        result = {}
        for sname in wb.sheet_names():
            ws = wb.sheet_by_name(sname)
            rows = []
            for r in range(ws.nrows):
                row = [str(ws.cell(r, c).value).strip() for c in range(ws.ncols)]
                rows.append(row)
            result[sname] = rows
        return result
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        result = {}
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c).strip() if c is not None else "" for c in row])
            result[sname] = rows
        return result


HEADER_SKIP = {"vatech", "wabag", "chennai", "murray", "alwarpet", "no. 11",
               "instrument datasheet", "issued for", "enquiry", "document"}
SECTION_KEYWORDS = {
    "general", "process", "sensor", "transmitter", "certification",
    "purchase", "options", "mechanical", "electrical", "safety",
    "tank", "notes", "upper fluid", "lower fluid", "operating",
    "loop powered", "external cage",
}
SKIP_PARAM_VALUES = {"parameter", "specification", "description",
                     "sl.no", "s.no", "#", "item", ":", "/", "-"}


def _meaningful_cells(row: list) -> list:
    """Return [(col_idx, value)] for non-empty, non-junk, non-pure-number cells."""
    out = []
    for ci, cell in enumerate(row):
        v = str(cell).strip() if cell else ""
        if not v or v.lower() in JUNK_VALUES:
            continue
        if v in (":", "/", "-", "%"):
            continue
        out.append((ci, v))
    return out


def _looks_like_section(text: str) -> bool:
    """True only for genuine section headers — NOT for params that happen to contain a keyword."""
    tl = text.strip().lower()
    # multi-line cells at col 0 are always section/header
    if "\n" in text and len(text) < 60:
        return True
    # company/doc header lines
    if any(kw in tl for kw in HEADER_SKIP):
        return True
    # section keywords that appear as the ENTIRE cell (not embedded in a longer phrase)
    # e.g. "General Data" → yes; "Operating Pressure" → no
    words = set(tl.replace("\n", " ").split())
    if any(kw in words for kw in SECTION_KEYWORDS):
        # only treat as section if it's short (≤ 4 words) and NOT a known compound param
        if len(words) <= 4 and not any(skip in tl for skip in ("pressure", "temperature", "protection", "classification", "current", "supply")):
            return True
    return False


def _detect_split_col(rows: list) -> int:
    """
    Find the column index that divides params (left) from values (right).
    Uses the most common column position of "Refer Annexure" and other values.
    Falls back to the median column of all non-empty content.
    """
    max_col = max((len(r) for r in rows if r), default=5)
    if max_col <= 5:
        return 1  # simple 2-col sheet — split after col 1

    # Vote: cells that are clearly values (annexure refs, known value patterns)
    value_col_votes: dict = {}
    all_content_cols: list = []

    for row in rows:
        cells = _meaningful_cells(row)
        for ci, v in cells:
            all_content_cols.append(ci)
            vl = v.lower()
            if ("refer annex" in vl or "annexure" in vl
                    or vl in ("na", "nil", "yes", "no", "liquid", "gas", "atex")
                    or any(unit in vl for unit in ("mA", "VDC", "VAC", "bar", "°C", "mm", "IP ", "SIL", "HART", "NPT"))):
                value_col_votes[ci] = value_col_votes.get(ci, 0) + 1

    if value_col_votes:
        # the most-voted column is the primary value column
        primary_val_col = max(value_col_votes, key=value_col_votes.get)
        # split is one before the earliest value column that has significant votes
        earliest_val = min(c for c, cnt in value_col_votes.items() if cnt >= 2) if any(cnt >= 2 for cnt in value_col_votes.values()) else primary_val_col
        return max(0, earliest_val - 1)

    # fallback: split at 60% of max_col
    if all_content_cols:
        sorted_cols = sorted(all_content_cols)
        split = sorted_cols[len(sorted_cols) * 6 // 10]
        return split

    return max_col // 2


def _extract_params(rows: list, sheet_name: str) -> list:
    """
    Universal extractor — handles any WABAG datasheet layout:
    - SOP template: col0=section, col1=param, col2=value
    - iEZ generated: col0=section, col5=num, col6=param, col9=param2, col13=val, col16=val2
    - Simple 2-col: col0=param, col1=value
    Auto-detects the param/value split column.
    """
    split = _detect_split_col(rows)
    current_section = ""
    results = []
    seen: set = set()

    in_header = True  # skip project header rows (PROJECT/CLIENT/etc.) at top

    for row in rows:
        if not row or all(not c for c in row):
            continue

        cells = _meaningful_cells(row)
        if not cells:
            continue

        # Update section from col-0 cell if it looks like a section header
        if cells[0][0] == 0 and _looks_like_section(cells[0][1]):
            new_section = cells[0][1].replace("\n", " ").strip()[:50]
            # once we hit a real instrument section, we're past the header
            if any(kw in new_section.lower() for kw in SECTION_KEYWORDS) and not any(kw in new_section.lower() for kw in HEADER_SKIP):
                in_header = False
            current_section = new_section
            cells = cells[1:]

        # Skip project/doc header rows at the top of the sheet
        if in_header:
            continue

        # Separate into left (params) and right (values)
        left = [(ci, v) for ci, v in cells if ci <= split and not v.replace(".", "").isdigit()]
        right = [(ci, v) for ci, v in cells if ci > split]

        if not left:
            continue

        # Pair by index: left[0]→right[0], left[1]→right[1], etc.
        for i, (pci, param) in enumerate(left):
            param = param.strip()
            if not param or len(param) > 80:
                continue
            if param.lower() in SKIP_PARAM_VALUES:
                continue
            # only skip as section if it's the ONLY thing in left (a stray header)
            if _looks_like_section(param) and len(left) == 1:
                current_section = param
                continue

            value = right[i][1] if i < len(right) else ""
            if value.lower() in JUNK_VALUES:
                value = ""

            key = (current_section.lower()[:20], param.lower()[:40])
            if key in seen:
                continue
            seen.add(key)

            results.append({
                "param": param,
                "value": value,
                "section": current_section,
                "source": sheet_name,
                "resolved": False,
            })

    return results


def _detect_instrument_type(specs: list[dict]) -> str:
    for row in specs:
        param = (row.get("param") or "").lower()
        value = (row.get("value") or "").lower()
        if any(k in param for k in ("instrument type", "instrument_type", "type of instrument", "measurement type")):
            for alias, canonical in TYPE_ALIASES.items():
                if alias in value:
                    return canonical
        # Fallback: check service / tag
        for alias, canonical in TYPE_ALIASES.items():
            if alias in value:
                return canonical
    return "non-contact radar level transmitter"  # safest default


def _match_vendor_db_key(instrument_type: str) -> Optional[str]:
    it = instrument_type.lower()
    if it in VENDOR_DB:
        return it
    from rapidfuzz import fuzz
    best, best_score = None, 0
    for key in VENDOR_DB:
        s = fuzz.partial_ratio(it, key)
        if s > best_score:
            best_score, best = s, key
    return best if best_score >= 60 else list(VENDOR_DB.keys())[0]


NOT_ACCEPTABLE_PHRASES = (
    "not available", "not supported", "not applicable for this model",
    "cannot", "not offered", "not provided", "does not support", "n/o",
)


def _evaluate_compliance(wabag: str, vendor: str) -> str:
    w, v = wabag.strip().lower(), vendor.strip().lower()
    if not v or v in ("n/a", "na", "tbd", "tbc", "-", ""):
        return "CLARIFICATION REQUIRED"
    # vendor explicitly states not available
    if any(phrase in v for phrase in NOT_ACCEPTABLE_PHRASES):
        return "NOT ACCEPTABLE"
    # exact / strong match
    if _fuzzy(wabag, vendor) >= 80:
        return "COMPLIES"
    # check numeric comparison for ranges / values
    try:
        import re
        wn = float(re.search(r"[\d.]+", w).group())
        vn = float(re.search(r"[\d.]+", v).group())
        # For accuracy/tolerance smaller is better
        if any(k in w for k in ("accuracy", "tolerance", "error", "mm", "±")):
            if vn <= wn:
                return "COMPLIES"
            elif vn <= wn * 1.5:
                return "DEVIATION"
            else:
                return "NOT ACCEPTABLE"
        return "EXCEEDS REQUIREMENT" if vn >= wn else "DEVIATION"
    except Exception:
        pass
    # substring check
    if w in v or v in w:
        return "COMPLIES"
    return "DEVIATION"


def _auto_reply(status: str) -> str:
    mapping = {
        "COMPLIES": "COMPLIES",
        "EXCEEDS REQUIREMENT": "EXCEEDS REQUIREMENT — TECHNICALLY ACCEPTABLE",
        "DEVIATION": "DEVIATION — CLARIFICATION / WAIVER REQUIRED",
        "CLARIFICATION REQUIRED": "CLARIFICATION REQUIRED",
        "NOT ACCEPTABLE": "NOT ACCEPTABLE — VENDOR TO PROPOSE ALTERNATIVE",
        "TECHNICALLY ACCEPTABLE": "TECHNICALLY ACCEPTABLE",
    }
    return mapping.get(status, status)


# ── /analyze ──────────────────────────────────────────────────────────────────

@router.post("/analyze")
async def analyze_datasheet(
    file: UploadFile = File(...),
    _: User = Depends(get_current_user),
):
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(400, "Only .xlsx, .xlsm, or .xls files are supported")

    try:
        content = await file.read()
        all_sheets = _load_workbook_sheets(content, file.filename or "")
    except Exception as e:
        raise HTTPException(400, f"Cannot open workbook: {e}")

    all_names = list(all_sheets.keys())
    sheet_names = [s for s in all_names
                   if not any(k in s.lower() for k in SKIP_SHEETS)]
    annexure_sheets = [s for s in all_names
                       if any(k in s.lower() for k in ANNEXURE_KEYWORDS)]

    # ── parse annexures ──
    annexures: dict = {}
    for aname in annexure_sheets:
        rows = all_sheets[aname]
        if not rows:
            continue
        # find header row (first row with ≥2 non-empty cells)
        header_row_idx, headers = 0, []
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c]
            if len(non_empty) >= 2:
                headers = row
                header_row_idx = i
                break
        if not headers:
            continue

        # find tag column
        tag_col_idx = 0
        for i, h in enumerate(headers):
            if _fuzzy("tag", h) >= 70 or _fuzzy("tag no", h) >= 70:
                tag_col_idx = i
                break

        rows_parsed = []
        for row in rows[header_row_idx + 1:]:
            if not any(c for c in row):
                continue
            tag_val = row[tag_col_idx] if tag_col_idx < len(row) else ""
            row_dict = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
            rows_parsed.append({"tag": tag_val, "cols": row_dict})

        annexures[aname] = {"headers": headers, "rows": rows_parsed}

    # ── parse instrument sheets ──
    specs_raw = []
    tag_numbers = []

    for sname in sheet_names:
        if any(k in sname.lower() for k in ANNEXURE_KEYWORDS):
            continue
        tag_numbers.append(sname)
        rows = all_sheets[sname]

        raw_params = _extract_params(rows, sname)

        for entry in raw_params:
            param = entry["param"]
            value = entry["value"]
            resolved = False
            source = sname

            # resolve annexure references
            if value and ("annexure" in value.lower() or "refer annex" in value.lower()):
                for aname, ann_data in annexures.items():
                    # find row whose tag matches this sheet name
                    matched_row = None
                    for r in ann_data["rows"]:
                        if _fuzzy(sname, r["tag"]) >= 75:
                            matched_row = r
                            break
                    if matched_row:
                        best_col, best_score = None, 0
                        for col_h in matched_row["cols"]:
                            s = _fuzzy(param, col_h)
                            if s > best_score:
                                best_score, best_col = s, col_h
                        if best_col and best_score >= 55 and matched_row["cols"].get(best_col):
                            value = matched_row["cols"][best_col]
                            source = aname
                            resolved = True
                            break

            specs_raw.append({
                "param": param,
                "value": value,
                "section": entry.get("section", ""),
                "source": source,
                "resolved": resolved,
            })

    instrument_type = _detect_instrument_type(specs_raw)

    # fallback: try filename for instrument type
    if instrument_type == "non-contact radar level transmitter":
        fn_lower = (file.filename or "").lower()
        for alias, canonical in TYPE_ALIASES.items():
            if alias in fn_lower:
                instrument_type = canonical
                break

    return {
        "instrument_type": instrument_type,
        "tag_numbers": tag_numbers,
        "specs": specs_raw,
        "sheet_names": sheet_names,
        "annexure_sheets": annexure_sheets,
    }


# ── /match ─────────────────────────────────────────────────────────────────────

class MatchRequest(BaseModel):
    instrument_type: str
    specs: list


@router.post("/match")
def match_vendors(body: MatchRequest, _: User = Depends(get_current_user)):
    db_key = _match_vendor_db_key(body.instrument_type)
    vendor_models = VENDOR_DB.get(db_key, [])

    results = []
    for vm in vendor_models:
        spec_results = []
        for row in body.specs:
            param = row.get("param", "")
            wabag_val = row.get("value", "")
            if not param or not wabag_val:
                continue

            # fuzzy-find best matching vendor spec key
            best_key, best_score = None, 0
            for vk in vm["specs"]:
                s = _fuzzy(param, vk)
                if s > best_score:
                    best_score, best_key = s, vk

            vendor_offer = vm["specs"].get(best_key, "") if best_key and best_score >= 60 else ""
            status = _evaluate_compliance(wabag_val, vendor_offer)

            spec_results.append({
                "param": param,
                "wabag_req": wabag_val,
                "vendor_offer": vendor_offer,
                "status": status,
                "auto_reply": _auto_reply(status),
            })

        if not spec_results:
            continue

        complies = sum(1 for r in spec_results if r["status"] in ("COMPLIES", "EXCEEDS REQUIREMENT"))
        match_pct = round(complies / len(spec_results) * 100) if spec_results else 0

        results.append({
            "vendor": vm["vendor"],
            "abbr": vm["abbr"],
            "model": vm["model"],
            "match_pct": match_pct,
            "spec_results": spec_results,
        })

    results.sort(key=lambda x: -x["match_pct"])
    return {"vendors": results[:5], "instrument_type": db_key}


# ── /generate ─────────────────────────────────────────────────────────────────

class GenerateRequest(BaseModel):
    instrument_type: str
    vendors: list                   # vendor result objects from /match
    tbe_replies: dict = {}          # {abbr: {param: reply}}
    deviation_severities: dict = {} # {abbr: {param: "Critical"|"Major"|"Minor"}}
    recommended_vendor: str = ""
    recommended_model: str = ""
    recommendation_reason: str = ""


@router.post("/generate")
def generate_tbe(body: GenerateRequest, _: User = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    session_id = str(uuid.uuid4())
    out_dir = TBE_TMP / session_id
    out_dir.mkdir(parents=True, exist_ok=True)

    vendors = body.vendors
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def cell_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    STATUS_COLOR = {
        "COMPLIES": "E8F5E9",
        "EXCEEDS REQUIREMENT": "E3F2FD",
        "DEVIATION": "FFEBEE",
        "CLARIFICATION REQUIRED": "FFF8E1",
        "NOT ACCEPTABLE": "FFCDD2",
        "TECHNICALLY ACCEPTABLE": "E0F7FA",
    }

    def _default_severity(status: str) -> str:
        if status == "NOT ACCEPTABLE":
            return "Critical"
        if status == "DEVIATION":
            return "Major"
        return "Minor"

    # ── TBE Report ──
    wb_tbe = openpyxl.Workbook()
    ws = wb_tbe.active
    ws.title = "TBE Report"

    # header row
    headers = ["Sl.No", "Specification", "WABAG Requirement"]
    for v in vendors:
        headers += [f"{v['abbr']} Offer", f"WABAG Reply ({v['abbr']})"]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill("1E3A5F")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    ws.row_dimensions[1].height = 30

    # collect all params
    all_params = []
    if vendors:
        seen = set()
        for r in vendors[0]["spec_results"]:
            if r["param"] not in seen:
                seen.add(r["param"])
                all_params.append(r["param"])

    # data rows
    deviation_rows = []
    for ri, param in enumerate(all_params, 1):
        row_data = [ri, param]
        wabag_req = ""
        for v in vendors:
            for sr in v["spec_results"]:
                if sr["param"] == param:
                    wabag_req = sr["wabag_req"]
                    break
            if wabag_req:
                break
        row_data.append(wabag_req)

        for v in vendors:
            sr = next((r for r in v["spec_results"] if r["param"] == param), None)
            offer = sr["vendor_offer"] if sr else ""
            status = sr["status"] if sr else "CLARIFICATION REQUIRED"
            reply = (body.tbe_replies.get(v["abbr"], {}).get(param)
                     or (sr["auto_reply"] if sr else _auto_reply(status)))

            c_offer = ws.cell(row=ri + 1, column=len(row_data) + 1, value=offer)
            c_offer.fill = cell_fill(STATUS_COLOR.get(status, "FFFFFF"))
            c_offer.font = Font(size=9)
            c_offer.border = border
            c_offer.alignment = Alignment(wrap_text=True)

            c_reply = ws.cell(row=ri + 1, column=len(row_data) + 2, value=reply)
            c_reply.fill = cell_fill("FFF9C4")
            c_reply.font = Font(size=9)
            c_reply.border = border
            c_reply.alignment = Alignment(wrap_text=True)
            row_data += [offer, reply]

            if status not in ("COMPLIES",):
                sev = (body.deviation_severities.get(v["abbr"], {}).get(param)
                       or _default_severity(status))
                deviation_rows.append({
                    "param": param,
                    "wabag_req": wabag_req,
                    "vendor": v["vendor"],
                    "model": v["model"],
                    "offer": offer,
                    "status": status,
                    "severity": sev,
                    "abbr": v["abbr"],
                })

        for ci, val in enumerate([ri, param, wabag_req], 1):
            c = ws.cell(row=ri + 1, column=ci, value=val)
            c.font = Font(size=9)
            c.border = border
            c.alignment = Alignment(wrap_text=True)

    # column widths
    for col in ws.columns:
        max_w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 2, 40)

    tbe_path = out_dir / "TBE_Report.xlsx"
    wb_tbe.save(tbe_path)

    # ── Deviation Report ──
    wb_dev = openpyxl.Workbook()
    ws_sum = wb_dev.active
    ws_sum.title = "Summary"
    dev_headers = ["Sl.No", "Specification", "WABAG Requirement",
                   "Vendor", "Model", "Vendor Offer", "Status", "Severity", "WABAG Comment"]
    for ci, h in enumerate(dev_headers, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill("B71C1C")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.border = border

    for ri, dr in enumerate(deviation_rows, 1):
        vals = [ri, dr["param"], dr["wabag_req"], dr["vendor"], dr["model"],
                dr["offer"], dr["status"], dr["severity"], ""]
        for ci, v in enumerate(vals, 1):
            c = ws_sum.cell(row=ri + 1, column=ci, value=v)
            c.fill = cell_fill(STATUS_COLOR.get(dr["status"], "FFFFFF"))
            c.font = Font(size=9)
            c.border = border

    # per-vendor sheets
    for v in vendors:
        vrows = [r for r in deviation_rows if r["abbr"] == v["abbr"]]
        if not vrows:
            continue
        ws_v = wb_dev.create_sheet(title=v["abbr"][:31])
        for ci, h in enumerate(dev_headers, 1):
            c = ws_v.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill("1E3A5F")
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.border = border
        for ri, dr in enumerate(vrows, 1):
            vals = [ri, dr["param"], dr["wabag_req"], dr["vendor"], dr["model"],
                    dr["offer"], dr["status"], dr["severity"], ""]
            for ci, val in enumerate(vals, 1):
                c = ws_v.cell(row=ri + 1, column=ci, value=val)
                c.fill = cell_fill(STATUS_COLOR.get(dr["status"], "FFFFFF"))
                c.font = Font(size=9)
                c.border = border

    dev_path = out_dir / "Deviation_Report.xlsx"
    wb_dev.save(dev_path)

    # ── Compliance Summary ──
    wb_cs = openpyxl.Workbook()
    ws_cs = wb_cs.active
    ws_cs.title = "Compliance Summary"
    cs_headers = ["Vendor", "Model", "Compliance %", "Complies", "Exceeds", "Deviations", "Clarifications"]
    for ci, h in enumerate(cs_headers, 1):
        c = ws_cs.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill("1B5E20")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.border = border

    for ri, v in enumerate(vendors, 1):
        sr = v["spec_results"]
        complies = sum(1 for r in sr if r["status"] == "COMPLIES")
        exceeds = sum(1 for r in sr if r["status"] == "EXCEEDS REQUIREMENT")
        devs = sum(1 for r in sr if r["status"] == "DEVIATION")
        clarif = sum(1 for r in sr if r["status"] == "CLARIFICATION REQUIRED")
        not_acc = sum(1 for r in sr if r["status"] == "NOT ACCEPTABLE")
        vals = [v["vendor"], v["model"], f"{v['match_pct']}%", complies, exceeds, devs, clarif, not_acc]
        for ci, val in enumerate(vals, 1):
            c = ws_cs.cell(row=ri + 1, column=ci, value=val)
            c.font = Font(size=9)
            c.border = border

    # Update header to include Not Acceptable column
    ws_cs.cell(row=1, column=8, value="Not Acceptable").fill = hdr_fill("1B5E20")
    ws_cs.cell(row=1, column=8).font = Font(bold=True, color="FFFFFF", size=10)
    ws_cs.cell(row=1, column=8).border = border

    # Recommendation sheet
    ws_rec = wb_cs.create_sheet("Recommendation")
    rec_data = [
        ("Recommended Vendor", body.recommended_vendor or (vendors[0]["vendor"] if vendors else "")),
        ("Recommended Model", body.recommended_model or (vendors[0]["model"] if vendors else "")),
        ("Compliance Score", f"{vendors[0]['match_pct']}%" if vendors else ""),
        ("Instrument Type", body.instrument_type),
        ("Reason", body.recommendation_reason or (
            f"Highest compliance score ({vendors[0]['match_pct']}%) among evaluated vendors"
            + (" with no critical deviations." if not any(
                r["status"] == "NOT ACCEPTABLE" for r in vendors[0].get("spec_results", [])
            ) else ".")
            if vendors else ""
        )),
        ("Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for ri, (k, v_val) in enumerate(rec_data, 1):
        c_k = ws_rec.cell(row=ri, column=1, value=k)
        c_k.font = Font(bold=True, size=10)
        c_k.fill = hdr_fill("1E3A5F")
        c_k.font = Font(bold=True, color="FFFFFF", size=10)
        c_k.border = border
        c_v = ws_rec.cell(row=ri, column=2, value=v_val)
        c_v.font = Font(size=10)
        c_v.border = border
        c_v.alignment = Alignment(wrap_text=True)
    ws_rec.column_dimensions["A"].width = 25
    ws_rec.column_dimensions["B"].width = 60

    cs_path = out_dir / "Compliance_Summary.xlsx"
    wb_cs.save(cs_path)

    return {
        "session_id": session_id,
        "tbe_rows": len(all_params),
        "deviation_rows": len(deviation_rows),
    }


# ── /approve ───────────────────────────────────────────────────────────────────

class ApproveRequest(BaseModel):
    session_id: str
    employee_id: str
    password: str
    instrument_type: str
    recommended_vendor: str
    recommended_model: str


@router.post("/approve")
def approve_tbe(body: ApproveRequest, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    admin = db.query(AdminUser).filter(
        AdminUser.employee_id == body.employee_id,
        AdminUser.status == "active",
    ).first()
    if not admin or not verify_password(body.password, admin.password_hash):
        raise HTTPException(401, "Invalid employee ID or password")

    tbe_number = f"TBE-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:4].upper()}"
    log = TBEApprovalLog(
        tbe_number=tbe_number,
        instrument_type=body.instrument_type,
        approved_by=admin.employee_name,
        employee_id=admin.employee_id,
        approval_date=datetime.now(timezone.utc),
        recommended_vendor=body.recommended_vendor,
        recommended_model=body.recommended_model,
        session_id=body.session_id,
    )
    db.add(log)
    db.commit()

    return {
        "status": "approved",
        "tbe_number": tbe_number,
        "approved_by": admin.employee_name,
        "employee_id": admin.employee_id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ── /download ──────────────────────────────────────────────────────────────────

FILE_MAP = {
    "tbe": "TBE_Report.xlsx",
    "deviation": "Deviation_Report.xlsx",
    "compliance": "Compliance_Summary.xlsx",
}


@router.get("/download/{session_id}/{file_type}")
def download_file(session_id: str, file_type: str, _: User = Depends(get_current_user)):
    if file_type not in FILE_MAP:
        raise HTTPException(400, "file_type must be tbe | deviation | compliance")
    # sanitize session_id to prevent path traversal
    if "/" in session_id or ".." in session_id:
        raise HTTPException(400, "Invalid session_id")
    path = TBE_TMP / session_id / FILE_MAP[file_type]
    if not path.exists():
        raise HTTPException(404, "File not found — please generate the TBE first")
    return FileResponse(path, filename=FILE_MAP[file_type],
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
