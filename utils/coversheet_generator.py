"""
coversheet_generator.py
Generates a Document Control / Revision Control cover sheet from the
LT_DS_coversheet.xlsx reference template, with an interactive preview
(layout + image positions) and final-generation step that applies user
edits (cell formatting overrides, image placements).
"""

import base64
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from utils.file_handler import workbook_to_bytes
from utils.coversheet_layout import (
    get_col_widths_px,
    get_row_heights_px,
    build_layout,
    extract_existing_images,
    place_image,
    apply_cell_overrides,
    _cumulative,
)

_HERE = os.path.dirname(os.path.abspath(__file__))

# The template lives at backend/app/templates/coversheet_template.xlsx.
# Inside the backend Docker container, /app/utils is a symlink to the
# host utils/ directory, so ".." from there does not resolve back to /app -
# use the known absolute container path first, then fall back to the local
# dev layout where utils/ and backend/ are siblings.
_CANDIDATE_TEMPLATE_PATHS = [
    "/app/app/templates/coversheet_template.xlsx",
    os.path.join(_HERE, "..", "backend", "app", "templates", "coversheet_template.xlsx"),
]

TEMPLATE_PATH = next(
    (p for p in _CANDIDATE_TEMPLATE_PATHS if os.path.exists(p)),
    _CANDIDATE_TEMPLATE_PATHS[0],
)

# Revision table row layout (oldest revision first).
# Each entry maps a revision slot to the row holding REV/DATE/DESCRIPTION/
# PREPARED BY/CHECKED BY/APPROVED BY for that slot.
_REVISION_ROWS = [22, 20, 19, 18, 17, 16, 15]
MAX_REVISIONS = len(_REVISION_ROWS)

# Column indices (0-based) for the prepared/checked/approved-by columns
# within a revision row, used for default signature placement.
_SIGNATURE_COLS = {
    "prepared_signature": 8,   # column I
    "checked_signature": 11,   # column L
    "approved_signature": 14,  # column O
}
_SIGNATURE_SIZE = (100, 35)

_PMC_LOGO_DEFAULT_SIZE = (150, 80)
_PMC_LOGO_ROW_IDX = 11  # row 12 (stamping area)

# The template marks input cells with a yellow "fill me in" highlight.
# This is removed from the output so the final sheet looks like a normal
# document (borders, fonts, merges, row/col sizing are left untouched).
_PLACEHOLDER_FILL_RGB = "FFFFFF00"
_NO_FILL = PatternFill(fill_type=None)


def _strip_placeholder_highlights(ws):
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill and fill.fill_type == "solid" and fill.fgColor and fill.fgColor.rgb == _PLACEHOLDER_FILL_RGB:
                cell.fill = _NO_FILL


def _fill_fields(ws, project_info, revisions):
    # --- Title block / ID strip ---
    ws["G2"] = project_info.get("doc_title_short", "")
    ws["A3"] = project_info.get("project_description", "")
    ws["G4"] = project_info.get("job_no", "")
    ws["H4"] = project_info.get("unit_no", "")
    ws["J4"] = project_info.get("project_no", "")
    ws["K4"] = project_info.get("doc_code", "")
    ws["M4"] = project_info.get("serial_no", "")
    ws["N4"] = revisions[-1]["rev_no"]
    ws["P4"] = project_info.get("page_no", "")

    # --- Org block ---
    ws["F6"] = project_info.get("client", "")
    ws["F7"] = project_info.get("pmc", "")
    ws["F8"] = project_info.get("contractor", "")

    # --- Document title ---
    ws["A10"] = "DOCUMENT TITLE\n\n\n\n\n\n" + project_info.get("doc_title_full", "")

    # --- Class / discipline ---
    ws["E13"] = project_info.get("doc_class", "")
    ws["E14"] = project_info.get("discipline", "")

    # --- Revision table (oldest -> _REVISION_ROWS[0], newest -> highest index) ---
    for idx, rev in enumerate(revisions):
        row = _REVISION_ROWS[idx]
        ws[f"A{row}"] = rev.get("rev_no", "")
        ws[f"B{row}"] = rev.get("date", "")
        ws[f"C{row}"] = rev.get("description", "")
        ws[f"I{row}"] = rev.get("prepared_by", "")
        ws[f"L{row}"] = rev.get("checked_by", "")
        ws[f"O{row}"] = rev.get("approved_by", "")


def _default_image_placements(col_widths, row_heights, revisions):
    col_offsets = _cumulative(col_widths)
    row_offsets = _cumulative(row_heights)

    placements = {
        "client_logo": {
            "x": 0, "y": 0,
            "width": sum(col_widths[:6]),
            "height": row_heights[0] + row_heights[1],
        },
        "pmc_logo": {
            "x": 10,
            "y": row_offsets[_PMC_LOGO_ROW_IDX] + 10,
            "width": _PMC_LOGO_DEFAULT_SIZE[0],
            "height": _PMC_LOGO_DEFAULT_SIZE[1],
        },
    }

    rev_row_idx = _REVISION_ROWS[len(revisions) - 1] - 1
    sig_y = row_offsets[rev_row_idx]
    for key, col_idx in _SIGNATURE_COLS.items():
        placements[key] = {
            "x": col_offsets[col_idx],
            "y": sig_y,
            "width": _SIGNATURE_SIZE[0],
            "height": _SIGNATURE_SIZE[1],
        }

    return placements


def build_preview(project_info: dict, revisions: list[dict], uploaded_images: dict | None = None, template_bytes: bytes | None = None):
    """
    Build a JSON-serialisable preview: the filled-in grid layout (cells,
    styles, merges, row/col sizes) plus image overlays (current WABAG logo
    position + default positions for client/PMC logos and signatures).

    uploaded_images: optional dict of {image_key: bytes} for images the
    user has already selected on the form (used to render the overlay
    thumbnail at its default/last-known position).

    Returns (preview_dict, error_message).
    """
    try:
        if not revisions:
            return None, "At least one revision entry is required."
        if len(revisions) > MAX_REVISIONS:
            return None, f"A maximum of {MAX_REVISIONS} revisions is supported."

        if template_bytes:
            import io as _io
            wb = load_workbook(_io.BytesIO(template_bytes))
            ws_name = wb.sheetnames[0]  # use first sheet of custom template
            ws = wb[ws_name]
        else:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb["Coversheet"]

        existing_images = extract_existing_images(ws)

        _fill_fields(ws, project_info, revisions)
        _strip_placeholder_highlights(ws)

        layout = build_layout(ws)
        col_widths, row_heights = layout["cols"], layout["rows"]

        images = {}
        if existing_images:
            images["wabag_logo"] = existing_images[0]

        images.update(_default_image_placements(col_widths, row_heights, revisions))

        uploaded_images = uploaded_images or {}
        for key, file_bytes in uploaded_images.items():
            if not file_bytes:
                continue
            entry = images.setdefault(key, {"x": 0, "y": 0, "width": 100, "height": 50})
            entry["data"] = base64.b64encode(file_bytes).decode("ascii")

        return {"layout": layout, "images": images}, None

    except Exception as e:
        return None, f"Failed to build cover sheet preview: {str(e)}"


def generate_cover_sheet(
    project_info: dict,
    revisions: list[dict],
    images: dict | None = None,
    cell_overrides: dict | None = None,
    template_bytes: bytes | None = None,
    merge_ranges: list[str] | None = None,
    unmerge_coords: list[str] | None = None,
    blank_mode: bool = False,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Fill the cover sheet template with project information, revision
    history, repositioned images, and cell-formatting overrides.

    merge_ranges: list of Excel range strings ("A1:C3") to merge
    unmerge_coords: list of Excel range strings ("A1:C3") to unmerge
    blank_mode: if True, start from a fresh blank workbook instead of template

    Returns (output_bytes, filename, error_message).
    """
    try:
        if not revisions:
            return None, None, "At least one revision entry is required."
        if len(revisions) > MAX_REVISIONS:
            return None, None, f"A maximum of {MAX_REVISIONS} revisions is supported."

        images = images or {}
        cell_overrides = cell_overrides or {}
        merge_ranges = merge_ranges or []
        unmerge_coords = unmerge_coords or []

        if blank_mode:
            from openpyxl import Workbook as _Workbook
            wb = _Workbook()
            ws = wb.active
            ws.title = "Coversheet"
            existing_images = []
            wabag_default = None
        elif template_bytes:
            import io as _io
            wb = load_workbook(_io.BytesIO(template_bytes))
            ws_name = wb.sheetnames[0]
            ws = wb[ws_name]
            existing_images = extract_existing_images(ws)
            wabag_default = existing_images[0] if existing_images else None
        else:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb["Coversheet"]
            existing_images = extract_existing_images(ws)
            wabag_default = existing_images[0] if existing_images else None

        if not blank_mode:
            _fill_fields(ws, project_info, revisions)
            _strip_placeholder_highlights(ws)

        apply_cell_overrides(ws, cell_overrides)

        # Apply unmerges (split template merges the user removed)
        for range_str in unmerge_coords:
            try:
                ws.unmerge_cells(range_str)
            except Exception:
                pass  # ignore if not actually merged

        # Apply new merges (user added via canvas)
        for range_str in merge_ranges:
            try:
                ws.merge_cells(range_str)
            except Exception:
                pass

        col_widths = get_col_widths_px(ws)
        row_heights = get_row_heights_px(ws)

        # Remove embedded images - re-added below at corrected placements
        ws._images = []

        if not blank_mode:
            defaults = _default_image_placements(col_widths, row_heights, revisions)

            # --- Client logo ---
            client_info = images.get("client_logo")
            if client_info and client_info.get("bytes"):
                placement = client_info.get("placement") or defaults["client_logo"]
                place_image(ws, client_info["bytes"], placement, col_widths, row_heights)
                ws["A1"] = ""
            else:
                ws["A1"] = project_info.get("client", "")

            # --- WABAG logo (from template, optionally repositioned/replaced) ---
            wabag_info = images.get("wabag_logo")
            wabag_bytes = (wabag_info or {}).get("bytes")
            if not wabag_bytes and wabag_default:
                wabag_bytes = base64.b64decode(wabag_default["data"])
            if wabag_bytes:
                placement = (wabag_info or {}).get("placement")
                if not placement and wabag_default:
                    placement = {k: wabag_default[k] for k in ("x", "y", "width", "height")}
                if placement:
                    place_image(ws, wabag_bytes, placement, col_widths, row_heights)

            # --- PMC logo (optional) ---
            pmc_info = images.get("pmc_logo")
            if pmc_info and pmc_info.get("bytes"):
                placement = pmc_info.get("placement") or defaults["pmc_logo"]
                place_image(ws, pmc_info["bytes"], placement, col_widths, row_heights)

            # --- Signatures (optional) ---
            for key in _SIGNATURE_COLS:
                info = images.get(key)
                if info and info.get("bytes"):
                    placement = info.get("placement") or defaults[key]
                    place_image(ws, info["bytes"], placement, col_widths, row_heights)

        # --- Custom / extra images (any id not in the known fixed keys) ---
        _known_keys = {"client_logo", "pmc_logo", "wabag_logo",
                       "prepared_signature", "checked_signature", "approved_signature"}
        for img_id, info in images.items():
            if img_id in _known_keys:
                continue
            if info and info.get("bytes") and info.get("placement"):
                place_image(ws, info["bytes"], info["placement"], col_widths, row_heights)

        out_bytes = workbook_to_bytes(wb)
        doc_code = project_info.get("doc_code", "COVER")
        serial_no = project_info.get("serial_no", "")
        filename = f"{doc_code}_{serial_no}_CoverSheet.xlsx".replace(" ", "_")
        return out_bytes, filename, None

    except Exception as e:
        return None, None, f"Failed to generate cover sheet: {str(e)}"
