"""
sop_datasheet.py
================
SOP-driven interactive datasheet generator — Phase 1 (analysis).

A SOP datasheet workbook (e.g. "LT Non contact Radar Type-R0.xls") contains,
per instrument type, a blank template sheet (e.g. "LT-RADAR") and a filled
example sheet (e.g. "LT-RADAR-EG"). This module reads those workbooks
(.xls via xlrd, .xlsx via openpyxl) into a uniform cell model and derives,
for a chosen instrument datasheet, the ordered list of technical-spec fields
the user must fill — classifying each field as either:

  * "iodb"       — value fetched from the uploaded IODB per tag
                   (EG cell empty or "REFER ANNEXURE …")
  * "predefined" — a fixed default carried from the EG example
                   (editable by the user)

It also extracts the distinct instrument types and matching tags from the
uploaded IODB.

Public API (Phase 1)
--------------------
    read_workbook(file_bytes, filename)          -> dict[str, SheetModel]
    list_datasheet_sheets(sheets)                -> list[DatasheetInfo]
    parse_field_spec(template, eg)               -> list[FieldSpec]
    get_iodb_instrument_types(iodb_bytes)        -> (types, error)
    get_iodb_tags(iodb_bytes, instrument_type)   -> (tags, error)
"""
from __future__ import annotations

import io
import os
import re
import json
from dataclasses import dataclass, field
from typing import Any

import pandas as pd

from .datasheet_generator import normalize

# ── Constants ────────────────────────────────────────────────────────────────
# Sheets that are never instrument datasheets.
_OMIT_SHEETS = ("cover sheet", "index", "spare", "checklist")
_ANNEXURE_RE = re.compile(r"annexure", re.IGNORECASE)
_EG_SUFFIX_RE = re.compile(r"[-_\s]*eg\s*$", re.IGNORECASE)  # "LT-RADAR-EG" -> example
_PLACEHOLDER_RE = re.compile(r"refer\s*annexure", re.IGNORECASE)

# Value region: label in cols B–E, value begins at column F (0-based index 5).
VALUE_COL = 5
# Split rows (Min/Nor/Max) put sub-values at these 0-based columns.
SPLIT_VALUE_COLS = (5, 7, 9)
SUB_LABEL_COLS = (2, 3, 4)
SOURCE_NOTE_COL = 11

IODB_TYPE_KEYS = ("instrument type", "instrument_type", "type")
IODB_NAME_KEYS = ("instrument name", "instrument_name", "instr name", "instr. name", "instr_name")
IODB_TAG_KEYS = ("tag no", "tag number", "tag_no", "tag_number", "tag")

# Field region in the EG example sheet (1-based rows 9–60 → 0-based 8–59).
DEFAULT_START_ROW = 8
DEFAULT_END_ROW = 60

CUSTOM_VALUE = "Custom Value"

# ── Dropdown knowledge base ─────────────────────────────────────────────────────
_KB_PATH = os.path.join(os.path.dirname(__file__), "datasheet_kb.json")
_kb_cache: dict | None = None


def load_kb() -> dict[str, list[str]]:
    """Load the dropdown knowledge base (normalized label -> option list)."""
    global _kb_cache
    if _kb_cache is None:
        try:
            with open(_KB_PATH, "r", encoding="utf-8") as fh:
                raw = json.load(fh)
            _kb_cache = {normalize(k): v for k, v in raw.items()}
        except Exception:
            _kb_cache = {}
    return _kb_cache


# ── Uniform cell model ─────────────────────────────────────────────────────────
@dataclass
class SheetModel:
    name: str
    nrows: int
    ncols: int
    values: dict[tuple[int, int], str]          # (row, col) -> text
    fills: dict[tuple[int, int], tuple | None]  # (row, col) -> (r,g,b) or None
    merged: list[tuple[int, int, int, int]]     # (rlo, rhi_excl, clo, chi_excl)
    col_widths: dict[int, float]                # col -> width (chars/units)
    row_heights: dict[int, float]               # row -> height (points)
    styles: dict[tuple[int, int], dict] = field(default_factory=dict)  # (row,col) -> style descriptor

    def value(self, r: int, c: int) -> str:
        return self.values.get((r, c), "")

    def fill(self, r: int, c: int) -> tuple | None:
        return self.fills.get((r, c))

    def style(self, r: int, c: int) -> dict:
        return self.styles.get((r, c), {})


@dataclass
class FieldSpec:
    section: str
    label: str
    row: int
    value_cols: list[int]
    sub_labels: list[str]          # e.g. ["Min", "Nor", "Max"]; [] for simple field
    source: str                    # "iodb" | "predefined"
    defaults: list[str]            # default value(s); aligned with value_cols
    source_note: str               # the EG "SOURCE" annotation, if any
    color: str                     # template fill color name: yellow/red/green/none
    input_type: str = "text"       # "iodb" | "dropdown" | "text"
    options: list[str] = field(default_factory=list)  # dropdown options (incl "Custom Value")
    mandatory: bool = False        # blocks generation until filled (only when no default)


@dataclass
class DatasheetInfo:
    sheet: str        # template sheet name, e.g. "LT-RADAR"
    eg_sheet: str | None
    title: str        # human title derived from the sheet content


# ── Color helpers ──────────────────────────────────────────────────────────────
def _classify_color(rgb: tuple | None) -> str:
    if not rgb:
        return "none"
    r, g, b = rgb[0], rgb[1], rgb[2]
    if r > 200 and g > 200 and b < 100:
        return "yellow"
    if r > 200 and g < 90 and b < 90:
        return "red"
    if g > 130 and r < 160 and b < 160:
        return "green"
    if r > 245 and g > 245 and b > 245:
        return "none"
    return "other"


# ── Workbook reading ─────────────────────────────────────────────────────────
def read_workbook(file_bytes: bytes, filename: str) -> dict[str, SheetModel]:
    """Read a .xls or .xlsx workbook into a uniform per-sheet cell model."""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext == ".xls":
        return _read_xls(file_bytes)
    return _read_xlsx(file_bytes)


def _read_xls(file_bytes: bytes) -> dict[str, SheetModel]:
    import xlrd

    book = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    cmap = book.colour_map
    out: dict[str, SheetModel] = {}

    for sh in book.sheets():
        values: dict[tuple[int, int], str] = {}
        fills: dict[tuple[int, int], tuple | None] = {}
        styles: dict[tuple[int, int], dict] = {}
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                if v not in ("", None):
                    values[(r, c)] = str(v)
                try:
                    xf = book.xf_list[sh.cell_xf_index(r, c)]
                    rgb = cmap.get(xf.background.pattern_colour_index)
                    if rgb:
                        fills[(r, c)] = rgb
                    styles[(r, c)] = _xls_style(book, xf, cmap)
                except Exception:
                    pass

        merged = [(rlo, rhi, clo, chi) for (rlo, rhi, clo, chi) in sh.merged_cells]
        col_widths = {
            c: (info.width / 256.0) for c, info in sh.colinfo_map.items()
        }
        row_heights = {
            r: (sh.rowinfo_map[r].height / 20.0)
            for r in sh.rowinfo_map
        }
        out[sh.name] = SheetModel(
            name=sh.name, nrows=sh.nrows, ncols=sh.ncols,
            values=values, fills=fills, merged=merged,
            col_widths=col_widths, row_heights=row_heights, styles=styles,
        )
    return out


# xlrd border-style code -> openpyxl border style name.
_XLS_BORDER = {0: None, 1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
               5: "thick", 6: "double", 7: "hair"}
# xlrd alignment codes -> openpyxl names.
_XLS_HALIGN = {0: "general", 1: "left", 2: "center", 3: "right", 4: "fill",
               5: "justify", 6: "centerContinuous", 7: "distributed"}
_XLS_VALIGN = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}


def _xls_style(book, xf, cmap) -> dict:
    """Normalize an xlrd XF record into a uniform style descriptor."""
    f = book.font_list[xf.font_index]
    frgb = cmap.get(f.colour_index) or (0, 0, 0)
    brd = xf.border

    def bstyle(code):
        return _XLS_BORDER.get(code, "thin" if code else None)

    return {
        "font": {
            "name": f.name or "Arial",
            "size": (f.height / 20.0) if f.height else 10.0,
            "bold": bool(f.weight and f.weight >= 700) or bool(getattr(f, "bold", 0)),
            "italic": bool(f.italic),
            "color": "%02X%02X%02X" % frgb,
        },
        "align": {
            "h": _XLS_HALIGN.get(xf.alignment.hor_align, "general"),
            "v": _XLS_VALIGN.get(xf.alignment.vert_align, "bottom"),
            "wrap": bool(xf.alignment.text_wrapped),
        },
        "border": {
            "top": bstyle(brd.top_line_style),
            "bottom": bstyle(brd.bottom_line_style),
            "left": bstyle(brd.left_line_style),
            "right": bstyle(brd.right_line_style),
        },
    }


def _openpyxl_style(cell) -> dict:
    """Normalize an openpyxl cell's style into the same descriptor shape."""
    f = cell.font
    color = "000000"
    try:
        if f.color and f.color.rgb and isinstance(f.color.rgb, str):
            color = f.color.rgb[-6:]
    except Exception:
        pass
    b = cell.border
    al = cell.alignment
    return {
        "font": {
            "name": f.name or "Arial",
            "size": float(f.sz or 10.0),
            "bold": bool(f.bold),
            "italic": bool(f.italic),
            "color": color,
        },
        "align": {
            "h": al.horizontal or "general",
            "v": al.vertical or "bottom",
            "wrap": bool(al.wrap_text),
        },
        "border": {
            "top": b.top.style if b.top else None,
            "bottom": b.bottom.style if b.bottom else None,
            "left": b.left.style if b.left else None,
            "right": b.right.style if b.right else None,
        },
    }


def _read_xlsx(file_bytes: bytes) -> dict[str, SheetModel]:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    out: dict[str, SheetModel] = {}
    for ws in wb.worksheets:
        values: dict[tuple[int, int], str] = {}
        fills: dict[tuple[int, int], tuple | None] = {}
        styles: dict[tuple[int, int], dict] = {}
        nrows = ws.max_row or 0
        ncols = ws.max_column or 0
        for row in ws.iter_rows():
            for cell in row:
                r, c = cell.row - 1, cell.column - 1
                if cell.value not in (None, ""):
                    values[(r, c)] = str(cell.value)
                rgb = _openpyxl_fill_rgb(cell)
                if rgb:
                    fills[(r, c)] = rgb
                styles[(r, c)] = _openpyxl_style(cell)

        merged = []
        for mr in ws.merged_cells.ranges:
            merged.append((mr.min_row - 1, mr.max_row, mr.min_col - 1, mr.max_col))

        col_widths = {}
        for letter, dim in ws.column_dimensions.items():
            if dim.width:
                try:
                    col_widths[column_index_from_string(letter) - 1] = dim.width
                except Exception:
                    pass
        row_heights = {r - 1: d.height for r, d in ws.row_dimensions.items() if d.height}

        out[ws.title] = SheetModel(
            name=ws.title, nrows=nrows, ncols=ncols,
            values=values, fills=fills, merged=merged,
            col_widths=col_widths, row_heights=row_heights, styles=styles,
        )
    return out


def _openpyxl_fill_rgb(cell) -> tuple | None:
    try:
        fill = cell.fill
        if not fill or fill.patternType not in ("solid",):
            return None
        argb = fill.fgColor.rgb
        if not argb or not isinstance(argb, str) or len(argb) < 6:
            return None
        hexpart = argb[-6:]
        return (int(hexpart[0:2], 16), int(hexpart[2:4], 16), int(hexpart[4:6], 16))
    except Exception:
        return None


# ── Datasheet sheet discovery ──────────────────────────────────────────────────
def list_datasheet_sheets(sheets: dict[str, SheetModel]) -> list[DatasheetInfo]:
    """Return instrument-datasheet template sheets, pairing each with its EG example."""
    names = list(sheets.keys())
    eg_map: dict[str, str] = {}
    for n in names:
        if _EG_SUFFIX_RE.search(n):
            base = _EG_SUFFIX_RE.sub("", n).strip()
            eg_map[normalize(base)] = n

    infos: list[DatasheetInfo] = []
    for n in names:
        low = n.strip().lower()
        if low in _OMIT_SHEETS or _ANNEXURE_RE.search(n) or _EG_SUFFIX_RE.search(n):
            continue
        eg = eg_map.get(normalize(n))
        infos.append(DatasheetInfo(sheet=n, eg_sheet=eg, title=_sheet_title(sheets[n])))
    return infos


def _sheet_title(sm: SheetModel) -> str:
    """Best-effort human title from the top rows of the sheet."""
    parts = []
    for r in range(0, 4):
        for c in range(0, sm.ncols):
            v = sm.value(r, c).strip()
            if v and "DATASHEET" in v.upper():
                # Title usually sits adjacent (the next non-empty cell after the header block)
                pass
    # The instrument descriptor (e.g. "LEVEL TRANSMITTER- NON CONTACT RADAR TYPE")
    for r in range(0, 5):
        for c in range(0, sm.ncols):
            v = sm.value(r, c).strip()
            if v and ("TRANSMITTER" in v.upper() or "GAUGE" in v.upper() or "SWITCH" in v.upper()):
                return re.sub(r"\s+", " ", v).strip()
    _ = parts
    return sm.name


# ── Field-spec parsing ──────────────────────────────────────────────────────────
def _is_split_row(sm: SheetModel, r: int) -> bool:
    subs = [sm.value(r, c).strip().lower() for c in SUB_LABEL_COLS]
    return subs[:2] == ["min", "nor"] or subs == ["min", "nor", "max"]


def _section_at(sm: SheetModel, r: int) -> str:
    return sm.value(r, 0).strip()


def parse_field_spec(
    template: SheetModel,
    eg: SheetModel | None,
    start_row: int = DEFAULT_START_ROW,
    end_row: int = DEFAULT_END_ROW,
) -> list[FieldSpec]:
    """
    Build the ordered list of spec fields for a datasheet, using the EG example
    as the source of truth for defaults and IODB-vs-predefined classification.

    Each field is further classified into an input_type for the config popup:
      * "iodb"     — value comes from the IODB per tag (yellow)
      * "dropdown" — label found in the knowledge base (green); options + default
      * "text"     — free-text mandatory input (red); blocking only when no default

    `start_row`/`end_row` bound the spec region (0-based, end exclusive) so the
    EG example sheet's data table (rows 9–60) is targeted precisely.
    """
    kb = load_kb()
    fields: list[FieldSpec] = []
    section = ""

    # Bound to the requested region, but never run past a NOTES/footer marker.
    auto_start, auto_end = _table_bounds(template)
    start = max(start_row, 0)
    end = min(end_row, template.nrows, auto_end if auto_end > start else template.nrows)
    if start >= end:
        start, end = auto_start, auto_end

    for r in range(start, end):
        sec = _section_at(template, r)
        if sec:
            section = re.sub(r"\s+", " ", sec).strip()
        label = template.value(r, 1).strip()
        if not label:
            continue
        if label.upper().startswith("NOTES"):
            break
        # Skip document-control placeholders (CLIENT/PROJECT header & footer doc-no
        # blocks carry an "XXXXXX" placeholder in the label column).
        if re.fullmatch(r"[Xx\s]+", label):
            continue

        label = re.sub(r"\s+", " ", label)
        split = _is_split_row(template, r)
        value_cols = list(SPLIT_VALUE_COLS) if split else [VALUE_COL]
        sub_labels = [template.value(r, c).strip() for c in SUB_LABEL_COLS] if split else []

        eg_sheet = eg if eg is not None else template
        defaults = [eg_sheet.value(r, c).strip() for c in value_cols]

        is_iodb = all((not d) or bool(_PLACEHOLDER_RE.search(d)) for d in defaults)
        source = "iodb" if is_iodb else "predefined"

        note = template.value(r, SOURCE_NOTE_COL).strip() or (
            eg_sheet.value(r, SOURCE_NOTE_COL).strip()
        )

        # Classify input_type for the configuration popup.
        primary_default = next((d for d in defaults if d), "")
        kb_options = kb.get(normalize(label))
        if is_iodb:
            input_type = "iodb"
            options: list[str] = []
            mandatory = False
            color = "yellow"
        elif kb_options:
            input_type = "dropdown"
            # Move the template default to the front, then append remaining options + Custom Value.
            opts = [o for o in kb_options]
            if primary_default and primary_default not in opts:
                opts = [primary_default] + opts
            elif primary_default in opts:
                opts = [primary_default] + [o for o in opts if o != primary_default]
            options = opts + [CUSTOM_VALUE]
            mandatory = False
            color = "green"
        else:
            input_type = "text"
            options = []
            # Blocking only when there is no EG default to fall back on.
            mandatory = not bool(primary_default)
            color = "red"

        fields.append(FieldSpec(
            section=section, label=label, row=r,
            value_cols=value_cols, sub_labels=sub_labels,
            source=source, defaults=defaults, source_note=note, color=color,
            input_type=input_type, options=options, mandatory=mandatory,
        ))
    return fields


def _table_bounds(sm: SheetModel) -> tuple[int, int]:
    start = None
    for r in range(sm.nrows):
        if _section_at(sm, r):
            start = r
            break
    if start is None:
        start = 0
    end = sm.nrows
    for r in range(start, sm.nrows):
        if sm.value(r, 1).strip().upper().startswith("NOTES"):
            end = r
            break
    return start, end


# ── IODB helpers ─────────────────────────────────────────────────────────────
def _find_iodb_sheet(xls: pd.ExcelFile) -> str:
    for s in xls.sheet_names:
        if "iodb" in s.strip().lower():
            return s
    return xls.sheet_names[0]


def _load_iodb_df(iodb_bytes: bytes) -> tuple[pd.DataFrame | None, str | None]:
    try:
        buf = io.BytesIO(iodb_bytes)
        xls = pd.ExcelFile(buf)
        sheet = _find_iodb_sheet(xls)
        # Probe up to the first few rows for the header row containing TAG + TYPE.
        raw = pd.read_excel(buf, sheet_name=sheet, header=None, nrows=10)
        header_row = 0
        for i in range(min(10, len(raw))):
            cells = [normalize(v) for v in raw.iloc[i].tolist()]
            if any(k in cells for k in IODB_TAG_KEYS):
                header_row = i
                break
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name=sheet, header=header_row)
        df = df.dropna(how="all").reset_index(drop=True)
        return df, None
    except Exception as exc:
        return None, str(exc)


def _find_column(df: pd.DataFrame, keys: tuple[str, ...]) -> str | None:
    norm_keys = [normalize(k) for k in keys]
    # exact-ish first
    for col in df.columns:
        if normalize(col) in norm_keys:
            return col
    for col in df.columns:
        n = normalize(col)
        if any(k in n for k in norm_keys):
            return col
    return None


_BLANK_VALS = {"", "nan", "none"}


def get_iodb_instrument_types(iodb_bytes: bytes) -> tuple[list[str], str | None]:
    """
    Return unique instrument type combinations from the IODB.

    When both INSTRUMENT NAME and INSTRUMENT TYPE columns exist, the dropdown
    value is the combined string:
        "INSTRUMENT NAME - INSTRUMENT TYPE"

    This uniquely identifies the instrument family (e.g.
    "LEVEL TRANSMITTER - NON CONTACT RADAR TYPE") so the same generic type
    word like "TRANSMITTER" is never ambiguous across technologies.

    Blank, NULL, and duplicate combinations are removed. Result is sorted.
    """
    df, err = _load_iodb_df(iodb_bytes)
    if err:
        return [], err
    type_col = _find_column(df, IODB_TYPE_KEYS)
    if type_col is None:
        return [], "No 'INSTRUMENT TYPE' column found in IODB."

    name_col = _find_column(df, IODB_NAME_KEYS)
    combined: set[str] = set()

    for _, row in df.iterrows():
        itype = str(row.get(type_col, "") or "").strip()
        if not itype or itype.lower() in _BLANK_VALS:
            continue
        if name_col is not None:
            name = str(row.get(name_col, "") or "").strip()
            if name and name.lower() not in _BLANK_VALS:
                combined.add(f"{name} - {itype}")
                continue
        combined.add(itype)

    return sorted(combined), None


def get_iodb_tags(iodb_bytes: bytes, instrument_type: str) -> tuple[list[str], str | None]:
    """
    Return tags whose rows match `instrument_type`.

    `instrument_type` is now a combined "INSTRUMENT NAME - INSTRUMENT TYPE"
    string (as produced by get_iodb_instrument_types). The function rebuilds
    every (name, type) pair from the IODB to find the exact match, then
    filters rows where BOTH columns equal the found values.

    Falls back to matching on INSTRUMENT TYPE alone for older callers or when
    the INSTRUMENT NAME column is absent.
    """
    df, err = _load_iodb_df(iodb_bytes)
    if err:
        return [], err
    type_col = _find_column(df, IODB_TYPE_KEYS)
    tag_col = _find_column(df, IODB_TAG_KEYS)
    if tag_col is None:
        return [], "No 'TAG NO' column found in IODB."
    if type_col is None:
        return [], "No 'INSTRUMENT TYPE' column found in IODB."

    name_col = _find_column(df, IODB_NAME_KEYS)
    it_lower = instrument_type.strip().lower()

    if name_col is not None:
        # Search every row for the (name, type) pair that produces the combined string.
        found_name: str | None = None
        found_type: str | None = None
        for _, row in df.iterrows():
            n = str(row.get(name_col, "") or "").strip()
            t = str(row.get(type_col, "") or "").strip()
            if n.lower() in _BLANK_VALS or t.lower() in _BLANK_VALS:
                continue
            if f"{n} - {t}".lower() == it_lower:
                found_name = n
                found_type = t
                break

        if found_name is not None:
            # Filter by both INSTRUMENT NAME and INSTRUMENT TYPE (exact, case-insensitive).
            name_mask = (
                df[name_col].fillna("").astype(str).str.strip().str.lower()
                == found_name.lower()
            )
            type_mask = (
                df[type_col].fillna("").astype(str).str.strip().str.lower()
                == found_type.lower()  # type: ignore[arg-type]
            )
            mask = name_mask & type_mask
        else:
            # Combined string not matched — fall back to INSTRUMENT TYPE only.
            mask = (
                df[type_col].fillna("").astype(str).str.strip().str.lower()
                == it_lower
            )
    else:
        mask = (
            df[type_col].fillna("").astype(str).str.strip().str.lower()
            == it_lower
        )

    tags = (
        df.loc[mask, tag_col].dropna().astype(str).str.strip()
        .loc[lambda s: s != ""].unique().tolist()
    )
    return sorted(tags), None


# ── Generation (Phase 2) ──────────────────────────────────────────────────────
import zipfile  # noqa: E402
from .datasheet_generator import fuzzy_match  # noqa: E402

# Recolor scheme for the generated output.
_FILL_YELLOW = "FFFF00"   # IODB-sourced
_FILL_GREEN = "C6EFCE"    # predefined (filled)
_FILL_RED = "FFC7CE"      # mandatory left blank


def _apply_cell_style(cell, style: dict, fill_hex: str | None) -> None:
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

    if not style:
        if fill_hex:
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        return

    f = style.get("font", {})
    cell.font = Font(
        name=f.get("name", "Arial"), size=f.get("size", 10.0),
        bold=f.get("bold", False), italic=f.get("italic", False),
        color=f.get("color", "000000"),
    )
    al = style.get("align", {})
    cell.alignment = Alignment(
        horizontal=(al.get("h") if al.get("h") not in (None, "general") else None),
        vertical=al.get("v", "bottom"),
        wrap_text=al.get("wrap", False),
    )
    b = style.get("border", {})

    def side(s):
        return Side(style=s) if s else Side()

    cell.border = Border(top=side(b.get("top")), bottom=side(b.get("bottom")),
                         left=side(b.get("left")), right=side(b.get("right")))

    if fill_hex:
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")


def build_datasheet_xlsx(template: SheetModel, fields: list[FieldSpec],
                         resolved: dict[tuple[int, int], tuple[str, str]]):
    """
    Rebuild a styled worksheet for one tag.

    `resolved` maps a value cell coordinate (row, col) -> (value, source) where
    source ∈ {"iodb", "predefined"} and drives the recolor.
    Returns an openpyxl Workbook.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = template.name[:31]

    # Column widths / row heights
    for c, w in template.col_widths.items():
        ws.column_dimensions[get_column_letter(c + 1)].width = w
    for r, h in template.row_heights.items():
        ws.row_dimensions[r + 1].height = h

    # Write every cell that has a value or a style. Only the VALUE changes for
    # resolved cells — the template's own fill/formatting is preserved (the
    # red/green/yellow scheme is a UI concept, not applied to the output).
    coords = set(template.values.keys()) | set(template.styles.keys())
    for (r, c) in coords:
        cell = ws.cell(row=r + 1, column=c + 1)
        if (r, c) in resolved:
            cell.value = resolved[(r, c)][0]
        else:
            cell.value = template.value(r, c) or None
        rgb = template.fill(r, c)
        fill = ("%02X%02X%02X" % rgb) if rgb else None
        _apply_cell_style(cell, template.style(r, c), fill)

    # Merge ranges (after values/styles are set).
    for (rlo, rhi, clo, chi) in template.merged:
        try:
            ws.merge_cells(start_row=rlo + 1, start_column=clo + 1,
                          end_row=rhi, end_column=chi)
        except Exception:
            pass

    return wb


def _build_iodb_colmap(df: pd.DataFrame, fields: list[FieldSpec],
                       threshold: int = 75) -> dict[int, str]:
    """Fuzzy-map each IODB-sourced field (by row) to the best IODB column."""
    choices = [str(col) for col in df.columns]
    colmap: dict[int, str] = {}
    for f in fields:
        if f.source != "iodb":
            continue
        best, score = fuzzy_match(f.label, choices, threshold)
        if best is not None:
            colmap[f.row] = best
    return colmap


def _resolve_values_for_tag(
    fields: list[FieldSpec],
    overrides: dict[str, list[str]],
    iodb_row: pd.Series | None,
    iodb_colmap: dict[int, str],
    tag: str,
) -> dict[tuple[int, int], tuple[str, str]]:
    """Resolve every field's value cell for one tag → {(row,col): (value, source)}."""
    resolved: dict[tuple[int, int], tuple[str, str]] = {}
    for f in fields:
        # Per-field user override keyed by label (falls back to defaults).
        ov = overrides.get(f.label)
        for idx, col in enumerate(f.value_cols):
            default = f.defaults[idx] if idx < len(f.defaults) else ""
            if f.source == "predefined":
                val = ov[idx] if (ov and idx < len(ov) and ov[idx] != "") else default
                resolved[(f.row, col)] = (val, "predefined")
            else:
                # IODB-sourced
                val = default  # usually "REFER ANNEXURE-I"
                if f.section.strip().upper() == "GENERAL" and normalize(f.label) == "tag no":
                    val = tag
                elif iodb_row is not None and f.row in iodb_colmap:
                    raw = iodb_row.get(iodb_colmap[f.row])
                    if raw is not None and str(raw).strip() not in ("", "nan"):
                        val = str(raw).strip()
                resolved[(f.row, col)] = (val, "iodb")
    return resolved


def _generate_xlsx_inplace(sop_bytes: bytes, sheet_name: str,
                           resolved: dict[tuple[int, int], tuple[str, str]]) -> bytes:
    """
    Fill ONLY the resolved value cells of `sheet_name` in the original .xlsx,
    keeping just that sheet — preserving logos/images/formulas/page setup/
    borders/merged cells exactly (nothing else is touched).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(sop_bytes))   # formulas + images preserved
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]

    for (r, c), (value, _src) in resolved.items():
        try:
            ws.cell(row=r + 1, column=c + 1).value = value
        except (AttributeError, ValueError):
            # Merged non-anchor cell — skip (anchor already written).
            pass

    # Keep only the target datasheet for a clean per-tag output.
    for sn in list(wb.sheetnames):
        if sn != sheet_name:
            del wb[sn]

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def generate(
    sop_bytes: bytes, sop_filename: str, datasheet_sheet: str,
    iodb_bytes: bytes, instrument_type: str, selected_tags: list[str],
    overrides: dict[str, list[str]] | None = None,
) -> tuple[bytes | None, str, str | None]:
    """
    Generate one datasheet (.xlsx) per selected tag, zipped.

    For .xlsx SOPs the original workbook is loaded and only value cells are
    written (full formatting/logo/formula preservation). For .xls SOPs the
    sheet is rebuilt from the parsed model (best-effort; no logos/images).

    Returns (zip_bytes, filename, error).
    """
    overrides = overrides or {}
    is_xlsx = (sop_filename.rsplit(".", 1)[-1].lower() in ("xlsx", "xlsm")) if "." in sop_filename else False

    try:
        sheets = read_workbook(sop_bytes, sop_filename)
    except Exception as exc:
        return None, "Datasheets.zip", f"Failed to read SOP workbook: {exc}"
    if datasheet_sheet not in sheets:
        return None, "Datasheets.zip", f"Sheet '{datasheet_sheet}' not found in SOP workbook."

    # Target the EG example sheet (it carries the predefined defaults + formatting).
    infos = {d.sheet: d for d in list_datasheet_sheets(sheets)}
    info = infos.get(datasheet_sheet)
    target_name = info.eg_sheet if (info and info.eg_sheet) else datasheet_sheet
    target_model = sheets.get(target_name) or sheets[datasheet_sheet]
    fields = parse_field_spec(target_model, target_model)

    df, err = _load_iodb_df(iodb_bytes)
    if err:
        return None, "Datasheets.zip", f"Failed to load IODB: {err}"
    tag_col = _find_column(df, IODB_TAG_KEYS)
    if tag_col is None:
        return None, "Datasheets.zip", "No 'TAG NO' column found in IODB."

    iodb_colmap = _build_iodb_colmap(df, fields)

    zip_buf = io.BytesIO()
    written = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for tag in selected_tags:
            iodb_row = None
            match = df[df[tag_col].astype(str).str.strip() == str(tag).strip()]
            if not match.empty:
                iodb_row = match.iloc[0]

            resolved = _resolve_values_for_tag(fields, overrides, iodb_row, iodb_colmap, tag)

            if is_xlsx:
                file_bytes = _generate_xlsx_inplace(sop_bytes, target_name, resolved)
            else:
                wb = build_datasheet_xlsx(target_model, fields, resolved)
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                file_bytes = buf.read()

            safe_tag = re.sub(r'[\\/*?:\[\]]', "_", str(tag))
            zf.writestr(f"{safe_tag} Datasheet.xlsx", file_bytes)
            written += 1

    if written == 0:
        return None, "Datasheets.zip", "No datasheets generated — no matching tags."

    zip_buf.seek(0)
    return zip_buf.read(), "Datasheets.zip", None
