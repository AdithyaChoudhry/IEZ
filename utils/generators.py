"""
generators.py
Core generation functions for each of the 5 modules.
All functions return (output_bytes, filename, error_message).
  - output_bytes: bytes of the generated Excel file (or None on error)
  - filename:     suggested download filename  (or None on error)
  - error_message: None on success, string description on failure
"""

import io
import copy
import re
import zipfile
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, PatternFill, Border

from utils.file_handler import workbook_to_bytes, dataframe_to_bytes

# Shared cell style applied to all output cells in list generators
_LIST_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LIST_FONT = Font(name='Calibri', size=14)
_HEADER_FONT = Font(name='Calibri', size=14, bold=True)


# ─────────────────────────────────────────────────────────────────────────────
# 1. Instrument List Generator
# ─────────────────────────────────────────────────────────────────────────────

def generate_instrument_list(
    df: pd.DataFrame,
    selected_columns: list[str],
) -> tuple[bytes | None, str | None, str | None]:
    """
    Extract selected columns from the IODB DataFrame and produce a formatted Excel file.
    Applies Calibri 14pt, center+wrap alignment to all cells, and auto-fits column widths.
    """
    try:
        available = [c for c in selected_columns if c in df.columns]
        if not available:
            return None, None, "None of the selected columns were found in the IODB file."
        result = df[available].copy()
        out_bytes = _build_formatted_list(result, "Sheet1")
        return out_bytes, "Instrument_List.xlsx", None
    except Exception as e:
        return None, None, f"Instrument List generation failed: {e}"


# ─────────────────────────────────────────────────────────────────────────────
# 2. I/O List Generator
# ─────────────────────────────────────────────────────────────────────────────

def generate_io_list(
    df: pd.DataFrame,
    selected_columns: list[str],
) -> tuple[bytes | None, str | None, str | None]:
    """
    Same logic as Instrument List but outputs IO_List.xlsx.
    Applies Calibri 14pt, center+wrap alignment to all cells, and auto-fits column widths.
    """
    try:
        available = [c for c in selected_columns if c in df.columns]
        if not available:
            return None, None, "None of the selected columns were found in the IODB file."
        result = df[available].copy()
        out_bytes = _build_formatted_list(result, "Sheet1")
        return out_bytes, "IO_List.xlsx", None
    except Exception as e:
        return None, None, f"I/O List generation failed: {e}"


def _build_formatted_list(df: pd.DataFrame, sheet_name: str) -> bytes:
    """
    Write a DataFrame to an Excel workbook with:
      - Calibri 14 font on every cell (bold for header row)
      - Center + center + wrap_text alignment on every cell
      - Auto-adjust column width to the longest value in each column
    Returns raw bytes of the xlsx file.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = wb[sheet_name]

        # Track max width per column for auto-fit
        col_widths: dict[int, int] = {}

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            is_header = row_idx == 1
            for cell in row:
                cell.alignment = _LIST_ALIGNMENT
                cell.font = _HEADER_FONT if is_header else _LIST_FONT

                # Measure display length (handle multi-line cell values)
                if cell.value is not None:
                    val_str = str(cell.value)
                    # For wrapped text, use longest line
                    line_len = max((len(line) for line in val_str.split('\n')), default=0)
                else:
                    line_len = 0
                col_widths[cell.column] = max(col_widths.get(cell.column, 0), line_len)

        # Apply column widths (cap between 10 and 60 chars)
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 4, 10), 60)

    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# 3. Data Sheet Generator
# ─────────────────────────────────────────────────────────────────────────────

# Values in datasheet value cells that are considered "empty placeholders".
# ONLY these are replaced; anything else ("NA", "VTS", numbers, etc.) is preserved.
_PLACEHOLDER_VALUES = {"refer annexure 1", "refer annexure1", "annexure 1", "", None}


def generate_datasheets(
    df: pd.DataFrame,
    template_wb: openpyxl.Workbook,
    tag_column: str,
    selected_tags: list[str],
    progress_callback=None,
) -> tuple[list[tuple[bytes, str]] | None, str | None]:
    """
    For each selected tag, copy the entire template workbook, then perform
    layout-aware field mapping on the 'Annexure' sheet:

      1. Scan EVERY cell in the sheet.
      2. If the cell text matches an IODB column name (case-insensitive, with
         fuzzy fallback), look rightward in the same row (up to 5 cells).
      3. Write the IODB value into the first rightward cell whose current value
         is a placeholder (None, empty, or "Refer Annexure 1").
      4. Never overwrite formulas, "NA", "VTS", or any other meaningful text.
      5. Write "TBA" when the IODB value is empty/NaN.

    Preserves all formatting, merged cells, and structure — only placeholder
    cells are touched.
    """
    try:
        # ───────────────────────────────────────────────────────────────────────
        # Locate Annexure sheet (case-insensitive, then partial match)
        # ───────────────────────────────────────────────────────────────────────
        ANNEXURE_SHEET = next(
            (s for s in template_wb.sheetnames if s.strip().lower() == "annexure"),
            None,
        )
        # Also accept "Annexure 1- NEW WWTP", "Annexure 1", etc.
        if ANNEXURE_SHEET is None:
            ANNEXURE_SHEET = next(
                (s for s in template_wb.sheetnames if "annexure" in s.strip().lower()),
                None,
            )
        if ANNEXURE_SHEET is None:
            return None, (
                f"Template does not contain a sheet named 'Annexure'. "
                f"Found sheets: {template_wb.sheetnames}"
            )

        if tag_column not in df.columns:
            return None, f"Tag column '{tag_column}' not found in IODB."

        # ───────────────────────────────────────────────────────────────────────
        # Build normalised IODB column index for fast lookup
        # ───────────────────────────────────────────────────────────────────────
        # Map: normalised_col_name -> original column name (preserves case for df access)
        iodb_col_norm: dict[str, str] = {
            str(c).strip().lower(): str(c) for c in df.columns
        }
        iodb_col_names_norm = list(iodb_col_norm.keys())

        # ───────────────────────────────────────────────────────────────────────
        # Pre-scan template sheet ONCE to build a label-position map.
        # For each cell that contains text matching an IODB column:
        #   label_map[norm_label] = (row, col_of_label_cell)
        # We defer finding the target write-cell to fill-time so we can
        # respect per-tag placeholder state.
        # ───────────────────────────────────────────────────────────────────────
        tmpl_ws = template_wb[ANNEXURE_SHEET]
        # label_map: norm_label -> (row, label_col)
        label_map: dict[str, tuple[int, int]] = {}
        for t_row in tmpl_ws.iter_rows():
            for t_cell in t_row:
                if t_cell.value is None:
                    continue
                norm = str(t_cell.value).strip().lower()
                if not norm:
                    continue
                # Exact match
                if norm in iodb_col_norm:
                    label_map[norm] = (t_cell.row, t_cell.column)
                else:
                    # Fuzzy: check if template label contains or is contained
                    # by an IODB column name (handles "Line Size (NB)" vs "Line size")
                    for iodb_norm in iodb_col_names_norm:
                        if (iodb_norm in norm or norm in iodb_norm) and len(norm) >= 3:
                            if norm not in label_map:
                                label_map[norm] = (t_cell.row, t_cell.column)
                            break

        results: list[tuple[bytes, str]] = []
        total = len(selected_tags)

        for idx, tag in enumerate(selected_tags):
            if progress_callback:
                progress_callback(idx, total)

            # Find IODB row for this tag
            mask = df[tag_column].astype(str).str.strip() == str(tag).strip()
            if not mask.any():
                continue
            row_data = df[mask].iloc[0].to_dict()

            # Build normalised lookup {norm_col -> value}
            iodb_values: dict[str, object] = {}
            for col, val in row_data.items():
                norm_key = str(col).strip().lower()
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    iodb_values[norm_key] = "TBA"
                else:
                    iodb_values[norm_key] = val

            # Deep-copy workbook so every tag gets pristine formatting
            new_wb = _copy_workbook(template_wb)
            ws = new_wb[ANNEXURE_SHEET]

            # ────────────────────────────────────────────────
            # Spatial fill: for each matched label cell,
            # scan rightward for a placeholder and write the IODB value.
            # ────────────────────────────────────────────────
            _spatial_fill(ws, label_map, iodb_values, iodb_col_norm)

            out_bytes = workbook_to_bytes(new_wb)
            safe_tag = re.sub(r'[\\/*?:\[\]]', '_', str(tag))
            filename = f"Datasheet_{safe_tag}.xlsx"
            results.append((out_bytes, filename))

        if progress_callback:
            progress_callback(total, total)

        return results, None

    except Exception as e:
        return None, f"Datasheet generation failed: {e}"


def _is_placeholder(value) -> bool:
    """
    Return True if a cell value is a placeholder that may be overwritten.
    Placeholders: None, empty string, or any string that contains
    'refer annexure' (case-insensitive).
    Any other value — 'NA', 'VTS', a number, a formula — is preserved.
    """
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return True
        if "refer annexure" in stripped.lower() or stripped.lower() == "annexure 1":
            return True
    return False


def _spatial_fill(
    ws,
    label_map: dict[str, tuple[int, int]],
    iodb_values: dict[str, object],
    iodb_col_norm: dict[str, str],
    max_right_search: int = 5,
):
    """
    Fill worksheet by spatial label-to-value mapping.

    For each (norm_label -> (row, label_col)) in label_map:
      1. Resolve which IODB column this label corresponds to.
      2. Get the IODB value for this tag.
      3. Scan cells to the RIGHT (same row) for the first placeholder.
      4. Write the value there.  Never overwrite formulas or meaningful content.
    """
    for norm_label, (r, label_col) in label_map.items():
        # Resolve IODB value for this label
        iodb_value = None
        if norm_label in iodb_values:
            iodb_value = iodb_values[norm_label]
        else:
            # Try fuzzy resolution: find the best-matching IODB column
            for iodb_norm, orig_col in iodb_col_norm.items():
                if (iodb_norm in norm_label or norm_label in iodb_norm) and len(norm_label) >= 3:
                    raw = iodb_values.get(iodb_norm)
                    if raw is not None:
                        iodb_value = raw
                    break

        if iodb_value is None:
            # No IODB column matched this label at all — do nothing
            continue

        # Search rightward for a placeholder cell to write into
        for offset in range(1, max_right_search + 1):
            target = ws.cell(row=r, column=label_col + offset)
            # Skip read-only merged-cell slaves (non-top-left cells in a merge)
            if target.__class__.__name__ == "MergedCell":
                continue
            # Never touch formula cells
            if isinstance(target.value, str) and target.value.startswith("="):
                break  # formula encountered — stop scanning right
            if _is_placeholder(target.value):
                target.value = iodb_value
                break  # wrote successfully — move on to next label


# ─────────────────────────────────────────────────────────────────────────────
# 4. Cable Schedule Generator
# ─────────────────────────────────────────────────────────────────────────────

# The target sheet inside the cable schedule template
CABLE_SCHEDULE_SHEET = "Cable Schedule -INST"

# Header row index (1-based) inside the cable schedule template
CABLE_SCHEDULE_HEADER_ROW = 7
# First data row (1-based)
CABLE_SCHEDULE_FIRST_DATA_ROW = 9
# Rows per tag in the template (the reference shows 3 rows per tag entry)
ROWS_PER_TAG = 3
# Max tags per JB section (kept for reference; no longer enforced)
MAX_TAGS_PER_JB = 12

# Cable Schedule cell formatting (Arial 12, centered, wrap)
_CS_FONT      = Font(name='Arial', size=12)
_CS_FONT_BOLD = Font(name='Arial', size=12, bold=True)
_CS_ALIGN     = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _adjust_formula_row(formula: str, src_row: int, dst_row: int) -> str:
    """
    When copying a formula from src_row to dst_row, adjust every *relative*
    row reference by the offset (dst_row - src_row).

    Rules (standard Excel copy-paste semantics):
      A9   / $A9   → row is relative  → shift by offset
      A$9  / $A$9  → row is absolute  → leave unchanged
    """
    if not isinstance(formula, str) or not formula.startswith('=') or dst_row == src_row:
        return formula
    offset = dst_row - src_row

    def _sub(m):
        col_tok  = m.group(1)   # e.g. "A" or "$A"
        abs_mark = m.group(2)   # "$" → absolute row, "" → relative row
        row_n    = int(m.group(3))
        if abs_mark:            # absolute row – leave as-is
            return m.group(0)
        return f'{col_tok}{row_n + offset}'

    return re.sub(r'(\$?[A-Za-z]+)(\$?)(\d+)', _sub, formula)


def _find_cs_sheet(template_wb) -> str | None:
    """Return the Cable Schedule sheet name; fuzzy-matches if exact name is missing."""
    if CABLE_SCHEDULE_SHEET in template_wb.sheetnames:
        return CABLE_SCHEDULE_SHEET
    for s in template_wb.sheetnames:
        if 'cable schedule' in s.lower() and 'inst' in s.lower():
            return s
    for s in template_wb.sheetnames:
        if 'cable' in s.lower():
            return s
    return None


def generate_cable_schedule(
    df: pd.DataFrame,
    template_wb: openpyxl.Workbook,
    jb_column: str = "JUNCTION BOX",
    tag_column: str = "TAG NO",
    progress_callback=None,
    template_bytes: bytes = None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Generate a Cable Schedule from IODB data using the uploaded template.

    Logic:
    - Filters IODB rows where JUNCTION BOX has a non-blank value other than "-".
    - Sorts ascending by JUNCTION BOX (natural / alphanumeric order).
    - Groups by JB value; all tags for a JB are written consecutively.
    - For each tag, writes one 3-row block matching the template structure:
        * Copies the template prototype block (rows 9-11) cell by cell.
        * Formula cells are carried forward with row references adjusted for the
          new row position (e.g. row-9 formula with ref to "row 9" → adjusts to
          destination row so sequential counters like "06" → "07" work correctly).
        * Non-formula cells are filled from the matching IODB column via
          case-insensitive / fuzzy header matching.
        * Missing IODB values → "TBA".
    - All generated cells: Arial 12, center/center alignment, wrap-text.

    Args:
        df: IODB DataFrame
        template_wb: openpyxl Workbook (may be read-only) — used to read structure.
        jb_column: IODB column containing Junction Box numbers.
        tag_column: IODB column containing tag numbers.
        progress_callback: optional callable(current, total).
        template_bytes: raw bytes of template file; when supplied a fresh
            writable copy is built from it (preserving all template formatting).

    Returns:
        (bytes, "Cable_Schedule.xlsx", None) on success
        (None, None, error_string) on failure
    """
    try:
        # ── Locate cable schedule sheet ──────────────────────────────────────
        target_sheet = _find_cs_sheet(template_wb)
        if target_sheet is None:
            return None, None, (
                f"Could not find Cable Schedule sheet. "
                f"Available: {template_wb.sheetnames}"
            )
        tmpl_ws = template_wb[target_sheet]

        # ── Build header map: col_idx → normalised lower-case header text ────
        # Combines main header row (7) and sub-header row (8).
        header_map: dict[int, str] = {}
        for hr_off in range(2):
            for row in tmpl_ws.iter_rows(
                min_row=CABLE_SCHEDULE_HEADER_ROW + hr_off,
                max_row=CABLE_SCHEDULE_HEADER_ROW + hr_off,
            ):
                for col_idx, cell in enumerate(row, 1):
                    val = getattr(cell, 'value', None)
                    if val is not None and col_idx not in header_map:
                        txt = str(val).strip()
                        if txt:
                            header_map[col_idx] = txt.lower()

        # Reverse: normalised text → col_idx
        rev_hdr: dict[str, int] = {v: k for k, v in header_map.items()}

        # ── Match IODB columns → template column indices (fuzzy) ─────────────
        iodb_norm: dict[str, str] = {str(c).strip().lower(): str(c) for c in df.columns}
        col_match: dict[str, int] = {}  # iodb_col_original → template_col_idx
        for iodb_n, iodb_orig in iodb_norm.items():
            if iodb_n in rev_hdr:
                col_match[iodb_orig] = rev_hdr[iodb_n]
            else:
                for h_txt, h_col in rev_hdr.items():
                    if (iodb_n in h_txt or h_txt in iodb_n) and len(iodb_n) >= 3:
                        if iodb_orig not in col_match:
                            col_match[iodb_orig] = h_col
                        break

        # ── Snapshot prototype block (template rows 9-11, ALL columns) ───────
        # proto[row_offset][col_idx] = cell_value_or_formula_string
        proto: dict[int, dict[int, object]] = {i: {} for i in range(ROWS_PER_TAG)}
        for row_off in range(ROWS_PER_TAG):
            src_r = CABLE_SCHEDULE_FIRST_DATA_ROW + row_off
            for row in tmpl_ws.iter_rows(min_row=src_r, max_row=src_r):
                for col_idx, cell in enumerate(row, 1):
                    val = getattr(cell, 'value', None)
                    if val is not None:
                        proto[row_off][col_idx] = val

        # ── Filter IODB: keep rows where JB has a real value (not "" or "-") ─
        jb_str  = df[jb_column].astype(str).str.strip()
        valid   = df[jb_column].notna() & (jb_str != "") & (jb_str != "-")
        if not valid.any():
            return None, None, (
                f"No rows found in IODB where '{jb_column}' has a "
                f"value other than blank or '-'."
            )
        work_df = df[valid].copy()
        work_df[jb_column] = work_df[jb_column].astype(str).str.strip()

        # Natural sort by JB column (ascending, handles mixed alpha-numeric)
        try:
            work_df["_jb_key"] = work_df[jb_column].apply(
                lambda x: [int(t) if t.isdigit() else t.lower()
                           for t in re.split(r'(\d+)', x)]
            )
            work_df = work_df.sort_values("_jb_key").drop(columns=["_jb_key"])
        except Exception:
            work_df = work_df.sort_values(jb_column)

        # ── Build output workbook ─────────────────────────────────────────────
        # Prefer: reload template_bytes in writable mode so the full template
        # layout (merged cells, column widths, header styles) is preserved.
        out_wb = None
        if template_bytes:
            for _opts in (
                dict(keep_vba=False, read_only=False, keep_links=False),
                dict(keep_vba=True,  read_only=False, keep_links=False),
            ):
                try:
                    out_wb = openpyxl.load_workbook(io.BytesIO(template_bytes), **_opts)
                    break
                except Exception:
                    continue

        if out_wb is None and not getattr(template_wb, 'read_only', True):
            # Template already loaded writable — deep-copy it
            try:
                out_wb = _copy_workbook(template_wb)
            except Exception:
                pass

        if out_wb is None:
            # Last resort: fresh workbook + reconstruct headers from prototype data
            out_wb = openpyxl.Workbook()
            out_ws_new = out_wb.active
            out_ws_new.title = target_sheet
            for h_off in range(CABLE_SCHEDULE_FIRST_DATA_ROW - 1):
                h_row = h_off + 1
                for row in tmpl_ws.iter_rows(min_row=h_row, max_row=h_row):
                    for cell in row:
                        if cell.value is not None:
                            oc = out_ws_new.cell(row=h_row, column=cell.column,
                                                 value=cell.value)
                            oc.font      = _CS_FONT_BOLD
                            oc.alignment = _CS_ALIGN

        out_ws = out_wb[target_sheet]

        # Clear all pre-existing data rows (≥ FIRST_DATA_ROW)
        for r_idx in range(CABLE_SCHEDULE_FIRST_DATA_ROW, out_ws.max_row + 1):
            for row in out_ws.iter_rows(min_row=r_idx, max_row=r_idx):
                for cell in row:
                    try:
                        cell.value = None
                    except Exception:
                        pass

        # ── Write data rows ───────────────────────────────────────────────────
        current_row = CABLE_SCHEDULE_FIRST_DATA_ROW
        sl_no       = 1
        jb_order    = list(dict.fromkeys(work_df[jb_column]))  # preserves sort order, deduped
        total       = len(jb_order)

        for jb_idx, jb_name in enumerate(jb_order):
            if progress_callback:
                progress_callback(jb_idx, total)

            tag_entries = work_df[work_df[jb_column] == jb_name].to_dict('records')

            for entry in tag_entries:
                # 1. Copy the prototype block (all 3 rows) with formula adjustment
                for row_off in range(ROWS_PER_TAG):
                    dst_r = current_row + row_off
                    src_r = CABLE_SCHEDULE_FIRST_DATA_ROW + row_off

                    for col_idx, proto_val in proto[row_off].items():
                        try:
                            c = out_ws.cell(row=dst_r, column=col_idx)
                        except Exception:
                            continue
                        if isinstance(proto_val, str) and proto_val.startswith('='):
                            c.value = _adjust_formula_row(proto_val, src_r, dst_r)
                        else:
                            c.value = proto_val
                        c.font      = _CS_FONT
                        c.alignment = _CS_ALIGN

                # 2. Fill IODB values in main row (row_off=0); preserve formulas
                dst_main = current_row

                # Serial number in column A (if not a formula)
                sn = out_ws.cell(row=dst_main, column=1)
                if not (isinstance(sn.value, str) and sn.value.startswith('=')):
                    sn.value     = sl_no
                    sn.font      = _CS_FONT
                    sn.alignment = _CS_ALIGN

                for iodb_col, tmpl_col_idx in col_match.items():
                    try:
                        c = out_ws.cell(row=dst_main, column=tmpl_col_idx)
                    except Exception:
                        continue
                    if isinstance(c.value, str) and c.value.startswith('='):
                        continue  # keep formula
                    raw = entry.get(iodb_col)
                    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                        c.value = "TBA"
                    elif str(raw).strip() in ("", "nan"):
                        c.value = "TBA"
                    else:
                        c.value = str(raw).strip() if isinstance(raw, str) else raw
                    c.font      = _CS_FONT
                    c.alignment = _CS_ALIGN

                sl_no      += 1
                current_row += ROWS_PER_TAG

        if progress_callback:
            progress_callback(total, total)
        # Compact tag blocks: detect blocks (main row contains TAG) and
        # move them upward so blocks are contiguous with no interleaving
        # empty rows. This preserves the 3-row-per-tag structure while
        # removing stray blank spacer rows introduced by some templates.
        try:
            # Find tag column index from header rows
            tag_col_idx = None
            for row in tmpl_ws.iter_rows(min_row=CABLE_SCHEDULE_HEADER_ROW, max_row=CABLE_SCHEDULE_HEADER_ROW + 1):
                for col_idx, cell in enumerate(row, 1):
                    val = getattr(cell, 'value', None)
                    if val and isinstance(val, str):
                        v = val.strip().lower()
                        if 'tag' in v and ('no' in v or 'number' in v):
                            tag_col_idx = col_idx
                            break
                if tag_col_idx:
                    break
            if tag_col_idx is None:
                tag_col_idx = 2

            # Detect block starts by scanning for non-empty tag cell
            block_starts: list[int] = []
            r = CABLE_SCHEDULE_FIRST_DATA_ROW
            last_row = out_ws.max_row
            while r <= last_row:
                tag_val = out_ws.cell(row=r, column=tag_col_idx).value
                if tag_val is not None and str(tag_val).strip() != '':
                    block_starts.append(r)
                    r += ROWS_PER_TAG
                else:
                    r += 1

            # Move blocks upward to eliminate gaps
            dest = CABLE_SCHEDULE_FIRST_DATA_ROW
            max_col = out_ws.max_column
            for src in block_starts:
                if src != dest:
                    for off in range(ROWS_PER_TAG):
                        for cidx in range(1, max_col + 1):
                            src_cell = out_ws.cell(row=src + off, column=cidx)
                            dst_cell = out_ws.cell(row=dest + off, column=cidx)
                            dst_cell.value = src_cell.value
                            try:
                                dst_cell.font = _CS_FONT
                                dst_cell.alignment = _CS_ALIGN
                            except Exception:
                                pass
                            src_cell.value = None
                dest += ROWS_PER_TAG
        except Exception:
            # non-fatal
            pass

        out_bytes = workbook_to_bytes(out_wb)
        return out_bytes, "Cable_Schedule.xlsx", None

    except Exception as e:
        return None, None, f"Cable Schedule generation failed: {e}"


def _read_formula_block(
    ws,
    start_row: int,
    num_rows: int,
    cols: tuple[str, ...] = ("AK", "AL", "AM", "AN"),
) -> dict[tuple[str, int], object]:
    """
    Read a rectangular block of cells from the template worksheet.
    Returns {(col_letter, row_offset): cell_value} for the given columns
    across `num_rows` rows starting at `start_row`.
    """
    result: dict[tuple[str, int], object] = {}
    for row_offset in range(num_rows):
        row_idx = start_row + row_offset
        for col_letter in cols:
            col_idx = column_index_from_string(col_letter)
            # iter_rows works in read-only mode; use direct cell access
            val = None
            for row in ws.iter_rows(
                min_row=row_idx, max_row=row_idx,
                min_col=col_idx, max_col=col_idx,
                values_only=True,
            ):
                val = row[0]
            result[(col_letter, row_offset)] = val
    return result


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    """
    Read a header row and return {normalised_header_text: column_index (1-based)}.
    """
    result = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            key = str(cell.value).strip().upper()
            result[key] = cell.column
    return result


def _write_cable_row(
    ws,
    row_idx: int,
    tag_entry: dict,
    header_map: dict,
    tag_col: str,
    jb_col: str,
    sl_no: int = 0,
):
    """
    Write one tag's data into the cable schedule worksheet at row_idx.
    Maps IODB column names → template header names using CABLE_SCHEDULE_COL_MAP.
    For SPARE entries, writes SPARE in the TAG NUMBER column and the JB in JB TAG No.
    """
    tag_value = tag_entry.get(tag_col, "")
    is_spare = str(tag_value).strip().upper() == "SPARE"

    if is_spare:
        for template_hdr, col_idx in header_map.items():
            if "TAG" in template_hdr and "NUMBER" in template_hdr:
                ws.cell(row=row_idx, column=col_idx).value = "SPARE"
            elif "JB TAG" in template_hdr or ("JB" in template_hdr and "NO" in template_hdr):
                ws.cell(row=row_idx, column=col_idx).value = tag_entry.get(jb_col, "")
        # Write serial number
        sl_col = header_map.get("SL. NO.  ") or header_map.get("SL. NO.") or 1
        if sl_no:
            ws.cell(row=row_idx, column=sl_col if isinstance(sl_col, int) else 1).value = ""
        return

    # Write serial number in column A
    ws.cell(row=row_idx, column=1).value = sl_no

    for iodb_col, template_col in CABLE_SCHEDULE_COL_MAP.items():
        value = tag_entry.get(iodb_col)
        template_key = template_col.strip().upper()
        col_idx = header_map.get(template_key)
        if col_idx is None:
            col_idx = next(
                (v for k, v in header_map.items() if template_key in k or k in template_key),
                None,
            )
        if col_idx is not None:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell.value = "TBA"
            else:
                cell.value = value


# ─────────────────────────────────────────────────────────────────────────────
# 5. Loop Wiring Generator
# ─────────────────────────────────────────────────────────────────────────────

# The template sheet to use in the Loop Wiring template
LOOP_WIRING_TEMPLATE_SHEET = "AI Instr"
LOOP_WIRING_HEADER_ROW = 18        # Row containing column headers (W→AR)
LOOP_WIRING_DATA_ROW   = 19        # Row where tag values are written
LOOP_WIRING_COL_START  = 23        # Column W
LOOP_WIRING_COL_END    = 44        # Column AR


def generate_loop_wiring(
    lw_df: pd.DataFrame,
    template_wb: openpyxl.Workbook,
    progress_callback=None,
    template_bytes: bytes = None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    For each tag in lw_df, copy the 'AI Instr' template sheet and fill
    row 19 (columns W→AR) with matched IODB values.

    Sheet preservation: copy_worksheet() keeps shapes, formulas, and
    formatting intact. Drawing XML is also re-injected at ZIP level for
    connectors / text-boxes that openpyxl cannot round-trip.

    Returns (bytes, "Loop_Wiring.xlsx", None) on success,
            (None, None, error_string) on failure.
    """
    try:
        # ── 1. Find template sheet with tolerant matching ───────────────────
        def _normalize_name(x: str) -> str:
            if x is None:
                return ""
            # lower, strip, remove non-alphanumeric characters for tolerant compare
            return re.sub(r'[^0-9a-z]', '', str(x).strip().lower())

        target_norm = _normalize_name(LOOP_WIRING_TEMPLATE_SHEET)
        template_sheet_name = None
        # exact tolerant match first
        for s in template_wb.sheetnames:
            if _normalize_name(s) == target_norm:
                template_sheet_name = s
                break
        # then try substring (tolerant) match
        if template_sheet_name is None:
            for s in template_wb.sheetnames:
                if target_norm in _normalize_name(s) or _normalize_name(s) in target_norm:
                    template_sheet_name = s
                    break

        if not template_sheet_name:
            return None, None, (
                f"Loop Wiring template does not contain sheet '{LOOP_WIRING_TEMPLATE_SHEET}'. "
                f"Found: {template_wb.sheetnames}"
            )

        # ── 2. Locate tag column in input DataFrame ──────────────────────────
        tag_col = next(
            (c for c in lw_df.columns
             if c.strip().lower() in ("tag no", "tag number", "tag no.", "tag_nummber")),
            None,
        )
        if tag_col is None:
            tag_col = next(
                (c for c in lw_df.columns if "tag" in c.strip().lower()),
                None,
            )
        if not tag_col:
            return None, None, "No 'Tag Number' column found in Loop Wiring input."

        tmpl_ws = template_wb[template_sheet_name]

        # ── 3. Read headers from row 18, cols W(23)→AR(44) ──────────────────
        header_col_map: dict[str, int] = {}
        for col in range(LOOP_WIRING_COL_START, LOOP_WIRING_COL_END + 1):
            hdr = tmpl_ws.cell(row=LOOP_WIRING_HEADER_ROW, column=col).value
            if hdr is not None:
                header_col_map[str(hdr).strip().lower()] = col

        tags = lw_df[tag_col].dropna().astype(str).str.strip().unique().tolist()
        total = len(tags)

        # ── 4. One copied sheet per tag ──────────────────────────────────────
        for idx, tag in enumerate(tags):
            if progress_callback:
                progress_callback(idx, total)

            new_ws = template_wb.copy_worksheet(tmpl_ws)
            safe_tag = re.sub(r'[\\/*?:\[\]]', '_', tag)
            new_ws.title = f"Loop_{safe_tag}"[:31]

            tag_rows = lw_df[lw_df[tag_col].astype(str).str.strip() == tag]
            if tag_rows.empty:
                continue
            norm_row = {str(k).strip().lower(): v for k, v in tag_rows.iloc[0].items()}

            for hdr_key, col in header_col_map.items():
                value = norm_row.get(hdr_key)
                if value is None:
                    for rk, rv in norm_row.items():
                        if hdr_key in rk or rk in hdr_key:
                            value = rv
                            break
                cell = new_ws.cell(row=LOOP_WIRING_DATA_ROW, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue  # preserve formulas
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    cell.value = "TBA"
                else:
                    cell.value = value

        # ── 5. Remove the original template sheet ────────────────────────────
        del template_wb[template_sheet_name]

        if not template_wb.sheetnames:
            return None, None, "No tag sheets were generated. Check that the input has tags."

        if progress_callback:
            progress_callback(total, total)

        out_bytes = workbook_to_bytes(template_wb)

        # ── 6. Re-inject drawing XML for shapes/connectors at ZIP level ──────
        if template_bytes is not None:
            out_bytes = _inject_drawings_from_template(
                out_bytes, template_bytes, template_sheet_name
            )

        return out_bytes, "Loop_Wiring.xlsx", None

    except Exception as e:
        return None, None, f"Loop Wiring generation failed: {e}"


# ─────────────────────────────────────────────────────────────────────────────
# Utility: deep-copy a workbook
# ─────────────────────────────────────────────────────────────────────────────

def _inject_drawings_from_template(
    out_bytes: bytes,
    tmpl_bytes: bytes,
    template_sheet_name: str,
) -> bytes:
    """
    Post-process the output xlsx (bytes) to inject the drawing XML from the
    matching template sheet into EVERY sheet of the output workbook.

    Works entirely at the ZIP/XML level, bypassing openpyxl's incomplete
    drawing support.  If anything goes wrong the original bytes are returned
    unchanged so the caller always gets a valid file.
    """
    NS_R_TYPE_DRAWING = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
    )
    try:
        from xml.etree import ElementTree as ET

        # ── 1. Locate and extract drawing XML from the template zip ──────────
        drawing_xml_bytes = None
        with zipfile.ZipFile(io.BytesIO(tmpl_bytes)) as tz:
            tz_names = tz.namelist()
            NS_SS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            NS_R  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

            # Find the rId for the target sheet in workbook.xml
            wb_root = ET.fromstring(tz.read('xl/workbook.xml'))
            sheet_rid = None
            for el in wb_root.iter(f'{{{NS_SS}}}sheet'):
                if el.get('name', '').strip().lower() == template_sheet_name.strip().lower():
                    sheet_rid = el.get(f'{{{NS_R}}}id')
                    break
            if sheet_rid is None:
                return out_bytes

            # Resolve rId → worksheet filename through workbook.xml.rels
            wb_rels = ET.fromstring(tz.read('xl/_rels/workbook.xml.rels'))
            sheet_file = None
            for rel in wb_rels:
                if rel.get('Id') == sheet_rid:
                    sheet_file = rel.get('Target')  # e.g. "worksheets/sheet1.xml"
                    break
            if sheet_file is None:
                return out_bytes

            # Find drawing reference in the sheet's rels file
            sheet_basename = sheet_file.split('/')[-1]     # "sheet1.xml"
            rels_path = f'xl/worksheets/_rels/{sheet_basename}.rels'
            if rels_path not in tz_names:
                return out_bytes
            sheet_rels = ET.fromstring(tz.read(rels_path))
            drawing_filename = None
            for rel in sheet_rels:
                if rel.get('Type') == NS_R_TYPE_DRAWING:
                    drawing_filename = rel.get('Target').split('/')[-1]  # "drawing1.xml"
                    break
            if drawing_filename is None:
                return out_bytes

            drawing_zip_path = f'xl/drawings/{drawing_filename}'
            if drawing_zip_path not in tz_names:
                return out_bytes
            drawing_xml_bytes = tz.read(drawing_zip_path)

        if drawing_xml_bytes is None:
            return out_bytes

        # ── 2. Read output zip into memory ───────────────────────────────────
        with zipfile.ZipFile(io.BytesIO(out_bytes)) as oz:
            entries: dict[str, bytes] = {n: oz.read(n) for n in oz.namelist()}

        # Collect sheet files sorted numerically
        sheet_paths = sorted(
            (n for n in entries
             if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')),
            key=lambda x: int(re.search(r'(\d+)', x.split('/')[-1]).group(1))
            if re.search(r'(\d+)', x.split('/')[-1]) else 0,
        )
        if not sheet_paths:
            return out_bytes

        # ── 3. Inject drawing into each sheet ────────────────────────────────
        ct_text = entries.get('[Content_Types].xml', b'').decode('utf-8')
        DRAWING_CT = 'application/vnd.openxmlformats-officedocument.drawing+xml'

        for i, sheet_path in enumerate(sheet_paths, start=1):
            drw_name  = f'drawing_lw{i}.xml'
            drw_zip   = f'xl/drawings/{drw_name}'
            drw_rel   = f'../drawings/{drw_name}'
            ct_part   = f'/xl/drawings/{drw_name}'
            rel_id    = 'rIdLwDrw'

            # Store the drawing XML bytes
            entries[drw_zip] = drawing_xml_bytes

            # Update or create the sheet's rels file
            sht_base  = sheet_path.split('/')[-1]           # "sheet1.xml"
            rels_path = f'xl/worksheets/_rels/{sht_base}.rels'
            if rels_path in entries:
                rels_text = entries[rels_path].decode('utf-8')
                if NS_R_TYPE_DRAWING not in rels_text:
                    new_rel = (
                        f'<Relationship Id="{rel_id}" '
                        f'Type="{NS_R_TYPE_DRAWING}" '
                        f'Target="{drw_rel}"/>'
                    )
                    rels_text = rels_text.replace(
                        '</Relationships>', new_rel + '</Relationships>'
                    )
                    entries[rels_path] = rels_text.encode('utf-8')
                else:
                    rel_id = None  # drawing rel already present; extract its id
            else:
                rels_content = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    f'<Relationship Id="{rel_id}" '
                    f'Type="{NS_R_TYPE_DRAWING}" '
                    f'Target="{drw_rel}"/>'
                    '</Relationships>'
                )
                entries[rels_path] = rels_content.encode('utf-8')

            # Add <drawing r:id="..."/> to the sheet XML if not already there
            sht_xml = entries[sheet_path].decode('utf-8')
            if '<drawing' not in sht_xml and rel_id is not None:
                ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                if 'xmlns:r=' not in sht_xml:
                    sht_xml = sht_xml.replace(
                        '<worksheet ', f'<worksheet xmlns:r="{ns_r}" ', 1
                    )
                sht_xml = sht_xml.replace(
                    '</worksheet>',
                    f'<drawing r:id="{rel_id}"/></worksheet>',
                )
                entries[sheet_path] = sht_xml.encode('utf-8')

            # Register in [Content_Types].xml
            if ct_part not in ct_text:
                ct_text = ct_text.replace(
                    '</Types>',
                    f'<Override PartName="{ct_part}" ContentType="{DRAWING_CT}"/></Types>',
                )

        entries['[Content_Types].xml'] = ct_text.encode('utf-8')

        # ── 4. Rebuild the zip ────────────────────────────────────────────────
        out_buf = io.BytesIO()
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in entries.items():
                zout.writestr(name, data)
        out_buf.seek(0)
        return out_buf.read()

    except Exception:
        return out_bytes  # always return a valid file even if injection fails


def _copy_workbook(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Serialize a workbook to bytes and reload it, achieving a deep copy.

    Image-ref fix: openpyxl's Image._data() closes img.ref (a BytesIO) after
    reading image bytes during wb.save().  That means a second call to
    _copy_workbook on the same wb would fail with "I/O operation on closed file".
    After each save we therefore restore wb's image refs from the freshly-loaded
    copy's independent BytesIO objects so every subsequent call works correctly.
    """
    buf = io.BytesIO()
    wb.save(buf)           # NOTE: closes every img.ref on wb
    raw = buf.getvalue()   # full bytes before seeking
    buf.seek(0)
    new_wb = load_workbook(buf)
    new_wb._copybuf = buf  # keep buf alive for new_wb's own save()

    # Restore img.refs on original wb (and give new_wb fresh independent refs)
    # so that (a) the next _copy_workbook call on wb works and (b) new_wb.save()
    # can write its images without a closed-file error.
    for ws_orig, ws_new in zip(wb.worksheets, new_wb.worksheets):
        orig_imgs = getattr(ws_orig, '_images', [])
        new_imgs  = getattr(ws_new,  '_images', [])
        for o_img, n_img in zip(orig_imgs, new_imgs):
            try:
                # getvalue() returns all bytes regardless of current stream pos
                img_bytes = (n_img.ref.getvalue()
                             if hasattr(n_img.ref, 'getvalue')
                             else (n_img.ref.seek(0) or n_img.ref.read()))
                o_img.ref = io.BytesIO(img_bytes)   # fresh ref for next copy
                n_img.ref = io.BytesIO(img_bytes)   # fresh ref for this copy's save
            except Exception:
                pass

    return new_wb


# ─────────────────────────────────────────────────────────────────────────────
# 6. IODB Validation Engine
# ─────────────────────────────────────────────────────────────────────────────

# Optional spell-check support — degrades gracefully if library not installed
try:
    from spellchecker import SpellChecker as _SpellChecker
    _SPELL_AVAILABLE = True
except ImportError:
    _SPELL_AVAILABLE = False

# TAG NO substring → expected Instrument Type (longer keys first to prevent
# shorter keys matching before longer ones, e.g. "FIT" before "DFIT")
_TAG_TYPE_MAP: dict[str, str] = {
    "DFIT":  "DIFFERENTIAL FLOW INDICATING TRANSMITTER",
    "DPIT":  "DIFFERENTIAL PRESSURE INDICATING TRANSMITTER",
    "DLIT":  "DIFFERENTIAL LEVEL INDICATING TRANSMITTER",
    "FIT":   "FLOW INDICATING TRANSMITTER",
    "LIT":   "LEVEL INDICATING TRANSMITTER",
    "TIT":   "TEMPERATURE INDICATING TRANSMITTER",
    "PIT":   "PRESSURE INDICATING TRANSMITTER",
    "AIT":   "ANALYZER INDICATING TRANSMITTER",
    "DPG":   "DIFFERENTIAL PRESSURE GAUGE",
    "LG":    "LEVEL GAUGE",
    "FG":    "FLOW GAUGE",
    "PG":    "PRESSURE GAUGE",
    "TG":    "TEMPERATURE GAUGE",
}
_TAG_KEYS_SORTED = sorted(_TAG_TYPE_MAP, key=len, reverse=True)

# Cell fill colours for highlighted output
_VAL_FILL_EMPTY  = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")  # light red   → empty/mandatory
_VAL_FILL_LOGIC  = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # amber       → logic rules
_VAL_FILL_SPELL  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue  → spelling
_VAL_FILL_FIXED  = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")  # light green → auto-corrected

_VAL_HDR_FILL  = PatternFill(start_color="1B5EA7", end_color="1B5EA7", fill_type="solid")
_VAL_ROW_FILL_A = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")
_VAL_ROW_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


# ── Low-level helpers ────────────────────────────────────────────────────────

def _val_addr(col_idx: int, row_idx: int) -> str:
    """Return Excel cell address for 1-based col/row indices (e.g. 'B5')."""
    return f"{get_column_letter(col_idx)}{row_idx}"


def _val_norm(val) -> str:
    """Strip + uppercase a value; NaN/None returns empty string."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip().upper()


def _val_empty(val) -> bool:
    """True when value is None, NaN, or blank string."""
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    return str(val).strip() == ""


def _val_find_col(df: pd.DataFrame, *keywords: str) -> str | None:
    """
    Find the first matching column in df.
    Strategy: exact match (case-insensitive) → substring match.
    """
    kws = [k.upper() for k in keywords]
    cols_map = {str(c).strip().upper(): c for c in df.columns}
    for kw in kws:                            # exact first
        if kw in cols_map:
            return cols_map[kw]
    for kw in kws:                            # substring fallback
        for cu, co in cols_map.items():
            if kw in cu:
                return co
    return None


def _val_find_col_exact(df: pd.DataFrame, *keywords: str) -> str | None:
    """Find a column by EXACT match only (case-insensitive). No substring."""
    kws = {k.upper() for k in keywords}
    for col in df.columns:
        if str(col).strip().upper() in kws:
            return col
    return None


def _val_col_idx(col_names: list[str], col_name: str) -> int:
    """Return 1-based column index; falls back to 1 if not found."""
    try:
        return col_names.index(col_name) + 1
    except ValueError:
        return 1


def _val_parse_range(val) -> tuple[float, float] | None:
    """
    Parse 'min-max' or 'min to max' strings (handles negative numbers).
    Returns (min_val, max_val) or None when parsing fails.
    """
    if _val_empty(val):
        return None
    s = str(val).strip()
    if s in ("-", "N/A", "NA", "TBA", ""):
        return None
    m = re.search(r'(-?\d+(?:\.\d+)?)\s*(?:[-\u2013~]|to)\s*(-?\d+(?:\.\d+)?)', s)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None


# ── Main validation function ─────────────────────────────────────────────────

def generate_iodb_validation(
    df: pd.DataFrame,
    wb: openpyxl.Workbook,
    auto_correct_spelling: bool = False,
) -> tuple[bytes | None, bytes | None, list[dict], str | None]:
    """
    Run all 12 IODB validation rules on *df* and return:

      error_log_bytes   — Formatted Excel validation report
      highlighted_bytes — Original workbook with error cells highlighted
                          (red = empty, amber = logic, blue = spelling,
                           green = auto-corrected spelling when requested)
      errors            — List of error dicts for Streamlit display
      error_message     — None on success; string description on failure

    Assumes:  df row i  →  workbook row (i + 2)
              (row 1 = header, row 2 = first data row)

    getCellAddress(col_idx, row_idx) implemented as _val_addr().
    """
    try:
        errors: list[dict] = []
        col_names = list(df.columns)

        # ── Locate key columns ────────────────────────────────────────────────
        sno_col    = _val_find_col(df, "S.NO", "S NO", "SNO", "S.NO.", "SERIAL NO", "SR NO", "SL NO")
        tag_col    = _val_find_col(df, "TAG NO", "TAG NUMBER", "TAG_NO", "TAG_NUMBER", "TAG")
        area_col   = _val_find_col(df, "AREA CLASSIFICATION")
        isnis_col  = _val_find_col(df, "IS/NIS TYPE", "IS/NIS", "IS NIS", "IS_NIS", "ISNIS")
        scope_col  = _val_find_col(df, "SCOPE OF SUPPLY", "SUPPLY SCOPE", "SCOPE")
        vendor_col = _val_find_col(df, "VENDOR PACKAGE", "VENDOR PKG", "VENDOR PACKAGE NAME")
        itype_col  = _val_find_col(df, "INSTRUMENT TYPE", "INST TYPE")
        power_col  = _val_find_col(df, "POWER SUPPLY")
        wire_col   = _val_find_col(df, "2 WIRE/4 WIRE", "2WIRE/4WIRE", "2WIRE 4WIRE", "WIRE TYPE")
        signal_col = _val_find_col(df, "SIGNAL I/O TYPE", "SIGNAL TYPE", "I/O TYPE", "IO TYPE", "SIGNAL IO TYPE")
        subsys_col = _val_find_col(df, "SUB SYSTEM", "SUBSYSTEM", "SUB-SYSTEM")
        jb_col     = _val_find_col(df, "JUNCTION BOX", "JB", "J/B")
        calr_col   = _val_find_col(df, "CALIBRATION RANGE", "CAL RANGE", "CALIB RANGE")
        instr_col  = _val_find_col(df, "INST RANGE", "INSTRUMENT RANGE", "INSTR RANGE")
        calu_col   = _val_find_col(df, "CALIBRATION UNIT", "CAL UNIT", "CALIB UNIT")
        instu_col  = _val_find_col(df, "INST RANGE UNIT", "INST UNIT", "INSTRUMENT UNIT", "INST RANGE UNITS")
        fail_col   = _val_find_col(df, "FAIL ACTION", "FAIL SAFE", "FAILURE ACTION")
        # Alarm columns: use EXACT match to avoid "FLOW" matching "LOW" etc.
        ll_col = _val_find_col_exact(df, "LOW LOW", "LOLO", "LL", "L.L.", "LOW LOW ALARM", "LLLL")
        l_col  = _val_find_col_exact(df, "LOW", "LO", "L", "LOW ALARM")
        h_col  = _val_find_col_exact(df, "HIGH", "HI", "H", "HIGH ALARM")
        hh_col = _val_find_col_exact(df, "HIGH HIGH", "HIHI", "HH", "H.H.", "HIGH HIGH ALARM", "HHHH")
        # Prevent l_col / h_col from aliasing the LL/HH column
        if l_col and l_col == ll_col:
            l_col = None
        if h_col and h_col == hh_col:
            h_col = None

        SKIP_EMPTY = {"STATUS", "REMARKS"}

        # ── Error builder ─────────────────────────────────────────────────────
        def _err(i: int, col: str, msg: str, rule: int) -> dict:
            """Build a standardised error dict for row i (0-based df index)."""
            row_idx = i + 2                                     # Excel row
            ci      = _val_col_idx(col_names, col)
            sno_v   = "UNKNOWN"
            tag_v   = "UNKNOWN"
            if sno_col:
                sv = df.iloc[i].get(sno_col, "")
                if not _val_empty(sv):
                    sno_v = str(sv).strip()
            if tag_col:
                tv = df.iloc[i].get(tag_col, "")
                if not _val_empty(tv):
                    tag_v = str(tv).strip()
            return {
                "row":     row_idx,
                "sno":     sno_v,
                "tag":     tag_v,
                "column":  col,
                "cell":    _val_addr(ci, row_idx),
                "message": msg,
                "rule":    rule,
            }

        # ══════════════════════════════════════════════════════════════════════
        # RULE 1 — Empty cell validation (except STATUS and REMARKS)
        # ══════════════════════════════════════════════════════════════════════
        for col in col_names:
            if col.strip().upper() in SKIP_EMPTY:
                continue
            for i in range(len(df)):
                if _val_empty(df.iloc[i][col]):
                    errors.append(_err(i, col, "Cell cannot be empty", 1))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 2 — AREA CLASSIFICATION vs IS/NIS TYPE
        # ══════════════════════════════════════════════════════════════════════
        if area_col and isnis_col:
            for i in range(len(df)):
                area  = _val_norm(df.iloc[i][area_col])
                isnis = _val_norm(df.iloc[i][isnis_col])
                if area == "HAZARDOUS" and isnis != "IS":
                    errors.append(_err(i, isnis_col,
                        "Instrument under Hazardous area should be IS type only", 2))
                elif area == "SAFE AREA" and isnis != "NIS":
                    errors.append(_err(i, isnis_col,
                        "Instrument under Safe area should be NIS type only", 2))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 3 — SCOPE OF SUPPLY vs VENDOR PACKAGE
        # ══════════════════════════════════════════════════════════════════════
        if scope_col and vendor_col:
            for i in range(len(df)):
                scope  = _val_norm(df.iloc[i][scope_col])
                vendor = df.iloc[i][vendor_col]
                if scope != "WABAG" and (_val_empty(vendor) or _val_norm(vendor) == "-"):
                    errors.append(_err(i, vendor_col,
                        "Instrument is under Vendor package scope so Vendor package name should be mentioned", 3))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 4 — TAG NO keyword vs INSTRUMENT TYPE (substring mapping)
        # ══════════════════════════════════════════════════════════════════════
        if tag_col and itype_col:
            for i in range(len(df)):
                tag_v  = _val_norm(df.iloc[i][tag_col])
                itype  = _val_norm(df.iloc[i][itype_col])
                if not tag_v or not itype:
                    continue
                for key in _TAG_KEYS_SORTED:
                    if key in tag_v:
                        expected = _TAG_TYPE_MAP[key].upper()
                        if itype != expected:
                            errors.append(_err(i, itype_col,
                                f"Instrument Type mismatch with TAG NO "
                                f"(TAG contains '{key}', expected: {_TAG_TYPE_MAP[key]})", 4))
                        break

        # ══════════════════════════════════════════════════════════════════════
        # RULE 4B — TAG NO must be fully uppercase
        # ══════════════════════════════════════════════════════════════════════
        if tag_col:
            for i in range(len(df)):
                tv = df.iloc[i][tag_col]
                if not _val_empty(tv):
                    ts = str(tv).strip()
                    if ts != ts.upper():
                        errors.append(_err(i, tag_col,
                            "Tag name shall be in capital letters", 4))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 5 — POWER SUPPLY vs 2 WIRE / 4 WIRE
        # ══════════════════════════════════════════════════════════════════════
        if power_col and wire_col:
            for i in range(len(df)):
                power = _val_norm(df.iloc[i][power_col])
                wire  = _val_norm(df.iloc[i][wire_col])
                if power not in ("24V DC", "-", "") and wire != "4 WIRE - EXT POWER":
                    errors.append(_err(i, wire_col,
                        "2 wire instrument is Loop powered and 24 V DC and "
                        "4 wire is External powered", 5))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 6 — SIGNAL I/O TYPE vs JUNCTION BOX
        # ══════════════════════════════════════════════════════════════════════
        if signal_col and jb_col:
            jb_msg = ("All AI signals should be connected to AJB and "
                      "All DI or DO signals should be connected to DJB")
            for i in range(len(df)):
                signal = _val_norm(df.iloc[i][signal_col])
                jb     = _val_norm(df.iloc[i][jb_col])
                subsys = _val_norm(df.iloc[i][subsys_col]) if subsys_col else ""
                # Only apply when instrument is NOT routed via a sub-system
                if subsys not in ("-", ""):
                    continue
                if signal == "AI" and "AJB" not in jb:
                    errors.append(_err(i, jb_col, jb_msg, 6))
                elif signal in ("DI", "DO") and "DJB" not in jb:
                    errors.append(_err(i, jb_col, jb_msg, 6))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 7 — CALIBRATION RANGE must not exceed INSTRUMENT RANGE
        # ══════════════════════════════════════════════════════════════════════
        if calr_col and instr_col:
            for i in range(len(df)):
                cal_r  = _val_parse_range(df.iloc[i][calr_col])
                inst_r = _val_parse_range(df.iloc[i][instr_col])
                if cal_r and inst_r and cal_r[1] > inst_r[1]:
                    errors.append(_err(i, calr_col,
                        f"Calibration range ({cal_r[0]}-{cal_r[1]}) exceeds "
                        f"Instrument range ({inst_r[0]}-{inst_r[1]})", 7))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 8 — CALIBRATION UNIT must match INST RANGE UNIT
        # ══════════════════════════════════════════════════════════════════════
        if calu_col and instu_col:
            for i in range(len(df)):
                cu = _val_norm(df.iloc[i][calu_col])
                iu = _val_norm(df.iloc[i][instu_col])
                if cu and iu and cu != iu:
                    errors.append(_err(i, calu_col,
                        f"Calibration Range unit '{cu}' should be same as "
                        f"Instrument Range unit '{iu}'", 8))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 9 — FAIL ACTION must be FO, FC, or LAST POS
        # ══════════════════════════════════════════════════════════════════════
        if fail_col:
            _valid_fa = {"FO", "FC", "LAST POS", "-"}
            for i in range(len(df)):
                fa = _val_norm(df.iloc[i][fail_col])
                if fa and fa not in _valid_fa:
                    errors.append(_err(i, fail_col,
                        f"Fail action '{fa}' is invalid — allowed: "
                        "FO (Fail Open), FC (Fail Close), LAST POS", 9))

        # ══════════════════════════════════════════════════════════════════════
        # RULE 10 — ALARM SETPOINT ORDER: LL < L < H < HH
        # ══════════════════════════════════════════════════════════════════════
        if all(c is not None for c in (ll_col, l_col, h_col, hh_col)):
            for i in range(len(df)):
                try:
                    vals: list[float] = []
                    skip = False
                    for ac in (ll_col, l_col, h_col, hh_col):
                        v = df.iloc[i][ac]
                        if _val_empty(v) or str(v).strip() in ("-", "N/A", "NA"):
                            skip = True
                            break
                        vals.append(float(str(v).strip()))
                    if not skip and len(vals) == 4:
                        if not (vals[0] < vals[1] < vals[2] < vals[3]):
                            errors.append(_err(i, ll_col,
                                f"Alarm set points are not in correct order — "
                                f"LL={vals[0]}, L={vals[1]}, H={vals[2]}, HH={vals[3]} "
                                f"(required: LL < L < H < HH)", 10))
                except (ValueError, TypeError):
                    pass

        # ══════════════════════════════════════════════════════════════════════
        # RULE 11 — S.NO must be strictly ascending
        # ══════════════════════════════════════════════════════════════════════
        if sno_col:
            prev_sno: float | None = None
            for i in range(len(df)):
                sv = df.iloc[i][sno_col]
                if _val_empty(sv):
                    continue
                try:
                    curr = float(str(sv).strip())
                    if prev_sno is not None and curr <= prev_sno:
                        errors.append(_err(i, sno_col,
                            f"S.NO is not in ascending order "
                            f"(current={curr}, previous={prev_sno})", 11))
                    prev_sno = curr
                except ValueError:
                    pass

        # ══════════════════════════════════════════════════════════════════════
        # RULE 12 — SPELL CHECK (requires pyspellchecker)
        # ══════════════════════════════════════════════════════════════════════
        # spell_corrections maps (row_idx, col_name) → [(wrong, correct), ...]
        # used for the auto-correct feature in the highlighted output.
        spell_corrections: dict[tuple[int, str], list[tuple[str, str]]] = {}

        if _SPELL_AVAILABLE:
            spell = _SpellChecker()
            # Load domain-specific terms to prevent false positives
            spell.word_frequency.load_words([
                "wabag", "iodb", "transmitter", "instrumentation", "calibration",
                "commissioning", "setpoint", "flowmeter", "thermowell",
                "manometer", "orifice", "actuator", "positioner", "solenoid",
                "junction", "datasheet", "annexure", "nozzle", "reducer",
                "flanged", "threaded", "diaphragm", "capillary", "impulse",
                "pulsation", "dampener", "manifold", "multipoint", "thermocouple",
                "rtd", "pvdf", "ptfe", "hastelloy", "inconel",
            ])
            # Skip columns that are code/numeric fields
            _SKIP_SPELL = {
                "S.NO", "S NO", "SNO", "TAG NO", "TAG NUMBER", "TAG",
                "JUNCTION BOX", "JB", "J/B", "POWER SUPPLY",
                "SIGNAL I/O TYPE", "SIGNAL TYPE", "I/O TYPE", "IO TYPE",
                "CALIBRATION RANGE", "INST RANGE", "CALIBRATION UNIT",
                "INST RANGE UNIT", "INST UNIT", "INSTRUMENT UNIT",
                "IS/NIS TYPE", "IS/NIS", "IS NIS", "FAIL ACTION",
                "LOW LOW", "LOW", "HIGH", "HIGH HIGH", "LL", "L", "H", "HH",
                "2 WIRE/4 WIRE", "SCOPE OF SUPPLY", "VENDOR PACKAGE",
                "AREA CLASSIFICATION", "P&ID NO", "LOOP NO", "LOOP NUMBER",
                "STATUS", "REMARKS", "SUB SYSTEM", "SUBSYSTEM",
            }
            for col in col_names:
                if col.strip().upper() in _SKIP_SPELL:
                    continue
                for i in range(len(df)):
                    v = df.iloc[i][col]
                    if _val_empty(v):
                        continue
                    # Only check plain alphabetic words (4+ chars, not all-caps,
                    # no digits) to avoid flagging codes/abbreviations/numbers.
                    words = re.findall(r'\b[a-zA-Z]{4,}\b', str(v))
                    words = [
                        w for w in words
                        if not w.isupper()
                        and not any(ch.isdigit() for ch in w)
                    ]
                    if not words:
                        continue
                    misspelled = spell.unknown(words)
                    for wrong in sorted(misspelled):
                        correct = spell.correction(wrong)
                        if correct and correct.lower() != wrong.lower():
                            errors.append(_err(i, col,
                                f"Possible spelling: '{wrong}' → "
                                f"suggested: '{correct}'", 12))
                            row_idx = i + 2
                            spell_corrections.setdefault((row_idx, col), []).append(
                                (wrong, correct)
                            )

        # ══════════════════════════════════════════════════════════════════════
        # BUILD HIGHLIGHTED WORKBOOK
        # Apply fills to every error cell; optionally auto-correct spelling.
        # ══════════════════════════════════════════════════════════════════════
        ws_hl = wb.worksheets[0]
        for e in errors:
            ci   = _val_col_idx(col_names, e["column"])
            cell = ws_hl.cell(row=e["row"], column=ci)
            if e["rule"] == 1:
                cell.fill = _VAL_FILL_EMPTY
            elif e["rule"] == 12:
                cell.fill = _VAL_FILL_SPELL
            else:
                cell.fill = _VAL_FILL_LOGIC

        if auto_correct_spelling and spell_corrections:
            for (row_idx, col), pairs in spell_corrections.items():
                ci   = _val_col_idx(col_names, col)
                cell = ws_hl.cell(row=row_idx, column=ci)
                if cell.value and not str(cell.value).startswith("="):
                    new_val = str(cell.value)
                    for wrong, correct in pairs:
                        new_val = re.sub(
                            r'\b' + re.escape(wrong) + r'\b',
                            correct,
                            new_val,
                            flags=re.IGNORECASE,
                        )
                    cell.value = new_val
                    cell.fill = _VAL_FILL_FIXED   # green = corrected

        highlighted_bytes = workbook_to_bytes(wb)

        # ══════════════════════════════════════════════════════════════════════
        # BUILD ERROR LOG EXCEL
        # Sheet 1: full error table  |  Sheet 2: summary
        # ══════════════════════════════════════════════════════════════════════
        from collections import Counter

        log_wb = openpyxl.Workbook()
        rpt_ws = log_wb.active
        rpt_ws.title = "Validation Report"

        _hfont   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        _dfont   = Font(name="Calibri", size=11)
        _c_aln   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _l_aln   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        log_cols   = ["Row #", "S.NO", "TAG NO", "Column", "Cell", "Rule", "Error Message"]
        log_widths = [8, 12, 24, 28, 8, 10, 65]

        for ci, (hdr_txt, w) in enumerate(zip(log_cols, log_widths), start=1):
            c = rpt_ws.cell(row=1, column=ci, value=hdr_txt)
            c.fill = _VAL_HDR_FILL
            c.font = _hfont
            c.alignment = _c_aln
            rpt_ws.column_dimensions[get_column_letter(ci)].width = w
        rpt_ws.row_dimensions[1].height = 22

        for ri, e in enumerate(errors, start=2):
            rfill = _VAL_ROW_FILL_A if ri % 2 == 0 else _VAL_ROW_FILL_B
            vals  = [
                e["row"], e["sno"], e["tag"],
                e["column"], e["cell"],
                f"Rule {e['rule']}", e["message"],
            ]
            for ci2, val in enumerate(vals, start=1):
                c = rpt_ws.cell(row=ri, column=ci2, value=val)
                c.font  = _dfont
                c.fill  = rfill
                c.alignment = _c_aln if ci2 <= 6 else _l_aln
            rpt_ws.row_dimensions[ri].height = 20

        # Freeze header row
        rpt_ws.freeze_panes = "A2"

        # Summary sheet
        sum_ws = log_wb.create_sheet("Summary")
        sum_ws.column_dimensions["A"].width = 42
        sum_ws.column_dimensions["B"].width = 10

        title_font = Font(name="Calibri", size=14, bold=True, color="1B5EA7")
        h2_font    = Font(name="Calibri", size=11, bold=True)
        d_font     = Font(name="Calibri", size=11)

        sum_ws.cell(row=1, column=1, value="IODB Validation Summary").font = title_font
        sum_ws.cell(row=2, column=1, value=f"Total Errors:      {len(errors)}").font = h2_font
        rows_with_err = len({e["row"] for e in errors})
        sum_ws.cell(row=3, column=1, value=f"Rows with Errors:  {rows_with_err}").font = d_font

        _rule_labels = {
            1:  "Empty Cell Validation",
            2:  "Area Classification vs IS/NIS",
            3:  "Scope of Supply vs Vendor Package",
            4:  "TAG NO vs Instrument Type / Capitalisation",
            5:  "Power Supply Logic",
            6:  "Signal Type vs Junction Box",
            7:  "Range Validation",
            8:  "Unit Match",
            9:  "Fail Action",
            10: "Alarm Setpoint Order",
            11: "S.NO Order",
            12: "Spelling Check",
        }
        rule_ct = Counter(e["rule"] for e in errors)
        sum_ws.cell(row=5, column=1, value="Errors by Rule").font = h2_font
        sum_ws.cell(row=5, column=2, value="Count").font = h2_font
        for sri, rn in enumerate(sorted(rule_ct), start=6):
            sum_ws.cell(row=sri, column=1,
                        value=f"Rule {rn}: {_rule_labels.get(rn, '')}").font = d_font
            sum_ws.cell(row=sri, column=2, value=rule_ct[rn]).font = d_font

        log_bytes = workbook_to_bytes(log_wb)
        return log_bytes, highlighted_bytes, errors, None

    except Exception as exc:
        import traceback
        return None, None, [], f"Validation failed: {exc}\n{traceback.format_exc()}"
