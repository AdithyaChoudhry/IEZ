"""
generate_datasheet_cli.py
=========================
Interactive CLI datasheet generator for WABAG LT Non-Contact Radar Type instruments.

Template color scheme (LT Non contact Radar Type 3.xls):
  GREEN  RGB(51,153,102)   → IODB auto-fill; prompt user if IODB value is empty
  RED    RGB(255,0,0)      → Vendor/spec value; use EG-sheet default if set, else prompt user
  GREY   RGB(239,239,239)  → Project header (CLIENT/CONSULTANT/PROJECT/LOCATION); always prompt
  VIOLET RGB(51,51,153)    → PURCHASE (Make/Model/Qty) and NOTES; always prompt

Output Excel: identical layout but ALL background colours stripped → clean white datasheet.

Usage:
    python generate_datasheet_cli.py

Files expected (same directory as this script or configured below):
    IODB.xlsx                           ← instrument database
    LT Non contact Radar Type 3.xls    ← blank template
    LT Non contact Radar Type-R0.xls   ← contains LT-RADAR-EG example values
"""
from __future__ import annotations

import os
import re
import sys

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
IODB_PATH     = os.path.join(BASE_DIR, "IODB.xlsx")
TEMPLATE_PATH = os.path.join(BASE_DIR, "LT Non contact Radar Type 3.xls")
EG_PATH       = os.path.join(BASE_DIR, "LT Non contact Radar Type-R0.xls")
EG_SHEET      = "LT-RADAR-EG"

# Color palette indices in the template XLS
IDX_GREEN  = 57   # RGB(51,153,102)   – IODB spec labels
IDX_RED    = 10   # RGB(255,0,0)      – Vendor / job-spec labels
IDX_GREY   = 41   # RGB(239,239,239)  – Project header value cells
IDX_VIOLET = 62   # RGB(51,51,153)    – PURCHASE / NOTES labels
IDX_WHITE  = 9    # RGB(255,255,255)  – value cells (already whitespace)

# Columns (0-based) that carry values in the template
VALUE_COL       = 5    # primary value column (F)
SPLIT_VAL_COLS  = [5, 7, 9]   # for Min / Nor / Max rows

# ── helpers ───────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _banner(text: str) -> None:
    width = max(60, len(text) + 4)
    print("\n" + "═" * width)
    print(f"  {text}")
    print("═" * width)


def _ask(prompt: str, default: str = "", required: bool = False) -> str:
    """Interactive prompt. Empty input uses default. Loops if required and nothing given."""
    hint = f" [{default}]" if default else ""
    req  = " *" if required else ""
    while True:
        raw = input(f"  {prompt}{hint}{req}: ").strip()
        if raw:
            return raw
        if default:
            return default
        if not required:
            return ""
        print("    ⚠  This field is required — please enter a value.")


# ── Template parsing ───────────────────────────────────────────────────────────

def _cell_color_idx(sh, wb, row, col) -> int:
    xf = wb.xf_list[sh.cell_xf_index(row, col)]
    return xf.background.pattern_colour_index


def parse_template(path: str) -> list[dict]:
    """
    Read the XLS template and return a list of spec-row descriptors.
    Each dict has: row, section, label, color, sub_labels, value_cols.
    """
    import xlrd
    wb = xlrd.open_workbook(path, formatting_info=True)
    sh = wb.sheet_by_index(0)

    rows: list[dict] = []
    current_section = ""

    for r in range(sh.nrows):
        col_a = str(sh.cell_value(r, 0)).strip()
        col_b = str(sh.cell_value(r, 1)).strip()

        if col_a and col_a.upper() not in ("NOTES :", "NOTES:"):
            current_section = col_a

        # Determine color of label cell (col B, index 1)
        if col_b:
            color_idx = _cell_color_idx(sh, wb, r, 1)
        else:
            # Header rows where label is in col A (GENERAL, PROCESS CONDITIONS …)
            color_idx = _cell_color_idx(sh, wb, r, 0)

        # --- Project header block (GREY value cells, VIOLET label) ---
        if col_a in ("CLIENT", "CONSULTANT", "PROJECT", "LOCATION") and not col_b:
            rows.append({
                "row": r, "section": col_a, "label": col_a,
                "color": "GREY",
                "sub_labels": [], "value_cols": [1],  # value goes in col B space
            })
            continue

        # Skip rows with no label in col B
        if not col_b:
            continue

        # NOTES section
        if current_section.upper().startswith("NOTES"):
            rows.append({
                "row": r, "section": "NOTES", "label": "Notes",
                "color": "VIOLET",
                "sub_labels": [], "value_cols": [VALUE_COL],
            })
            continue

        # Determine semantic color
        if color_idx == IDX_GREEN:
            color = "GREEN"
        elif color_idx == IDX_RED:
            color = "RED"
        elif color_idx == IDX_VIOLET:
            color = "VIOLET"
        else:
            color = "OTHER"

        # Skip structural/fixed rows (TRANSMITTER Type row 38 has a pre-filled value — handled separately)
        if color == "OTHER":
            continue

        # Check for Min / Nor / Max split row
        c2 = str(sh.cell_value(r, 2)).strip()
        c3 = str(sh.cell_value(r, 3)).strip()
        c4 = str(sh.cell_value(r, 4)).strip()
        is_split = any(v in ("Min", "Nor", "Max") for v in [c2, c3, c4])
        sub_labels = [c2, c3, c4] if is_split else []
        value_cols = SPLIT_VAL_COLS if is_split else [VALUE_COL]

        rows.append({
            "row": r, "section": current_section, "label": col_b,
            "color": color, "sub_labels": sub_labels, "value_cols": value_cols,
        })

    return rows


# ── EG sheet defaults ──────────────────────────────────────────────────────────

def load_eg_defaults(path: str, sheet_name: str) -> dict[str, list[str]]:
    """
    Load predefined values from the EG example sheet.
    Returns {label_lower: [val_for_col5, val_for_col7, val_for_col9]}.
    """
    import xlrd
    wb = xlrd.open_workbook(path, formatting_info=False)
    if sheet_name not in wb.sheet_names():
        return {}
    sh = wb.sheet_by_name(sheet_name)

    defaults: dict[str, list[str]] = {}
    for r in range(sh.nrows):
        label = str(sh.cell_value(r, 1)).strip()
        if not label:
            continue
        v5 = str(sh.cell_value(r, 5)).strip() if sh.ncols > 5 else ""
        v7 = str(sh.cell_value(r, 7)).strip() if sh.ncols > 7 else ""
        v9 = str(sh.cell_value(r, 9)).strip() if sh.ncols > 9 else ""
        # Treat "REFER ANNEXURE*" as no default
        def clean(v):
            return "" if re.match(r"refer\s*annexure", v, re.I) else v
        vals = [clean(v5), clean(v7), clean(v9)]
        if any(vals):
            defaults[_norm(label)] = vals
    return defaults


# ── IODB loading ───────────────────────────────────────────────────────────────

def load_iodb(path: str):
    """
    Returns (df, tag_col, type_col).
    Filters to rows where IO_Type_Name == 'AI' and (Sub_System contains '-' if possible).
    """
    import pandas as pd

    df = pd.read_excel(path, sheet_name=0, header=1)  # row 1 is the real header
    df = df.dropna(how="all").reset_index(drop=True)

    def find_col(priority_keys):
        """
        Two-pass: exact normalised match first (in key priority order),
        then substring match.  This ensures specific keys like
        'tag_number_new' beat generic ones like 'tag'.
        """
        norm = lambda s: re.sub(r"[^a-z0-9]", "", str(s).lower())
        col_norms = {col: norm(col) for col in df.columns}
        # Pass 1 — exact match (check keys in priority order)
        for k in priority_keys:
            kn = norm(k)
            for col, cn in col_norms.items():
                if cn == kn:
                    return col
        # Pass 2 — substring (check keys in priority order, skip "old" columns)
        for k in priority_keys:
            kn = norm(k)
            for col, cn in col_norms.items():
                if kn in cn and "old" not in cn:
                    return col
        # Pass 3 — substring without 'old' restriction
        for k in priority_keys:
            kn = norm(k)
            for col, cn in col_norms.items():
                if kn in cn:
                    return col
        return None

    tag_col    = find_col(["tag_number_new", "tag number new", "tag no", "tag_no", "tag"])
    type_col   = find_col(["instrument_type_desc", "instrument type desc", "instrument type", "instrument_type"])
    signal_col = find_col(["io_type_name", "io type name", "signal i/o type", "io type"])
    subsys_col = find_col(["sub_system", "sub system", "subsystem"])

    if tag_col is None or signal_col is None:
        print(f"  Columns: {list(df.columns[:25])}")
        raise RuntimeError("TAG or SIGNAL column not found in IODB.")

    # Filter: AI signal type
    ai_mask = df[signal_col].fillna("").astype(str).str.strip().str.upper() == "AI"

    # Try to apply subsystem filter; fall back if it would leave nothing
    if subsys_col:
        ss_mask = df[subsys_col].fillna("").astype(str).str.contains("-", na=False)
        combined = df.loc[ai_mask & ss_mask]
        if len(combined) > 0:
            df = combined.reset_index(drop=True)
        else:
            # subsystem filter removed everything — use AI filter only
            df = df.loc[ai_mask].reset_index(drop=True)
            print("  ⚠  No rows with Sub_System containing '-'. Using AI filter only.")
    else:
        df = df.loc[ai_mask].reset_index(drop=True)

    return df, tag_col, type_col


def get_tag_row(df, tag_col, tag: str):
    """Return the first IODB row matching the tag, as a dict."""
    mask = df[tag_col].astype(str).str.strip() == tag.strip()
    if mask.any():
        return df.loc[mask].iloc[0].to_dict()
    return {}


def iodb_value_for_label(label: str, iodb_row: dict, df_cols: list[str]) -> str:
    """Fuzzy-match the label to an IODB column and return the value."""
    try:
        from rapidfuzz import fuzz, process as rfp
        norm_cols = [re.sub(r"[^a-z0-9 ]", "", c.lower()) for c in df_cols]
        norm_q    = re.sub(r"[^a-z0-9 ]", "", label.lower())
        result    = rfp.extractOne(norm_q, norm_cols, scorer=fuzz.WRatio, score_cutoff=65)
        if result:
            _, score, idx = result
            val = iodb_row.get(df_cols[idx])
            if val is not None and str(val).strip() not in ("", "nan", "None"):
                return str(val).strip()
    except ImportError:
        pass
    return ""


# ── Output generation ──────────────────────────────────────────────────────────

def _col_letter(idx: int) -> str:
    """0-based column index to Excel letter (A=0, B=1, ...)."""
    result = ""
    n = idx
    while True:
        result = chr(65 + n % 26) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


def generate_output(template_path: str, filled_rows: list[dict], project_info: dict, output_path: str) -> None:
    """
    Build a clean output .xlsx: same layout as template, NO background colors.
    """
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    FILL_NONE = PatternFill(fill_type=None)  # no fill
    THIN      = Side(style="thin")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    FONT_HDR  = Font(name="Arial", bold=True, size=10)
    FONT_LBL  = Font(name="Arial", bold=True, size=9)
    FONT_VAL  = Font(name="Arial", size=9)
    ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Datasheet"

    row_out = 1

    # ── Company & document header ─────────────────────────────────────────────
    ws.row_dimensions[row_out].height = 28
    ws.merge_cells(f"A{row_out}:E{row_out}")
    c = ws[f"A{row_out}"]
    c.value = "VATECH WABAG LIMITED"
    c.font  = Font(name="Arial", bold=True, size=12)
    c.alignment = ALIGN_CTR
    c.fill = FILL_NONE

    ws.merge_cells(f"F{row_out}:K{row_out}")
    c2 = ws[f"F{row_out}"]
    c2.value = "INSTRUMENT DATASHEET — LEVEL TRANSMITTER (NON-CONTACT RADAR)"
    c2.font  = Font(name="Arial", bold=True, size=10)
    c2.alignment = ALIGN_CTR
    c2.fill = FILL_NONE
    row_out += 1

    ws.append([])  # blank row
    row_out += 1

    # ── Project header block ──────────────────────────────────────────────────
    for key in ("CLIENT", "CONSULTANT", "PROJECT", "LOCATION"):
        ws.row_dimensions[row_out].height = 16
        ws.merge_cells(f"A{row_out}:E{row_out}")
        lc = ws[f"A{row_out}"]
        lc.value    = key
        lc.font     = FONT_LBL
        lc.fill     = FILL_NONE
        lc.alignment = ALIGN_L

        ws.merge_cells(f"F{row_out}:K{row_out}")
        vc = ws[f"F{row_out}"]
        vc.value    = project_info.get(key, "")
        vc.font     = FONT_VAL
        vc.fill     = FILL_NONE
        vc.alignment = ALIGN_L
        row_out += 1

    ws.append([])
    row_out += 1

    # ── Column headers ────────────────────────────────────────────────────────
    ws.row_dimensions[row_out].height = 16
    for col_letter, label in [("A", "SECTION"), ("B", "PARAMETER"), ("F", "VALUE")]:
        c = ws[f"{col_letter}{row_out}"]
        c.value     = label
        c.font      = FONT_HDR
        c.fill      = FILL_NONE
        c.alignment = ALIGN_CTR
        c.border    = BORDER
    row_out += 1

    # ── Spec rows ─────────────────────────────────────────────────────────────
    current_sec = ""
    for spec in filled_rows:
        sec   = spec["section"]
        label = spec["label"]
        vals  = spec.get("filled_values", [""])
        subs  = spec.get("sub_labels", [])

        ws.row_dimensions[row_out].height = 16

        # Section cell (only write when section changes)
        sec_cell = ws[f"A{row_out}"]
        if sec != current_sec:
            sec_cell.value = sec
            sec_cell.font  = Font(name="Arial", bold=True, size=9)
            current_sec = sec
        sec_cell.fill      = FILL_NONE
        sec_cell.alignment = ALIGN_CTR
        sec_cell.border    = BORDER

        # Label cell
        lc = ws[f"B{row_out}"]
        lc.value     = label
        lc.font      = FONT_LBL
        lc.fill      = FILL_NONE
        lc.alignment = ALIGN_L
        lc.border    = BORDER

        if subs and len(vals) == 3:
            # Min / Nor / Max split
            for i, (sub, val) in enumerate(zip(subs, vals)):
                sub_col = get_column_letter(6 + i * 2)  # F, H, J
                sc = ws[f"{sub_col}{row_out}"]
                sc.value     = f"{sub}: {val}" if sub else val
                sc.font      = FONT_VAL
                sc.fill      = FILL_NONE
                sc.alignment = ALIGN_L
                sc.border    = BORDER
        else:
            ws.merge_cells(f"F{row_out}:K{row_out}")
            vc = ws[f"F{row_out}"]
            vc.value     = vals[0] if vals else ""
            vc.font      = FONT_VAL
            vc.fill      = FILL_NONE
            vc.alignment = ALIGN_L
            vc.border    = BORDER

        row_out += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    for col in ["C", "D", "E"]:
        ws.column_dimensions[col].width = 4
    for col in ["F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 18

    wb_out.save(output_path)
    print(f"\n  ✔  Saved: {output_path}")


# ── Main flow ──────────────────────────────────────────────────────────────────

def main():
    _banner("WABAG iEZ — Datasheet Generator (CLI)")

    # Check files exist
    for path, name in [(IODB_PATH, "IODB"), (TEMPLATE_PATH, "Template"), (EG_PATH, "EG workbook")]:
        if not os.path.exists(path):
            print(f"\n  ✗  {name} file not found: {path}")
            print("     Update the path constants at the top of this script.")
            sys.exit(1)

    # ── Load IODB ─────────────────────────────────────────────────────────────
    print("\n  Loading IODB…", end="", flush=True)
    df_iodb, tag_col, type_col = load_iodb(IODB_PATH)
    iodb_cols = list(df_iodb.columns)
    print(f" {len(df_iodb)} AI instruments loaded.")

    # ── Select instrument type ────────────────────────────────────────────────
    types = sorted(df_iodb[type_col].dropna().astype(str).str.strip().unique().tolist()) if type_col else []
    if types:
        print("\n  Available instrument types:")
        for i, t in enumerate(types, 1):
            print(f"    {i:3d}. {t}")
        while True:
            raw = input("\n  Select instrument type number (or type part of the name): ").strip()
            if raw.isdigit() and 1 <= int(raw) <= len(types):
                selected_type = types[int(raw) - 1]
                break
            matches = [t for t in types if _norm(raw) in _norm(t)]
            if len(matches) == 1:
                selected_type = matches[0]
                break
            elif len(matches) > 1:
                print(f"  Matches: {matches[:5]} — please be more specific.")
            else:
                print("  No match found. Try again.")

        type_mask = df_iodb[type_col].astype(str).str.strip() == selected_type
        df_filtered = df_iodb.loc[type_mask].reset_index(drop=True)
    else:
        df_filtered = df_iodb
        selected_type = "N/A"

    # ── Select tags ───────────────────────────────────────────────────────────
    available_tags = sorted(df_filtered[tag_col].dropna().astype(str).str.strip().unique().tolist())
    if not available_tags:
        print(f"\n  ✗  No tags found for instrument type: {selected_type}")
        sys.exit(1)

    print(f"\n  Tags for [{selected_type}]:")
    for i, t in enumerate(available_tags, 1):
        print(f"    {i:3d}. {t}")

    selected_tags = []
    raw = input("\n  Enter tag number(s) [space/comma separated] or 'all': ").strip()
    if raw.lower() == "all":
        selected_tags = available_tags
    else:
        parts = re.split(r"[\s,]+", raw)
        for p in parts:
            if p.isdigit() and 1 <= int(p) <= len(available_tags):
                selected_tags.append(available_tags[int(p) - 1])
            elif p in available_tags:
                selected_tags.append(p)
    if not selected_tags:
        print("  No valid tags selected. Exiting.")
        sys.exit(1)
    print(f"  Processing: {selected_tags}")

    # ── Load template and EG defaults ─────────────────────────────────────────
    print("\n  Parsing template…", end="", flush=True)
    spec_rows = parse_template(TEMPLATE_PATH)
    print(f" {len(spec_rows)} spec rows found.")

    print("  Loading EG defaults…", end="", flush=True)
    eg_defaults = load_eg_defaults(EG_PATH, EG_SHEET)
    print(f" {len(eg_defaults)} predefined values.")

    # ── Project header info (asked once for all tags) ─────────────────────────
    _banner("Project Header Information")
    print("  These values appear in the header of every generated datasheet.\n")
    project_info = {}
    for key in ("CLIENT", "CONSULTANT", "PROJECT", "LOCATION"):
        project_info[key] = _ask(key, required=False)

    # ── Process each tag ──────────────────────────────────────────────────────
    output_files = []

    for tag in selected_tags:
        _banner(f"Tag: {tag}  ({selected_type})")
        iodb_row = get_tag_row(df_filtered, tag_col, tag)

        # Check if we have a row for this tag
        if not iodb_row:
            print(f"  ⚠  Tag '{tag}' not found in IODB — all IODB values will need manual entry.")

        filled_rows = []
        notes_asked = False

        for spec in spec_rows:
            label    = spec["label"]
            color    = spec["color"]
            sub_labs = spec["sub_labels"]
            eg       = eg_defaults.get(_norm(label), [""] * 3)

            # ── NOTES (optional free text) ─────────────────────────────────
            if label == "Notes":
                if notes_asked:
                    continue
                notes_asked = True
                print(f"\n  [NOTES]")
                notes_val = _ask("Enter notes / remarks (Enter to leave blank)", required=False)
                filled_rows.append({**spec, "filled_values": [notes_val]})
                continue

            # ── GREY — header (CLIENT/CONSULTANT etc.) ─────────────────────
            if color == "GREY":
                continue  # already captured in project_info header block

            sec_label = f"{spec['section']} › {label}" if spec['section'] else label
            prefix = {"GREEN": "🟢 IODB", "RED": "🔴 SPEC", "VIOLET": "🟣 PURCHASE"}

            # ── GREEN — auto-fill from IODB; prompt if not found ───────────
            if color == "GREEN":
                if sub_labs:
                    vals = []
                    for i, sub in enumerate(sub_labs):
                        sub_key = f"{label} {sub}"
                        val = (iodb_value_for_label(sub_key, iodb_row, iodb_cols) or
                               iodb_value_for_label(label, iodb_row, iodb_cols)) if iodb_row else ""
                        if val:
                            print(f"\n  🟢  {sec_label} ({sub})  →  {val!r}  ✔ IODB")
                            vals.append(val)
                        else:
                            print(f"\n  🟢  {sec_label} ({sub})  →  NOT in IODB")
                            v = _ask(f"     Enter value for [{label} — {sub}]",
                                     default=eg[i] if i < len(eg) else "",
                                     required=True)
                            vals.append(v)
                else:
                    val = iodb_value_for_label(label, iodb_row, iodb_cols) if iodb_row else ""
                    if val:
                        print(f"\n  🟢  {sec_label}  →  {val!r}  ✔ IODB")
                        vals = [val]
                    else:
                        print(f"\n  🟢  {sec_label}  →  NOT in IODB")
                        v = _ask(f"     Enter value for [{label}]",
                                 default=eg[0] if eg else "",
                                 required=True)
                        vals = [v]

            # ── RED — vendor spec; show EG default, must confirm or override ─
            elif color == "RED":
                if sub_labs:
                    vals = []
                    for i, sub in enumerate(sub_labs):
                        eg_val = eg[i] if i < len(eg) else ""
                        if eg_val:
                            print(f"\n  🔴  {sec_label} ({sub})  →  {eg_val!r}  (EG default — confirm or change)")
                        else:
                            print(f"\n  🔴  {sec_label} ({sub})  →  No default — enter value")
                        v = _ask(f"     [{label} — {sub}]", default=eg_val, required=True)
                        vals.append(v)
                else:
                    eg_val = eg[0] if eg else ""
                    if eg_val:
                        print(f"\n  🔴  {sec_label}  →  {eg_val!r}  (EG default — confirm or change)")
                    else:
                        print(f"\n  🔴  {sec_label}  →  No default — enter value")
                    v = _ask(f"     [{label}]", default=eg_val, required=True)
                    vals = [v]

            # ── VIOLET — PURCHASE Make/Model/Qty (always prompt) ───────────
            elif color == "VIOLET":
                print(f"\n  🟣  {sec_label}")
                v = _ask(f"     Enter [{label}]", required=True)
                vals = [v]

            else:
                continue

            filled_rows.append({**spec, "filled_values": vals})

        # ── Generate output file ──────────────────────────────────────────────
        safe_tag = re.sub(r'[\\/*?:\[\]<>|]', "_", tag)
        output_path = os.path.join(BASE_DIR, f"{safe_tag}_Datasheet.xlsx")
        generate_output(TEMPLATE_PATH, filled_rows, project_info, output_path)
        output_files.append(output_path)

    # ── Done ──────────────────────────────────────────────────────────────────
    _banner("Generation Complete")
    print(f"  Generated {len(output_files)} datasheet(s):\n")
    for f in output_files:
        print(f"    {f}")
    print()


if __name__ == "__main__":
    main()
