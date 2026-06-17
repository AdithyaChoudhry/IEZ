"""
Smart Datasheet Intelligence Engine (SDIE) — Phase 1: OCR extraction.

Pipeline:
    file bytes (PDF/JPG/PNG/TIFF)
        -> extract_pages()     : render each page to a PIL Image
        -> ocr_page()           : pytesseract word-level OCR -> lines with confidence
        -> parse_specifications(): regex "label : value" extraction per line
        -> normalize_specifications(): fuzzy-match labels to canonical WABAG fields
"""
from __future__ import annotations

import io
import re
from typing import Any

import os
import shutil

from PIL import Image
import pytesseract
from pytesseract import Output

# Resolve tesseract binary: prefer explicit env var, then common Linux path (Docker/Render),
# then macOS Homebrew path, then whatever is on PATH.
def _find_tesseract() -> str | None:
    if env := os.environ.get("TESSERACT_CMD"):
        return env
    for candidate in (
        "/usr/bin/tesseract",           # apt (Debian/Ubuntu Docker image)
        "/usr/local/bin/tesseract",     # some Linux builds
        "/opt/homebrew/bin/tesseract",  # macOS Homebrew (Apple Silicon)
        "/usr/local/homebrew/bin/tesseract",
    ):
        if os.path.isfile(candidate):
            return candidate
    return shutil.which("tesseract")

_tess = _find_tesseract()
if _tess:
    pytesseract.pytesseract.tesseract_cmd = _tess

import logging

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .datasheet_generator import (
    normalize,
    fuzzy_match,
    _get_datasheet_ws,
    _is_explicit_placeholder,
    FUZZY_THRESHOLD,
)

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Limits (keep memory/CPU bounded on Render's free tier — 512MB RAM, throttled CPU)
# ─────────────────────────────────────────────────────────────────────────────
MAX_PAGES = 5
PDF_DPI = 120

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}

# ─────────────────────────────────────────────────────────────────────────────
# Canonical WABAG datasheet fields + synonym dictionary
# ─────────────────────────────────────────────────────────────────────────────
CANONICAL_FIELDS = [
    "INSTRUMENT TYPE",
    "MODEL NUMBER",
    "MEASURING RANGE",
    "ACCURACY",
    "OUTPUT SIGNAL",
    "SUPPLY VOLTAGE",
    "PROCESS TEMPERATURE",
    "PROCESS PRESSURE",
    "WETTED MATERIAL",
    "ENCLOSURE PROTECTION",
    "HAZARDOUS AREA CERTIFICATION",
    "COMMUNICATION PROTOCOL",
    "MOUNTING TYPE",
    "PROCESS CONNECTION",
]

SYNONYMS: dict[str, list[str]] = {
    "INSTRUMENT TYPE": ["instrument type", "type", "device type", "transmitter type"],
    "MODEL NUMBER": ["model number", "model no", "model", "part number", "part no", "type number"],
    "MEASURING RANGE": [
        "measuring range", "range", "measurement range", "operating range",
        "span", "calibration range",
    ],
    "ACCURACY": ["accuracy", "accuracy class", "precision", "measurement accuracy"],
    "OUTPUT SIGNAL": [
        "output signal", "output", "signal output", "electrical output",
        "output type",
    ],
    "SUPPLY VOLTAGE": [
        "supply voltage", "power supply", "power", "voltage supply",
        "operating voltage", "input voltage",
    ],
    "PROCESS TEMPERATURE": [
        "process temperature", "operating temperature", "temperature range",
        "ambient temperature", "temperature",
    ],
    "PROCESS PRESSURE": [
        "process pressure", "operating pressure", "pressure rating",
        "pressure range", "pressure",
    ],
    "WETTED MATERIAL": [
        "wetted material", "wetted parts", "material of construction",
        "wetted parts material", "body material",
    ],
    "ENCLOSURE PROTECTION": [
        "enclosure protection", "ingress protection", "protection class",
        "ip rating", "enclosure rating", "housing protection",
    ],
    "HAZARDOUS AREA CERTIFICATION": [
        "hazardous area certification", "hazardous area approval",
        "explosion protection", "ex certification", "area classification",
        "certification",
    ],
    "COMMUNICATION PROTOCOL": [
        "communication protocol", "communication", "protocol",
        "digital communication", "fieldbus",
    ],
    "MOUNTING TYPE": ["mounting type", "mounting", "mounting style", "installation type"],
    "PROCESS CONNECTION": [
        "process connection", "connection type", "process connection size",
        "fitting", "connection",
    ],
}

# Build a flat list of (synonym, canonical_field) for fuzzy matching.
_SYNONYM_CHOICES: list[str] = []
_SYNONYM_TO_CANONICAL: dict[str, str] = {}
for _canonical, _terms in SYNONYMS.items():
    for _term in [_canonical] + _terms:
        norm_term = normalize(_term)
        _SYNONYM_CHOICES.append(norm_term)
        _SYNONYM_TO_CANONICAL[norm_term] = _canonical

MATCH_THRESHOLD = 80


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — page extraction
# ─────────────────────────────────────────────────────────────────────────────
def extract_pages(file_bytes: bytes, filename: str) -> list[Image.Image]:
    """Render an uploaded PDF/image into a list of PIL Images (one per page)."""
    ext = ""
    if "." in filename:
        ext = "." + filename.rsplit(".", 1)[-1].lower()

    if ext == ".pdf":
        from pdf2image import convert_from_bytes
        pages = convert_from_bytes(file_bytes, dpi=PDF_DPI, fmt="png")
        return pages[:MAX_PAGES]

    # Treat everything else as a single-page image (JPG/PNG/TIFF/BMP/...)
    img = Image.open(io.BytesIO(file_bytes))
    # TIFF files may contain multiple frames/pages
    pages: list[Image.Image] = []
    try:
        while True:
            pages.append(img.convert("RGB"))
            img.seek(img.tell() + 1)
    except EOFError:
        pass
    if not pages:
        pages = [img.convert("RGB")]
    return pages[:MAX_PAGES]


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — OCR a single page into lines with per-line confidence
# ─────────────────────────────────────────────────────────────────────────────
def ocr_page(image: Image.Image) -> list[dict[str, Any]]:
    """
    Run OCR on a page image and group words into lines.

    Returns a list of {"text": str, "confidence": float} — confidence is the
    average of the OCR word-level confidences (0-100) for that line.
    """
    if not _tess:
        raise RuntimeError(
            "Tesseract OCR binary not found. "
            "Install it with: apt-get install tesseract-ocr (Linux) or brew install tesseract (macOS)."
        )
    data = pytesseract.image_to_data(image, output_type=Output.DICT)

    lines: dict[tuple[int, int, int], dict[str, Any]] = {}
    n = len(data.get("text", []))
    for i in range(n):
        word = data["text"][i].strip()
        if not word:
            continue
        conf_raw = data["conf"][i]
        try:
            conf = float(conf_raw)
        except (TypeError, ValueError):
            conf = -1.0
        if conf < 0:
            continue

        key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
        entry = lines.setdefault(key, {"words": [], "confs": []})
        entry["words"].append(word)
        entry["confs"].append(conf)

    result = []
    for entry in lines.values():
        text = " ".join(entry["words"]).strip()
        if not text:
            continue
        avg_conf = sum(entry["confs"]) / len(entry["confs"])
        result.append({"text": text, "confidence": round(avg_conf, 1)})

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — parse "label : value" style specifications from OCR'd lines
# ─────────────────────────────────────────────────────────────────────────────
# Matches "Label : Value", "Label - Value", "Label = Value", or a label
# followed by 2+ spaces and a value (common in tabular vendor datasheets).
_SPEC_LINE_RE = re.compile(
    r"^\s*([A-Za-z][A-Za-z0-9/().,\- ]{1,60}?)\s*(?:[:=]|-{1,2}|\s{2,})\s*(\S.*)$"
)


def parse_specifications(pages: list[list[dict[str, Any]]]) -> list[dict[str, Any]]:
    """
    Extract {raw_label, value, confidence, page} dicts from OCR'd page lines.

    `pages` is a list of per-page line lists, as returned by ocr_page().
    """
    specs: list[dict[str, Any]] = []
    for page_idx, lines in enumerate(pages, start=1):
        for line in lines:
            text = line["text"]
            m = _SPEC_LINE_RE.match(text)
            if not m:
                continue
            raw_label = m.group(1).strip(" :-=")
            value = m.group(2).strip()
            if not raw_label or not value:
                continue
            # Skip labels that are themselves just numbers/symbols (false positives)
            if not re.search(r"[A-Za-z]", raw_label):
                continue
            specs.append({
                "raw_label": raw_label,
                "value": value,
                "confidence": line["confidence"],
                "page": page_idx,
            })
    return specs


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 — normalize extracted labels to canonical WABAG fields
# ─────────────────────────────────────────────────────────────────────────────
def normalize_specifications(specs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Attach `canonical_field` (or None) and `match_score` to each extracted spec
    by fuzzy-matching the raw label against the synonym dictionary.
    """
    normalized = []
    for spec in specs:
        norm_label = normalize(spec["raw_label"])
        canonical_field: str | None = None
        match_score = 0.0

        if norm_label in _SYNONYM_TO_CANONICAL:
            canonical_field = _SYNONYM_TO_CANONICAL[norm_label]
            match_score = 100.0
        else:
            best, score = fuzzy_match(norm_label, _SYNONYM_CHOICES, MATCH_THRESHOLD)
            if best is not None:
                canonical_field = _SYNONYM_TO_CANONICAL[best]
                match_score = float(score)

        normalized.append({
            **spec,
            "canonical_field": canonical_field,
            "match_score": round(match_score, 1),
        })
    return normalized


# ─────────────────────────────────────────────────────────────────────────────
# Orchestration
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# Step 5 — map normalized specs onto a WABAG template
# ─────────────────────────────────────────────────────────────────────────────
def _best_specs_by_canonical_field(specs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """For each canonical field, keep the spec with the highest OCR confidence."""
    best: dict[str, dict[str, Any]] = {}
    for spec in specs:
        field = spec.get("canonical_field")
        if not field:
            continue
        current = best.get(field)
        if current is None or spec["confidence"] > current["confidence"]:
            best[field] = spec
    return best


def map_specs_to_template(
    ws,
    specs: list[dict[str, Any]],
    threshold: int = FUZZY_THRESHOLD,
    log: list[dict] | None = None,
) -> None:
    """
    Populate a WABAG "Datasheet" worksheet in-place using extracted vendor specs.

    Mirrors the inverted-scan strategy of
    `datasheet_generator.map_data_to_template`: scans the sheet for explicit
    "Refer Annexure …" placeholder cells, walks leftward along the row to find
    the heading text, fuzzy-matches that heading against the SDIE synonym
    dictionary to resolve a canonical field, and fills the placeholder with the
    matching extracted spec's value (if any).
    """
    specs_by_field = _best_specs_by_canonical_field(specs)

    for row in ws.iter_rows():
        for cell in row:
            if not _is_explicit_placeholder(cell.value):
                continue
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                continue
            if cell.__class__.__name__ == "MergedCell":
                continue

            # Collect non-empty, non-placeholder labels to the left (left-to-right order)
            left_labels: list[str] = []
            for check_col in range(cell.column - 1, 0, -1):
                check_cell = ws.cell(row=cell.row, column=check_col)
                if check_cell.__class__.__name__ == "MergedCell":
                    continue
                cv = check_cell.value
                if cv is None or not str(cv).strip():
                    continue
                cv_str = str(cv).strip()
                if cv_str.startswith("=") or _is_explicit_placeholder(cv_str):
                    continue
                left_labels.insert(0, cv_str)

            if not left_labels:
                logger.debug(
                    "map_specs_to_template: no heading found left of %s%d — skipping",
                    get_column_letter(cell.column), cell.row,
                )
                continue

            heading_candidates = [" ".join(left_labels[start:]) for start in range(len(left_labels))]
            heading_text = heading_candidates[0]

            # ── Resolve canonical field — try each heading candidate ───────────
            canonical_field: str | None = None
            match_score = 0.0
            for candidate in heading_candidates:
                nc = normalize(candidate)
                if nc in _SYNONYM_TO_CANONICAL:
                    canonical_field = _SYNONYM_TO_CANONICAL[nc]
                    match_score = 100.0
                    break
                best, score = fuzzy_match(nc, _SYNONYM_CHOICES, threshold)
                if best is not None and score > match_score:
                    canonical_field = _SYNONYM_TO_CANONICAL[best]
                    match_score = float(score)

            spec = specs_by_field.get(canonical_field) if canonical_field else None

            if spec is None:
                if log is not None:
                    log.append({
                        "heading": heading_text,
                        "canonical_field": canonical_field,
                        "score": match_score,
                        "value": "N/A",
                        "status": "UNMATCHED",
                    })
                continue

            cell.value = spec["value"]
            if log is not None:
                log.append({
                    "heading": heading_text,
                    "canonical_field": canonical_field,
                    "score": match_score,
                    "value": spec["value"],
                    "status": "MATCHED",
                })


def generate_populated_datasheet(
    template_bytes: bytes,
    specs: list[dict[str, Any]],
    threshold: int = FUZZY_THRESHOLD,
) -> tuple[bytes, list[dict], str | None]:
    """
    Populate a WABAG datasheet template with extracted vendor specs.

    Returns (excel_bytes, mapping_log, error). On error, excel_bytes is b"".
    """
    wb = load_workbook(io.BytesIO(template_bytes))

    ws, err = _get_datasheet_ws(wb)
    if err:
        return b"", [], err

    mapping_log: list[dict] = []
    map_specs_to_template(ws, specs, threshold, mapping_log)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), mapping_log, None


# ─────────────────────────────────────────────────────────────────────────────
# Few-shot examples (one per common WABAG instrument type)
# These teach Groq the exact output format and field vocabulary without training.
# ─────────────────────────────────────────────────────────────────────────────
_FEW_SHOT_EXAMPLES = """\
=== EXAMPLE 1: Multi-section tender (Pressure + Level) ===
Input: "1. PRESSURE INSTRUMENTS\nTag PT-201 | Pressure Tx | Range 0-10 bar | 4-20mA HART | 24VDC | Accuracy ±0.075% | IP67 | Safe Area | SS316 | Yokogawa EJA110E\n2. LEVEL INSTRUMENTS\nTag LT-101 | Non Contact Radar | Raw Water Tank | FMCW 80GHz | Beam 3° | Tank Ht 6000mm | 4-20mA+HART | IP67 | ATEX | E+H FMR60B"
Output:
{"sections":[{"heading":"1. PRESSURE INSTRUMENTS","instrument_type":"Pressure Transmitter","specs":{"Tag Number":{"values":["PT-201"],"confidence":100},"Instrument Type":{"values":["Pressure Transmitter"],"confidence":99},"Measuring Range":{"values":["0-10 bar"],"confidence":98},"Output Signal":{"values":["4-20mA HART"],"confidence":99},"Supply Voltage":{"values":["24 VDC"],"confidence":99},"Accuracy":{"values":["±0.075%"],"confidence":99},"Enclosure Protection":{"values":["IP67"],"confidence":99},"Area Classification":{"values":["Safe Area"],"confidence":97},"Wetted Material":{"values":["SS316"],"confidence":98},"Make":{"values":["Yokogawa"],"confidence":92},"Model Number":{"values":["EJA110E"],"confidence":90}}},{"heading":"2. LEVEL INSTRUMENTS","instrument_type":"Non Contact Radar Level Transmitter","specs":{"Tag Number":{"values":["LT-101"],"confidence":100},"Location":{"values":["Raw Water Tank"],"confidence":98},"Sensor Type":{"values":["FMCW Radar"],"confidence":98},"Frequency":{"values":["80 GHz"],"confidence":99},"Beam Angle":{"values":["3°"],"confidence":96},"Tank Height":{"values":["6000 mm"],"confidence":96},"Output Signal":{"values":["4-20mA + HART"],"confidence":99},"Enclosure Protection":{"values":["IP67"],"confidence":99},"Area Certification":{"values":["ATEX"],"confidence":97},"Make":{"values":["Endress+Hauser"],"confidence":92},"Model Number":{"values":["FMR60B"],"confidence":90}}}]}

=== EXAMPLE 2: Three sections (Flow + Pressure + Level) ===
Input: "1. FLOW INSTRUMENTS\nFIT-101 | Mag Flow | Treated Water | DN300 | 250 m³/hr | PTFE lining | SS316L electrodes | ±0.5% | ABB ProcessMaster FEP300\n2. PRESSURE INSTRUMENTS\nPT-201 | 0-10 bar | 4-20mA HART | 24VDC | IP67 | Emerson 3051\n3. LEVEL INSTRUMENTS\nLT-201 | Guided Wave Radar | Sludge | SS316L probe | 0-10m | ±5mm | Emerson Rosemount 3308"
Output:
{"sections":[{"heading":"1. FLOW INSTRUMENTS","instrument_type":"Magnetic Flow Meter","specs":{"Tag Number":{"values":["FIT-101"],"confidence":100},"Fluid":{"values":["Treated Water"],"confidence":98},"Pipe Size":{"values":["DN300"],"confidence":98},"Flow Rate":{"values":["250 m³/hr"],"confidence":97},"Lining Material":{"values":["PTFE"],"confidence":98},"Electrode Material":{"values":["SS316L"],"confidence":98},"Accuracy":{"values":["±0.5%"],"confidence":99},"Make":{"values":["ABB"],"confidence":92},"Model Number":{"values":["ProcessMaster FEP300"],"confidence":90}}},{"heading":"2. PRESSURE INSTRUMENTS","instrument_type":"Pressure Transmitter","specs":{"Tag Number":{"values":["PT-201"],"confidence":100},"Measuring Range":{"values":["0-10 bar"],"confidence":98},"Output Signal":{"values":["4-20mA HART"],"confidence":99},"Supply Voltage":{"values":["24 VDC"],"confidence":99},"Enclosure Protection":{"values":["IP67"],"confidence":99},"Make":{"values":["Emerson"],"confidence":92},"Model Number":{"values":["3051"],"confidence":89}}},{"heading":"3. LEVEL INSTRUMENTS","instrument_type":"Guided Wave Radar Level Transmitter","specs":{"Tag Number":{"values":["LT-201"],"confidence":100},"Fluid":{"values":["Sludge"],"confidence":98},"Probe Material":{"values":["SS316L"],"confidence":98},"Measuring Range":{"values":["0-10 m"],"confidence":97},"Accuracy":{"values":["±5 mm"],"confidence":97},"Make":{"values":["Emerson"],"confidence":92},"Model Number":{"values":["Rosemount 3308"],"confidence":90}}}]}

=== EXAMPLE 3: Single instrument (no section heading found) ===
Input: "Differential Pressure Transmitter, Tag PDT-101, Range 0-250 mbar, Accuracy ±0.05%, Rangeability 150:1, 4-20mA+HART, 24VDC, Wetted SS316, Emerson Rosemount 3051CD"
Output:
{"sections":[{"heading":"Unidentified Instrument","instrument_type":"Differential Pressure Transmitter","specs":{"Tag Number":{"values":["PDT-101"],"confidence":100},"Instrument Type":{"values":["Differential Pressure Transmitter"],"confidence":99},"Pressure Range":{"values":["0-250 mbar"],"confidence":98},"Accuracy":{"values":["±0.05%"],"confidence":99},"Rangeability":{"values":["150:1"],"confidence":97},"Output Signal":{"values":["4-20mA + HART"],"confidence":99},"Supply Voltage":{"values":["24 VDC"],"confidence":99},"Wetted Material":{"values":["SS316"],"confidence":98},"Make":{"values":["Emerson"],"confidence":92},"Model Number":{"values":["Rosemount 3051CD"],"confidence":91}}}]}
"""

_GROQ_SYSTEM = f"""You are an Instrumentation Engineering Expert for EPC Water/Wastewater/Desalination projects (WABAG standard).

A tender document may describe MULTIPLE instrument types under numbered section headings (e.g. "1. PRESSURE INSTRUMENTS", "2. LEVEL INSTRUMENTS", "3. FLOW INSTRUMENTS").

TASK:
1. Identify all instrument sections separated by numbered/lettered headings.
2. For each section, extract ALL technical specifications from the text UNDER that heading until the next heading.
3. If no section headings found, create ONE section with heading "Unidentified Instrument".

Return ONLY valid JSON — no explanation, no markdown, no prefix:
{{
  "sections": [
    {{
      "heading": "exact heading from document or 'Unidentified Instrument'",
      "instrument_type": "specific type e.g. 'Pressure Transmitter'",
      "specs": {{
        "Field Name": {{"values": ["primary value", "alternate value if found elsewhere in same section"], "confidence": 95}},
        ...
      }}
    }}
  ]
}}

RULES:
- Specs within a section come ONLY from text under that heading, not from other sections.
- If the same field appears multiple times WITHIN a section with DIFFERENT values, list ALL values.
- Extract EVERY field per section: tag number, instrument type, fluid, measuring range, output signal, supply voltage, accuracy, repeatability, IP rating, area classification, certifications, make, model, materials (body/wetted/electrode/lining/probe), process connection, pipe size, flow rate, frequency, beam angle, sensor type, protocol, rangeability, tank height, specific gravity, and everything else.
- Confidence: 95-100 = explicitly stated, 80-94 = clearly implied, 60-79 = inferred.
- Omit fields where value is empty, N/A, or unknown.
- Keep heading text exactly as it appears in the document.

EXAMPLES:
{_FEW_SHOT_EXAMPLES}"""


def extract_template_fields(template_bytes: bytes) -> list[str]:
    """
    Read the target field names from a WABAG datasheet template (sheet 2 or 3).
    Scans for 'Refer Annexure' placeholder cells and collects the heading text
    to their left — these are the fields the template expects to be filled.
    Returns an empty list if template cannot be parsed.
    """
    try:
        wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
        # Try the standard datasheet worksheet first, then fall back to sheets 2/3
        ws, err = _get_datasheet_ws(wb)
        if err:
            sheets = wb.sheetnames
            candidates = [wb[s] for s in sheets[1:3]]  # sheets 2 and 3
            ws = candidates[0] if candidates else None
        if ws is None:
            return []

        fields: list[str] = []
        seen: set[str] = set()
        for row in ws.iter_rows():
            for cell in row:
                if not _is_explicit_placeholder(cell.value):
                    continue
                # Walk left to find the heading
                for col in range(cell.column - 1, 0, -1):
                    c = ws.cell(row=cell.row, column=col)
                    cv = c.value
                    if cv and str(cv).strip() and not _is_explicit_placeholder(str(cv)):
                        label = str(cv).strip()
                        if label not in seen:
                            fields.append(label)
                            seen.add(label)
                        break
        logger.info("SDIE: extracted %d template fields", len(fields))
        return fields
    except Exception as exc:
        logger.warning("SDIE: template field extraction failed: %s", exc)
        return []


_SKIP_VALUES = {"", "n/a", "na", "none", "null", "not specified", "unknown", "-", "—"}


def _parse_specs_dict(specs_dict: dict) -> list[dict[str, Any]]:
    """Convert a raw AI specs dict into a list of ExtractedSpec-compatible dicts."""
    result = []
    for field, data in specs_dict.items():
        if isinstance(data, dict):
            raw_vals = data.get("values") or [data.get("value", "")]
            conf = int(data.get("confidence", 90))
        elif isinstance(data, list):
            raw_vals = data
            conf = 85
        else:
            raw_vals = [str(data)]
            conf = 85

        values: list[str] = []
        seen: set[str] = set()
        for v in raw_vals:
            s = str(v).strip()
            if s.lower() not in _SKIP_VALUES and s not in seen:
                values.append(s)
                seen.add(s)

        if not values:
            continue

        result.append({
            "raw_label": field,
            "values": values,
            "value": values[0],
            "confidence": min(100, max(0, conf)),
            "page": 1,
            "canonical_field": field,
            "match_score": float(conf),
            "source": "ai",
        })
    return result


def _run_ai_extraction(
    ocr_text: str,
    template_fields: list[str] | None = None,
) -> tuple[list[dict[str, Any]], str]:
    """
    Call Groq API — returns (sections, summary_instrument_type).
    Each section: {heading, instrument_type, specs: [ExtractedSpec-compatible dicts]}
    """
    groq_key = os.environ.get("GROQ_API_KEY", "")
    if not groq_key:
        logger.warning("SDIE: GROQ_API_KEY not set — AI extraction skipped")
        return [], ""

    import json as _json
    import requests as _requests

    user_content = ocr_text[:8000]
    if template_fields:
        field_list = ", ".join(template_fields[:40])
        user_content = (
            f"TARGET FIELDS TO EXTRACT (from the datasheet template):\n{field_list}\n\n"
            f"DOCUMENT TEXT:\n{ocr_text[:7500]}"
        )

    try:
        resp = _requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {groq_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "messages": [
                    {"role": "system", "content": _GROQ_SYSTEM},
                    {"role": "user", "content": user_content},
                ],
                "temperature": 0.1,
                "max_tokens": 3000,
            },
            timeout=60,
        )
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"].strip()
    except Exception as exc:
        logger.error("SDIE: Groq API call failed: %s", exc)
        return [], ""

    import json as _json
    try:
        ai_result = _json.loads(raw)
    except _json.JSONDecodeError:
        import re as _re
        m = _re.search(r"\{.*\}", raw, _re.DOTALL)
        if not m:
            logger.error("SDIE: Could not parse JSON from Groq response")
            return [], ""
        try:
            ai_result = _json.loads(m.group())
        except _json.JSONDecodeError as e:
            logger.error("SDIE: JSON decode failed: %s", e)
            return [], ""

    # ── Parse multi-section response ──────────────────────────────────────────
    raw_sections = ai_result.get("sections", [])
    if raw_sections and isinstance(raw_sections, list):
        sections = []
        for sec in raw_sections:
            heading = str(sec.get("heading", "Unidentified Instrument")).strip()
            instr_type = str(sec.get("instrument_type", "")).strip()
            specs = _parse_specs_dict(sec.get("specs", {}))
            sections.append({"heading": heading, "instrument_type": instr_type, "specs": specs})
    else:
        # Fallback: AI returned old flat format (no "sections" key)
        instr_type = str(ai_result.pop("_instrument_type", "")).strip()
        specs = _parse_specs_dict(ai_result)
        sections = [{"heading": instr_type or "Unidentified Instrument", "instrument_type": instr_type, "specs": specs}]

    summary_type = ", ".join(s["instrument_type"] for s in sections if s["instrument_type"])
    total_specs = sum(len(s["specs"]) for s in sections)
    logger.info("SDIE: AI extracted %d section(s), %d spec(s) total, types=%r", len(sections), total_specs, summary_type)
    return sections, summary_type


def extract_specifications(
    file_bytes: bytes,
    filename: str,
    template_bytes: bytes | None = None,
) -> tuple[list[dict[str, Any]], int, str]:
    """
    Run the full extraction pipeline on an uploaded vendor datasheet.

    Renders pages → OCR text → Groq AI (section-aware extraction).
    Returns (sections, page_count, summary_instrument_type).
    sections = [{heading, instrument_type, specs: [ExtractedSpec-compatible dicts]}]
    """
    logger.info("SDIE: extracting from '%s' (%d bytes)", filename, len(file_bytes))
    pages = extract_pages(file_bytes, filename)
    logger.info("SDIE: rendered %d page(s)", len(pages))

    ocr_text_parts = []
    for idx, page in enumerate(pages, start=1):
        lines = ocr_page(page)
        logger.info("SDIE: OCR page %d/%d -> %d line(s)", idx, len(pages), len(lines))
        for line in lines:
            ocr_text_parts.append(line.get("text", ""))

    ocr_text = "\n".join(line for line in ocr_text_parts if len(line.strip()) > 3)
    logger.info("SDIE: sending %d chars to Groq AI", len(ocr_text))

    template_fields: list[str] = []
    if template_bytes:
        template_fields = extract_template_fields(template_bytes)

    sections, summary_type = _run_ai_extraction(ocr_text, template_fields or None)
    total = sum(len(s["specs"]) for s in sections)
    logger.info("SDIE: final %d section(s), %d spec(s), type=%r", len(sections), total, summary_type)
    return sections, len(pages), summary_type


def specs_to_excel_bytes(specs_or_sections: list[dict[str, Any]]) -> bytes:
    """Export extracted specifications to a single-sheet .xlsx workbook.

    Accepts either a flat list of spec dicts or a list of section dicts
    ({heading, instrument_type, specs}).
    """
    from openpyxl import Workbook

    # Flatten sections if needed
    if specs_or_sections and "specs" in specs_or_sections[0]:
        specs: list[dict] = []
        for sec in specs_or_sections:
            for sp in sec.get("specs", []):
                specs.append({**sp, "_section": sec.get("heading", "")})
    else:
        specs = specs_or_sections

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Specs"

    headers = ["Section", "Raw Label", "Value", "Mapped Field", "Confidence (%)", "Page"]
    ws.append(headers)
    for spec in specs:
        ws.append([
            spec.get("_section", ""),
            spec.get("raw_label", ""),
            spec.get("value", ""),
            spec.get("canonical_field") or "Other technical specification",
            round(float(spec.get("confidence", 0)), 1),
            spec.get("page", ""),
        ])

    widths = [28, 32, 40, 30, 16, 6]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
