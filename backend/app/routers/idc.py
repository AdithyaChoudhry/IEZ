"""
IDC (Inter Discipline Check) REST API router.
Handles sessions, documents, comments, annotations, approvals, and report generation.
"""
import json
import logging
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session as DBSession

from ..auth.models import AdminUser
from ..auth.utils import verify_password
from ..database import SessionLocal
from ..deps import get_current_user, get_db
from ..auth.models import User
from ..models.idc import (
    IDCAnnotation, IDCApproval, IDCComment, IDCCommentReply,
    IDCDiscipline, IDCDocument, IDCSession,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/idc", tags=["IDC"])

UPLOAD_DIR = Path("/var/www/iez/idc_uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

DISCIPLINE_COLORS = {
    "Civil Engineering":            "#8B4513",
    "Mechanical Engineering":       "#1E90FF",
    "Electrical Engineering":       "#FF8C00",
    "Piping Engineering":           "#9932CC",
    "Instrumentation Engineering":  "#228B22",
    "Technical Coordinator":        "#000000",
}

ALLOWED_EXT = {".pdf", ".dwg", ".dxf", ".dgn", ".zip"}


# ── helpers ────────────────────────────────────────────────────────────────────
def _session_or_404(session_id: int, db: DBSession) -> IDCSession:
    s = db.query(IDCSession).filter(IDCSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "IDC session not found")
    return s


def _verify_emp(employee_id: str, password: str, db: DBSession) -> AdminUser:
    emp = db.query(AdminUser).filter(
        AdminUser.employee_id == employee_id, AdminUser.status == "active"
    ).first()
    if not emp or not verify_password(password, emp.password_hash):
        raise HTTPException(401, "Invalid Employee ID or Password")
    return emp


def _next_comment_number(session_id: int, db: DBSession) -> str:
    count = db.query(IDCComment).filter(IDCComment.session_id == session_id).count()
    return f"IDC-{count + 1:03d}"


def _session_dict(s: IDCSession) -> dict:
    approvals = {a.discipline: {"employee_name": a.employee_name, "approved_at": a.approved_at.isoformat()} for a in s.approvals}
    disciplines = [d.discipline for d in s.disciplines]
    all_approved = all(d in approvals for d in disciplines)
    return {
        "id": s.id, "idc_number": s.idc_number, "project_name": s.project_name,
        "document_number": s.document_number, "document_title": s.document_title,
        "revision_number": s.revision_number, "document_category": s.document_category,
        "due_date": s.due_date, "remarks": s.remarks, "status": s.status,
        "created_by_emp": s.created_by_emp, "created_by_name": s.created_by_name,
        "created_at": s.created_at.isoformat(),
        "frozen_at": s.frozen_at.isoformat() if s.frozen_at else None,
        "frozen_by_name": s.frozen_by_name,
        "disciplines": disciplines,
        "approvals": approvals,
        "all_approved": all_approved,
        "documents": [{"id": d.id, "original_filename": d.original_filename, "file_type": d.file_type, "file_size": d.file_size} for d in s.documents],
        "comment_count": len(s.comments),
    }


# ── Session CRUD ───────────────────────────────────────────────────────────────
@router.get("/sessions")
def list_sessions(db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    sessions = db.query(IDCSession).order_by(IDCSession.id.desc()).all()
    return [_session_dict(s) for s in sessions]


@router.post("/sessions")
async def create_session(
    idc_number: str = Form(...),
    project_name: str = Form(...),
    document_number: str = Form(...),
    document_title: str = Form(...),
    revision_number: str = Form(...),
    document_category: str = Form(...),
    due_date: str = Form(...),
    remarks: str = Form(""),
    disciplines: str = Form(...),           # JSON array
    employee_id: str = Form(...),
    password: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    emp = _verify_emp(employee_id, password, db)
    if db.query(IDCSession).filter(IDCSession.idc_number == idc_number).first():
        raise HTTPException(400, "IDC number already exists")

    session = IDCSession(
        idc_number=idc_number, project_name=project_name,
        document_number=document_number, document_title=document_title,
        revision_number=revision_number, document_category=document_category,
        due_date=due_date, remarks=remarks,
        created_by_emp=emp.employee_id, created_by_name=emp.employee_name,
    )
    db.add(session); db.flush()

    disc_list = json.loads(disciplines)
    for d in disc_list:
        db.add(IDCDiscipline(session_id=session.id, discipline=d))

    for f in files:
        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_EXT:
            continue
        stored = f"{uuid.uuid4().hex}{ext}"
        dest = UPLOAD_DIR / stored
        content = await f.read()
        dest.write_bytes(content)
        db.add(IDCDocument(
            session_id=session.id, filename=stored,
            original_filename=f.filename, file_type=ext.lstrip("."),
            file_size=len(content),
            uploaded_by_emp=emp.employee_id, uploaded_by_name=emp.employee_name,
        ))

    db.commit(); db.refresh(session)
    return _session_dict(session)


@router.get("/sessions/{session_id}")
def get_session(session_id: int, db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    return _session_dict(_session_or_404(session_id, db))


@router.delete("/sessions/{session_id}", status_code=204)
def delete_session(
    session_id: int,
    employee_id: str,
    password: str,
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = _session_or_404(session_id, db)
    emp = _verify_emp(employee_id, password, db)
    if emp.role not in ("Admin",) and s.created_by_emp != emp.employee_id:
        raise HTTPException(403, "Only the session creator or Admin can delete")
    # delete uploaded files from disk
    for doc in s.documents:
        path = UPLOAD_DIR / doc.filename
        if path.exists():
            path.unlink()
    db.delete(s)
    db.commit()


# ── Document serve ─────────────────────────────────────────────────────────────
@router.get("/documents/{doc_id}/file")
def serve_document(doc_id: int, db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    doc = db.query(IDCDocument).filter(IDCDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(404, "Document not found")
    path = UPLOAD_DIR / doc.filename
    if not path.exists():
        raise HTTPException(404, "File not found on disk")
    media = "application/pdf" if doc.file_type == "pdf" else "application/octet-stream"
    return FileResponse(path, media_type=media, filename=doc.original_filename)


# ── Annotations REST ───────────────────────────────────────────────────────────
@router.get("/sessions/{session_id}/annotations")
def list_annotations(session_id: int, db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    anns = db.query(IDCAnnotation).filter(
        IDCAnnotation.session_id == session_id, IDCAnnotation.is_deleted == False
    ).all()
    return [{
        "id": a.id, "ann_uuid": a.ann_uuid, "document_id": a.document_id,
        "tool_type": a.tool_type, "page_number": a.page_number,
        "x": a.x, "y": a.y, "width": a.width, "height": a.height,
        "data_json": a.data_json, "color": a.color,
        "author_emp": a.author_emp, "author_name": a.author_name, "discipline": a.discipline,
        "created_at": a.created_at.isoformat(),
    } for a in anns]


# ── Comments REST ──────────────────────────────────────────────────────────────
@router.get("/sessions/{session_id}/comments")
def list_comments(session_id: int, db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    comments = db.query(IDCComment).filter(IDCComment.session_id == session_id).order_by(IDCComment.id).all()
    return [{
        "id": c.id, "comment_number": c.comment_number, "ann_uuid": c.ann_uuid,
        "page_number": c.page_number, "author_emp": c.author_emp, "author_name": c.author_name,
        "discipline": c.discipline, "comment_text": c.comment_text,
        "priority": c.priority, "status": c.status, "category": c.category,
        "created_at": c.created_at.isoformat(),
        "resolved_by": c.resolved_by,
        "resolved_at": c.resolved_at.isoformat() if c.resolved_at else None,
        "replies": [{
            "id": r.id, "author_name": r.author_name, "discipline": r.discipline,
            "reply_text": r.reply_text, "created_at": r.created_at.isoformat(),
        } for r in c.replies],
    } for c in comments]


@router.post("/sessions/{session_id}/comments")
def add_comment(
    session_id: int,
    ann_uuid: Optional[str] = Form(None),
    page_number: int = Form(1),
    employee_id: str = Form(...),
    password: str = Form(...),
    comment_text: str = Form(...),
    priority: str = Form("Normal"),
    category: str = Form("General"),
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = _session_or_404(session_id, db)
    if s.status == "frozen":
        raise HTTPException(403, "Session is frozen")
    emp = _verify_emp(employee_id, password, db)
    disc = DISCIPLINE_COLORS  # just to validate discipline
    # find discipline for this emp in session
    discipline = emp.department or "Instrumentation Engineering"
    comment = IDCComment(
        session_id=session_id, ann_uuid=ann_uuid,
        comment_number=_next_comment_number(session_id, db),
        page_number=page_number, author_emp=emp.employee_id, author_name=emp.employee_name,
        discipline=discipline, comment_text=comment_text, priority=priority, category=category,
    )
    db.add(comment); db.commit(); db.refresh(comment)
    return {"id": comment.id, "comment_number": comment.comment_number}


@router.patch("/sessions/{session_id}/comments/{comment_id}")
def update_comment_status(
    session_id: int, comment_id: int,
    status: str = Form(...),
    employee_id: str = Form(...),
    password: str = Form(...),
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    emp = _verify_emp(employee_id, password, db)
    c = db.query(IDCComment).filter(IDCComment.id == comment_id, IDCComment.session_id == session_id).first()
    if not c:
        raise HTTPException(404, "Comment not found")
    c.status = status; c.updated_at = datetime.utcnow()
    if status in ("Resolved", "Closed"):
        c.resolved_by = emp.employee_name; c.resolved_at = datetime.utcnow()
    db.commit()
    return {"status": status}


@router.post("/sessions/{session_id}/comments/{comment_id}/reply")
def add_reply(
    session_id: int, comment_id: int,
    employee_id: str = Form(...),
    password: str = Form(...),
    reply_text: str = Form(...),
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = _session_or_404(session_id, db)
    if s.status == "frozen":
        raise HTTPException(403, "Session is frozen")
    emp = _verify_emp(employee_id, password, db)
    discipline = emp.department or "Instrumentation Engineering"
    reply = IDCCommentReply(
        comment_id=comment_id, author_emp=emp.employee_id,
        author_name=emp.employee_name, discipline=discipline, reply_text=reply_text,
    )
    db.add(reply); db.commit()
    return {"ok": True}


# ── Approval ───────────────────────────────────────────────────────────────────
@router.post("/sessions/{session_id}/approve")
def approve_session(
    session_id: int,
    discipline: str = Form(...),
    employee_id: str = Form(...),
    password: str = Form(...),
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = _session_or_404(session_id, db)
    if s.status == "frozen":
        raise HTTPException(403, "Session already frozen")
    emp = _verify_emp(employee_id, password, db)
    existing = db.query(IDCApproval).filter(
        IDCApproval.session_id == session_id, IDCApproval.discipline == discipline
    ).first()
    if existing:
        raise HTTPException(400, f"{discipline} has already approved")
    db.add(IDCApproval(
        session_id=session_id, discipline=discipline,
        employee_id=emp.employee_id, employee_name=emp.employee_name,
    ))
    db.commit()
    return {"approved": True, "employee_name": emp.employee_name}


# ── Freeze ─────────────────────────────────────────────────────────────────────
@router.post("/sessions/{session_id}/freeze")
def freeze_session(
    session_id: int,
    employee_id: str = Form(...),
    password: str = Form(...),
    db: DBSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = _session_or_404(session_id, db)
    if s.status == "frozen":
        raise HTTPException(400, "Already frozen")
    emp = _verify_emp(employee_id, password, db)
    disciplines = [d.discipline for d in s.disciplines]
    approvals = {a.discipline for a in s.approvals}
    missing = [d for d in disciplines if d not in approvals]
    if missing:
        raise HTTPException(400, f"Pending approvals: {', '.join(missing)}")
    s.status = "frozen"
    s.frozen_at = datetime.utcnow()
    s.frozen_by_emp = emp.employee_id
    s.frozen_by_name = emp.employee_name
    db.commit()
    return {"frozen": True}


# ── Comment Register Export ────────────────────────────────────────────────────
@router.get("/sessions/{session_id}/export/comments")
def export_comment_register(session_id: int, db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    s = _session_or_404(session_id, db)
    comments = db.query(IDCComment).filter(IDCComment.session_id == session_id).order_by(IDCComment.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IDC Comment Register"

    # Header info
    ws.merge_cells("A1:K1")
    ws["A1"] = f"IDC COMMENT REGISTER — {s.idc_number} | {s.document_title} | Rev {s.revision_number}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0D1B2A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Comment No.", "Page", "Discipline", "Category", "Priority", "Comment", "Reply", "Status", "Reviewer", "Date"]
    col_widths = [4, 12, 6, 20, 18, 10, 50, 50, 15, 20, 12]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = w

    disc_fill = {
        "Civil Engineering":            "8B4513",
        "Mechanical Engineering":       "1E90FF",
        "Electrical Engineering":       "FF8C00",
        "Piping Engineering":           "9932CC",
        "Instrumentation Engineering":  "228B22",
        "Technical Coordinator":        "555555",
    }
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, c in enumerate(comments, 3):
        replies = " | ".join(r.reply_text for r in c.replies) if c.replies else ""
        values = [
            row_i - 2, c.comment_number, c.page_number, c.discipline, c.category,
            c.priority, c.comment_text, replies, c.status, c.author_name,
            c.created_at.strftime("%Y-%m-%d"),
        ]
        fill_color = disc_fill.get(c.discipline, "EEEEEE")
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_i == 4:  # discipline column
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(color="FFFFFF", bold=True)
        ws.row_dimensions[row_i].height = 40

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"IDC_{s.idc_number}_CommentRegister.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Dashboard stats ────────────────────────────────────────────────────────────
@router.get("/sessions/{session_id}/stats")
def session_stats(session_id: int, db: DBSession = Depends(get_db), _: User = Depends(get_current_user)):
    s = _session_or_404(session_id, db)
    comments = s.comments
    by_disc = {}
    for c in comments:
        by_disc.setdefault(c.discipline, {"total": 0, "open": 0, "resolved": 0})
        by_disc[c.discipline]["total"] += 1
        if c.status in ("Open", "Under Review", "Need Clarification"):
            by_disc[c.discipline]["open"] += 1
        elif c.status in ("Resolved", "Closed"):
            by_disc[c.discipline]["resolved"] += 1
    by_status = {}
    for c in comments:
        by_status[c.status] = by_status.get(c.status, 0) + 1
    approvals = {a.discipline: a.employee_name for a in s.approvals}
    disciplines = [d.discipline for d in s.disciplines]
    return {
        "total_comments": len(comments),
        "by_discipline": by_disc,
        "by_status": by_status,
        "disciplines": disciplines,
        "approvals": approvals,
        "pending_approvals": [d for d in disciplines if d not in approvals],
    }
