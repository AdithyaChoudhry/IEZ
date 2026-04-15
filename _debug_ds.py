import sys, io, traceback
sys.path.insert(0, '/Users/adithyachoudhrym/ProjectIEZ')

import openpyxl, pandas as pd
from utils.file_handler import load_workbook_from_upload
from utils.generators import generate_datasheets, _copy_workbook

TMPL = 'Pressure Transmitter.xlsx'
IODB = 'IODB Source File.xlsx'

with open(IODB, 'rb') as f:
    iodb_buf = io.BytesIO(f.read())
with open(TMPL, 'rb') as f:
    tmpl_data = f.read()
    tmpl_buf = io.BytesIO(tmpl_data)

# Load IODB
iodb_buf.seek(0)
df = pd.read_excel(iodb_buf, sheet_name="IODB", header=0).dropna(how='all')
tags = df['TAG NO'].dropna().astype(str).str.strip().unique()[:1].tolist()
print(f"Tags to test: {tags}")

# Load template
wb, buf_ref, err = load_workbook_from_upload(tmpl_buf)
print(f"Template loaded: {wb is not None}, err={err}, sheets={wb.sheetnames if wb else None}")
print(f"buf_ref is BytesIO? {isinstance(buf_ref, io.BytesIO)}")
print(f"buf_ref closed? {buf_ref.closed if buf_ref else 'N/A'}")

# Try _copy_workbook on its own
print("\nTesting _copy_workbook...")
try:
    new_wb = _copy_workbook(wb)
    print(f"  _copy_workbook OK: sheets={new_wb.sheetnames}")
except Exception as e:
    print(f"  _copy_workbook FAILED: {e}")
    traceback.print_exc()

# Try generate_datasheets with a single tag
print("\nTesting generate_datasheets...")
try:
    tmpl_buf2 = io.BytesIO(tmpl_data)
    wb2, buf2, err2 = load_workbook_from_upload(tmpl_buf2)
    results, err3 = generate_datasheets(df, wb2, 'TAG NO', tags)
    if err3:
        print(f"  FAILED: {err3}")
    else:
        print(f"  OK: {len(results)} results, first={results[0][1]}")
except Exception as e:
    print(f"  EXCEPTION: {e}")
    traceback.print_exc()
