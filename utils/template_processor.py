"""
template_processor.py
Higher-level helpers that combine file_handler + generators,
providing simple single-call interfaces for each module
that Streamlit main.py can use directly.
"""

import io
import zipfile
import openpyxl
import pandas as pd

from utils.file_handler import (
    read_iodb,
    read_loop_wiring_input,
    load_workbook_from_upload,
    dataframe_to_bytes,
)
from utils.generators import (
    generate_instrument_list,
    generate_io_list,
    generate_datasheets,
    generate_cable_schedule,
    generate_loop_wiring,
    generate_iodb_validation,
)


# ─────────────────────────────────────────────────────────────────────────────
# Instrument List
# ─────────────────────────────────────────────────────────────────────────────

def process_instrument_list(
    iodb_file,
    selected_columns: list[str],
    filters: dict | None = None,
    template_file=None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Full pipeline: read IODB → (optionally filter) → generate instrument list.
    If `template_file` is provided, output is appended as a new sheet to that workbook.

    Returns (bytes, filename, error)
    """
    df, err = read_iodb(iodb_file)
    if err:
        return None, None, f"Failed to read IODB: {err}"

    template_wb = None
    if template_file is not None:
        template_wb, _, err = load_workbook_from_upload(template_file)
        if err:
            return None, None, f"Failed to load template: {err}"

    return generate_instrument_list(df, selected_columns, filters=filters, template_wb=template_wb)


# ─────────────────────────────────────────────────────────────────────────────
# I/O List
# ─────────────────────────────────────────────────────────────────────────────

def process_io_list(
    iodb_file,
    selected_columns: list[str],
    filters: dict | None = None,
    template_file=None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Full pipeline: read IODB → (optionally filter) → generate I/O list.
    If `template_file` is provided, output is appended as a new sheet to that workbook.

    Returns (bytes, filename, error)
    """
    df, err = read_iodb(iodb_file)
    if err:
        return None, None, f"Failed to read IODB: {err}"

    template_wb = None
    if template_file is not None:
        template_wb, _, err = load_workbook_from_upload(template_file)
        if err:
            return None, None, f"Failed to load template: {err}"

    return generate_io_list(df, selected_columns, filters=filters, template_wb=template_wb)


# ─────────────────────────────────────────────────────────────────────────────
# Data Sheet
# ─────────────────────────────────────────────────────────────────────────────

def process_datasheets(
    iodb_file,
    template_file,
    tag_column: str,
    selected_tags: list[str],
    progress_callback=None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Full pipeline: read IODB + template → generate datasheets → zip all files.

    Returns (zip_bytes, "Datasheets.zip", error)
    """
    df, err = read_iodb(iodb_file)
    if err:
        return None, None, f"Failed to read IODB: {err}"

    wb, _tmpl_buf, err = load_workbook_from_upload(template_file)
    if err:
        return None, None, f"Failed to load template: {err}"
    # Keep _tmpl_buf alive so openpyxl can copy embedded images/drawings
    # from the source ZIP when wb.save() is called inside generate_datasheets.

    results, err = generate_datasheets(df, wb, tag_column, selected_tags, progress_callback)
    if err:
        return None, None, err
    if not results:
        return None, None, "No datasheets were generated. Check that the selected tags exist in the IODB."

    # Package all individual Excel files into a ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_bytes, filename in results:
            zf.writestr(filename, file_bytes)
    zip_buf.seek(0)
    return zip_buf.read(), "Datasheets.zip", None


# ─────────────────────────────────────────────────────────────────────────────
# Data Sheet v2  (two-row header · AI filter · Column-D heading mapping · rapidfuzz)
# ─────────────────────────────────────────────────────────────────────────────

def process_datasheets_v2(
    iodb_file,
    template_file,
    selected_tags: list[str],
    two_row_header: bool = True,
    fuzzy_threshold: int = 70,
    progress_callback=None,
) -> tuple[bytes | None, str | None, str | None, list[dict]]:
    """
    New datasheet pipeline using the datasheet_generator module.

    Reads IODB with optional two-row combined header, populates the template's
    "Datasheet" (Sheet 2) by fuzzy-matching Column D headings to IODB columns,
    and returns all files as a ZIP.

    Returns (zip_bytes, "Datasheets.zip", error_or_None, mapping_logs)
    """
    from utils.datasheet_generator import batch_generate_zip  # local import to avoid circular deps

    zip_bytes, filename, err, logs = batch_generate_zip(
        iodb_file=iodb_file,
        template_file=template_file,
        selected_tags=selected_tags,
        two_row_header=two_row_header,
        threshold=fuzzy_threshold,
        progress_callback=progress_callback,
    )
    return zip_bytes, filename, err, logs


def get_ai_tags_from_iodb(
    iodb_file,
    two_row_header: bool = True,
) -> tuple[list[str], str | None]:
    """
    Return (ai_tag_list, error) using the new datasheet_generator loader.
    Used by the Streamlit UI to populate the tag selector with AI-only tags.
    """
    from utils.datasheet_generator import load_iodb, get_ai_tags  # local import
    df, col_names, err = load_iodb(iodb_file, two_row_header=two_row_header)
    if err:
        return [], err
    return get_ai_tags(df, col_names)


# ─────────────────────────────────────────────────────────────────────────────
# Cable Schedule
# ─────────────────────────────────────────────────────────────────────────────

def process_cable_schedule(
    iodb_file,
    template_file,
    jb_column: str = "JUNCTION BOX",
    tag_column: str = "TAG NO",
    progress_callback=None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Full pipeline: read IODB + template → generate cable schedule.

    The template is loaded twice:
      - In read-only mode for safe structure/prototype reading.
      - The raw bytes are passed to the generator so it can reload the template
        in writable mode (preserving header formatting, merged cells, etc.).
    Falls back to a fresh workbook if the writable load fails.

    Returns (bytes, "Cable_Schedule.xlsx", error)
    """
    df, err = read_iodb(iodb_file)
    if err:
        return None, None, f"Failed to read IODB: {err}"

    # Read raw bytes so the generator can build a writable output workbook
    try:
        template_file.seek(0)
        tmpl_bytes = template_file.read()
        template_file.seek(0)
    except Exception:
        tmpl_bytes = None

    # Load read-only for safe structure reading (avoids corrupt-style loops)
    wb, _, err = load_workbook_from_upload(
        io.BytesIO(tmpl_bytes) if tmpl_bytes else template_file,
        read_only=True,
    )
    if err:
        return None, None, f"Failed to load template: {err}"

    return generate_cable_schedule(
        df, wb, jb_column, tag_column, progress_callback,
        template_bytes=tmpl_bytes,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Loop Wiring
# ─────────────────────────────────────────────────────────────────────────────

def process_loop_wiring(
    loop_input_file,
    template_file,
    progress_callback=None,
) -> tuple[bytes | None, str | None, str | None]:
    """
    Full pipeline: read Loop Wiring input + template → generate loop wiring sheets.

    Returns (bytes, "Loop_Wiring.xlsx", error)
    """
    df, err = read_loop_wiring_input(loop_input_file)
    if err:
        return None, None, f"Failed to read Loop Wiring Input: {err}"

    wb, tmpl_buf, err = load_workbook_from_upload(template_file)
    if err:
        return None, None, f"Failed to load Loop Wiring template: {err}"

    tmpl_buf.seek(0)
    tmpl_bytes = tmpl_buf.read()
    return generate_loop_wiring(df, wb, progress_callback, template_bytes=tmpl_bytes)


# ─────────────────────────────────────────────────────────────────────────────
# Helper: get IODB columns for UI dropdowns
# ─────────────────────────────────────────────────────────────────────────────

def get_iodb_columns(iodb_file) -> tuple[list[str] | None, str | None]:
    """
    Load IODB and return the list of column names for UI multiselect.

    Returns (columns_list, None) or (None, error_string)
    """
    df, err = read_iodb(iodb_file)
    if err:
        return None, err
    return list(df.columns), None


def get_iodb_tags(iodb_file, tag_column: str = "TAG NO") -> tuple[list[str] | None, str | None]:
    """
    Load IODB and return a list of unique, non-null tag numbers for UI selection.

    Returns (tags_list, None) or (None, error_string)
    """
    df, err = read_iodb(iodb_file)
    if err:
        return None, err

    # If requested tag column missing, attempt to auto-detect a Tag-like
    # column by name or by probing header rows in the source file.
    if tag_column not in df.columns:
        # try case-insensitive name match
        cols_lower = {str(c).strip().lower(): c for c in df.columns}
        found = None
        for k in (tag_column.lower(), 'tag no', 'tag', 'tag number', 'tag_number', 'tag_nummber'):
            if k in cols_lower:
                found = cols_lower[k]
                break
        if found is None:
            # probe header rows across sheets in the original file object
            try:
                # ensure we have a bytes buffer
                iodb_file.seek(0)
                xls = pd.ExcelFile(iodb_file)
                for sheet in xls.sheet_names:
                    for hr in range(0, 11):
                        try:
                            iodb_file.seek(0)
                            cand = pd.read_excel(iodb_file, sheet_name=sheet, header=hr, nrows=10)
                            iodb_file.seek(0)
                        except Exception:
                            iodb_file.seek(0)
                            continue
                        cand_cols = [str(c).strip().lower() for c in cand.columns]
                        for k in ('tag no', 'tag', 'tag number', 'tag_number', 'tag_nummber'):
                            if k in cand_cols:
                                found = cand.columns[cand_cols.index(k)]
                                # re-read full df with detected header
                                iodb_file.seek(0)
                                df = pd.read_excel(iodb_file, sheet_name=sheet, header=hr)
                                break
                        if found is not None:
                            break
                    if found is not None:
                        break
            except Exception:
                pass

        if found is None:
            return None, f"Column '{tag_column}' not found in IODB. Available: {list(df.columns)}"
        tag_column = found

    tags = df[tag_column].dropna().astype(str).str.strip()
    tags = sorted(tags[tags != ""].unique().tolist())
    return tags, None


def get_iodb_dataframe(iodb_file) -> tuple[pd.DataFrame | None, str | None]:
    """Return the full IODB dataframe for preview purposes."""
    return read_iodb(iodb_file)


# ─────────────────────────────────────────────────────────────────────────────
# IODB Validation
# ─────────────────────────────────────────────────────────────────────────────

def process_iodb_validation(
    raw_bytes: bytes,
    auto_correct_spelling: bool = False,
    dynamic_rules=None,          # list[DynamicRule] | None
) -> tuple[bytes | None, bytes | None, bytes | None, list[dict], str | None]:
    """
    Drive the full IODB validation pipeline.

    Returns
    -------
    error_log_bytes   : Excel validation report (two sheets)
    highlighted_bytes : Original workbook with error cells colour-coded
    tba_bytes         : Excel listing every cell whose value is "TBA"
    errors            : List of error dicts for Streamlit display
    error_message     : None on success; human-readable string on failure
    """
    try:
        buf = io.BytesIO(raw_bytes)
        # keep_default_na=False so literal "NA", "N/A", "TBA", "-" etc. are
        # kept as strings and don't get silently treated as empty cells.
        df  = pd.read_excel(buf, sheet_name=0, header=0,
                            keep_default_na=False, na_values=[''])
        df.columns = [str(c).strip() for c in df.columns]
        # Preserve original index so Excel row numbers are correct after dropna
        df = df.dropna(how="all")

        buf2 = io.BytesIO(raw_bytes)
        wb, _, err = load_workbook_from_upload(buf2)
        if err:
            return None, None, [], err

        return generate_iodb_validation(df, wb, auto_correct_spelling, dynamic_rules=dynamic_rules)
    except Exception as exc:
        import traceback
        return None, None, None, [], f"Failed to process file: {exc}\n{traceback.format_exc()}"
